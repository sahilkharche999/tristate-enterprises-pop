"""
Generate Budget from Income Statement

This script READS values from the Income Statement sheet and CALCULATES
the budget using the rules defined in the VBA macros (Modules 1-8).

NO VALUES ARE COPIED - everything is calculated from source data.
"""

import argparse
import builtins
from openpyxl import load_workbook, Workbook
from datetime import datetime, timezone
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter
import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from .services.income_statement_parser import _match_section_header


QUIET_MODE = True


def _print_quiet(*args, **kwargs):
    if not QUIET_MODE:
        builtins.print(*args, **kwargs)


print = _print_quiet

# Default paths/config
DEFAULT_INPUT_PATH = "data/2024_Income_Statement.xlsx"
DEFAULT_OUTPUT_PATH = "output/Budget_Pipeline.xlsx"
DEFAULT_TEMPLATE_PATH = "data/Budget_Template_Reusable.xlsx"
# No hardcoded growth-factor fallback. It must be inferred from input or passed explicitly.
DEFAULT_GROWTH_FACTOR = None

_MONTH_TOKENS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
_DATE_TOKEN_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")

def _extract_month_from_text(text: str) -> Optional[int]:
    if not text:
        return None
    date_matches = re.findall(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", text)
    if date_matches:
        month = int(date_matches[-1][0])
        if 1 <= month <= 12:
            return month
    month_matches = re.findall(
        r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b",
        text,
        flags=re.IGNORECASE,
    )
    if month_matches:
        return _MONTH_TOKENS.get(month_matches[-1].lower())
    return None


def _extract_year_from_text(text: str) -> Optional[int]:
    if not text:
        return None
    date_matches = re.findall(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", text)
    if date_matches:
        year = int(date_matches[-1][2])
        if year < 100:
            year += 2000
        if 1900 <= year <= 2200:
            return year
    year_matches = re.findall(r"\b(19\d{2}|20\d{2}|21\d{2})\b", text)
    if year_matches:
        return int(year_matches[-1])
    return None


def _normalize_year(year: int) -> int:
    return year + 2000 if year < 100 else year


def _months_covered_from_date_tokens(text: str) -> Optional[int]:
    """Infer inclusive month coverage from first/last date tokens in a string."""
    if not text:
        return None

    matches = _DATE_TOKEN_RE.findall(text)
    if len(matches) < 2:
        return None

    start_month, _, start_year = matches[0]
    end_month, _, end_year = matches[-1]

    start_m = int(start_month)
    end_m = int(end_month)
    start_y = _normalize_year(int(start_year))
    end_y = _normalize_year(int(end_year))

    if not (1 <= start_m <= 12 and 1 <= end_m <= 12):
        return None

    months = (end_y - start_y) * 12 + (end_m - start_m) + 1
    if months <= 0 or months > 24:
        return None
    return months


def _elapsed_months_from_fiscal_start(report_month: int, fiscal_year_start_month: int) -> int:
    return ((report_month - fiscal_year_start_month) % 12) + 1


def _read_header_cells(input_path: str, sheet_name: str = "Income Statement") -> list:
    """Read first 7 rows x first 9 columns for growth factor / year inference.

    Supports .xlsx (openpyxl), .xls (xlrd), and .pdf (pdfplumber text).
    Returns a list of lists where each inner list is a row of cell values.
    """
    ext = Path(input_path).suffix.lower()

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(input_path)
        try:
            ws = wb.sheet_by_name(sheet_name)
        except xlrd.XLRDError:
            ws = wb.sheet_by_index(0)
        cells = []
        for r in range(min(7, ws.nrows)):
            row = []
            for c in range(min(9, ws.ncols)):
                v = ws.cell_value(r, c)
                row.append(None if v == "" else v)
            cells.append(row)
        return cells

    elif ext == ".pdf":
        import pdfplumber
        with pdfplumber.open(input_path) as pdf:
            if not pdf.pages:
                return []
            page = pdf.pages[0]
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if len(words) < 5:
                return []
            lines_by_y: dict = {}
            for w in words:
                y_key = round(w['top'] / 2) * 2
                lines_by_y.setdefault(y_key, []).append(w)
            rows = []
            for y_key in sorted(lines_by_y.keys())[:7]:
                line_words = sorted(lines_by_y[y_key], key=lambda w: w['x0'])
                line_text = ' '.join(w['text'] for w in line_words)
                rows.append([line_text])
            return rows

    else:
        # .xlsx or .xlsm — use openpyxl
        wb = load_workbook(input_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            cells = []
            for r in range(1, 8):
                row = []
                for c in range(1, 10):
                    row.append(ws.cell(row=r, column=c).value)
                cells.append(row)
            return cells
        finally:
            wb.close()


def infer_growth_factor_from_input(
    input_path: str,
    default_factor: Optional[float] = DEFAULT_GROWTH_FACTOR,
    sheet_name: str = "Income Statement",
    fiscal_year_start_month: int = 1,
) -> Tuple[float, int, str]:
    if not (1 <= fiscal_year_start_month <= 12):
        raise ValueError("fiscal_year_start_month must be between 1 and 12.")

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    cells = _read_header_cells(input_path, sheet_name)
    for row_idx, row in enumerate(cells):
        for col_idx, value in enumerate(row):
            if not isinstance(value, str):
                continue
            text = value.strip()
            covered_months = _months_covered_from_date_tokens(text)
            if covered_months and covered_months > 1:
                factor = 12.0 / covered_months
                source = f"row{row_idx + 1}!col{col_idx + 1}"
                return factor, covered_months, source

            month = _extract_month_from_text(text)
            if month:
                elapsed_months = _elapsed_months_from_fiscal_start(
                    month, fiscal_year_start_month
                )
                factor = 12.0 / elapsed_months
                source = f"row{row_idx + 1}!col{col_idx + 1}"
                return factor, elapsed_months, source

    month_from_filename = _extract_month_from_text(os.path.basename(input_path))
    if month_from_filename:
        elapsed_months = _elapsed_months_from_fiscal_start(
            month_from_filename, fiscal_year_start_month
        )
        return 12.0 / elapsed_months, elapsed_months, "filename"

    if default_factor is not None:
        fallback_months = int(round(12.0 / default_factor))
        return default_factor, fallback_months, "default"

    raise ValueError(
        "Could not infer month coverage from workbook header or filename. "
        "Provide --growth-factor explicitly (example: --growth-factor 1.5)."
    )


def infer_statement_year_from_input(
    input_path: str,
    sheet_name: str = "Income Statement",
) -> Tuple[int, str]:
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    cells = _read_header_cells(input_path, sheet_name)
    for row_idx, row in enumerate(cells):
        for col_idx, value in enumerate(row):
            if not isinstance(value, str):
                continue
            year = _extract_year_from_text(value.strip())
            if year:
                source = f"row{row_idx + 1}!col{col_idx + 1}"
                return year, source

    filename_year = _extract_year_from_text(os.path.basename(input_path))
    if filename_year:
        return filename_year, "filename"

    raise ValueError(f"Could not infer statement year from: {input_path}")


def _debug_enabled_for(tag: str) -> bool:
    """Return True when debug breakpoints are enabled for the given tag."""
    enabled = os.getenv("BUDGET_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return False

    tags_raw = os.getenv("BUDGET_DEBUG_TAGS", "ALL").strip()
    if not tags_raw or tags_raw.upper() == "ALL":
        return True

    selected_tags = {t.strip() for t in tags_raw.split(",") if t.strip()}
    return tag in selected_tags


def _debug_breakpoint(tag: str, context: Optional[dict] = None):
    """
    Conditional breakpoint helper.

    Use with:
      BUDGET_DEBUG=1 BUDGET_DEBUG_TAGS=ALL
      BUDGET_DEBUG=1 BUDGET_DEBUG_TAGS=GEN_READ_UTILITY_REFUND
    """
    if not _debug_enabled_for(tag):
        return

    print(f"[DEBUG:{tag}] breakpoint hit")
    if context:
        for key, value in context.items():
            print(f"  {key}={value}")
    breakpoint()


@dataclass
class LineItem:
    """Represents a single line item from the Income Statement"""
    row_num: int
    label: str
    annual_budget: float
    row_type: str  # 'item', 'title', 'total', 'reserve_item', 'reserve_title', 'empty'
    section: str
    ytd_actual: float = 0.0
    account_code: str = ''


class BudgetGenerator:
    """Calculate budget rows from an Income Statement workbook."""

    MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    COL_LABEL = 'B'           # Account name/label
    COL_SECTION = 'A'         # Section headers
    COL_ANNUAL_BUDGET = 'AG'  # Annual Budget (original) values
    COL_PROPOSED = 'AN'       # Proposed values (after % Change adjustment)
    # Legacy constant — kept for reference but no longer used.
    # Section detection now delegates to income_statement_parser._match_section_header()
    RESERVE_SECTION_START_TITLES = {
        "reserve income",
        "reserve expense",
        "reserve expenses (per reserve study)",
    }

    COL_YTD_ACTUAL = 'T'      # YTD Actual values

    def __init__(self, input_path: str, output_path: str,
                 use_proposed: bool = True, growth_factor: float = None,
                 template_path: str = None, label_aliases: dict = None,
                 sheet_name: str = "Income Statement",
                 hoa_name: str = '',
                 col_overrides: dict = None):
        """
        Initialize the generator.

        Args:
            input_path: Path to income statement Excel file
            output_path: Path for output budget file
            use_proposed: If True, read from Proposed column (AN) which has manual adjustments.
                          If False, read from Annual Budget column (AG) and apply rules.
            growth_factor: Growth factor for projections (e.g. 12/8=1.5 for 8 months)
            template_path: Path to empty template file (structure only). If provided,
                           output is written into a copy of this template instead of from scratch.
            label_aliases: Dict mapping source labels to template labels for items
                           renamed between files (e.g. account code changes).
            hoa_name: HOA property name for the budget header.
            col_overrides: Dict with 1-based column positions from enricher (COL_T, COL_AG, COL_AN).
        """
        self.input_path = input_path
        self.output_path = output_path
        self.use_proposed = use_proposed
        if growth_factor is None:
            raise ValueError(
                "Growth factor is required. Pass --growth-factor or use auto-detection before initialization."
            )
        self.GROWTH_FACTOR = growth_factor
        self.template_path = template_path
        self.label_aliases = label_aliases or {}
        self.sheet_name = sheet_name
        self.hoa_name = hoa_name
        self.line_items: List[LineItem] = []
        self._net_income = 0  # Will be calculated
        self.DATA_START_ROW = 6  # default, overridden by col_overrides
        # Apply enricher column overrides if provided
        if col_overrides:
            self.COL_YTD_ACTUAL = get_column_letter(col_overrides["COL_T"])
            self.COL_ANNUAL_BUDGET = get_column_letter(col_overrides["COL_AG"])
            self.COL_PROPOSED = get_column_letter(col_overrides["COL_AN"])
            if "DATA_START_ROW" in col_overrides:
                self.DATA_START_ROW = col_overrides["DATA_START_ROW"]

    @staticmethod
    def _is_reserve_section_start(label: str) -> bool:
        result = _match_section_header(label)
        return result is not None and result.startswith("reserve")

    @staticmethod
    def _is_reserve_label(label: str) -> bool:
        # Section-based: delegate to section header matcher instead of keyword search
        result = _match_section_header(label)
        return result is not None and result.startswith("reserve")

    def _infer_budget_year(self) -> Optional[int]:
        """Infer budget year as statement year + 1 from input workbook metadata."""
        try:
            statement_year, _ = infer_statement_year_from_input(self.input_path)
        except Exception:
            return None
        return statement_year + 1

    def _source_column_info(self) -> Tuple[str, str]:
        source_col = self.COL_PROPOSED if self.use_proposed else self.COL_ANNUAL_BUDGET
        description = (
            "Proposed (AN) - with manual adjustments"
            if self.use_proposed
            else "Annual Budget (AG) - original values"
        )
        return source_col, description

    @staticmethod
    def _parse_budget_value(raw_value) -> float:
        if raw_value is None:
            return 0.0
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        if isinstance(raw_value, str):
            clean = raw_value.replace(',', '').replace('$', '').replace('-', '0')
            try:
                return float(clean) if clean else 0.0
            except ValueError:
                return 0.0
        return 0.0

    def _set_month_values(self, row_data: Dict, monthly_value: Optional[float]):
        for month in self.MONTHS:
            row_data[month] = monthly_value

    def _month_dict(self, monthly_value: float) -> Dict[str, float]:
        return {month: monthly_value for month in self.MONTHS}

    @staticmethod
    def _ensure_min_column_width(ws, col_name: str, min_width: float):
        existing = ws.column_dimensions[col_name].width
        if existing is None or existing < min_width:
            ws.column_dimensions[col_name].width = min_width

    def _base_row_data(self, item: LineItem) -> Dict:
        return {
            'label': item.label,
            'type': item.row_type,
            'section': item.section,
            'source_row': item.row_num,
            'annual_budget_read': item.annual_budget,
            'ytd_actual': item.ytd_actual,
            'account_code': item.account_code,
        }

    def _build_item_row(self, item: LineItem) -> Dict:
        annual_budget = item.annual_budget
        projection = annual_budget * self.GROWTH_FACTOR
        percent_diff = (projection - annual_budget) / annual_budget if annual_budget else 0.0
        percent_change = 0.0
        proposed = annual_budget
        monthly = proposed / 12

        row_data = self._base_row_data(item)
        row_data['projection'] = projection
        row_data['percent_diff'] = percent_diff
        row_data['percent_change'] = percent_change
        row_data['proposed'] = proposed
        row_data['monthly'] = monthly
        row_data['budget_total'] = proposed
        self._set_month_values(row_data, monthly)

        if item.label == '41170 - Utility Refund (Conservice)':
            _debug_breakpoint("GEN_CALC_UTILITY_REFUND", {
                "source_row": item.row_num,
                "annual_budget_read": annual_budget,
                "growth_factor": self.GROWTH_FACTOR,
                "projection": projection,
                "percent_diff": percent_diff,
                "percent_change": percent_change,
                "proposed": proposed,
                "monthly": monthly,
            })

        return row_data

    def _build_total_row(self, item: LineItem) -> Dict:
        row_data = self._base_row_data(item)
        row_data['budget_total'] = item.annual_budget
        monthly = item.annual_budget / 12 if item.annual_budget else 0.0
        self._set_month_values(row_data, monthly)
        return row_data

    def _build_non_numeric_row(self, item: LineItem) -> Dict:
        row_data = self._base_row_data(item)
        row_data['budget_total'] = None
        self._set_month_values(row_data, None)
        return row_data

    @staticmethod
    def _is_total_row(item: LineItem) -> bool:
        return item.row_type in ('total', 'reserve_total')

    def _print_sample_items(self, limit: int = 10):
        print("  Sample of data read (first 10 items with values):")
        print("  " + "-" * 65)
        count = 0
        for item in self.line_items:
            if item.annual_budget != 0 and count < limit:
                print(f"    Row {item.row_num}: {item.label[:40]:<40} ${item.annual_budget:>12,.2f}")
                count += 1
        print()

    def _print_sample_calculations(self, budget_rows: List[Dict], limit: int = 5):
        print("  Sample calculations (first 5 items):")
        print("  " + "-" * 80)
        print(f"  {'Label':<35} {'Annual Budget':>15} {'Proposed':>15} {'Monthly':>12}")
        print("  " + "-" * 80)
        count = 0
        for row in budget_rows:
            if row['type'] == 'item' and row.get('proposed') and count < limit:
                print(
                    f"    {row['label'][:33]:<33} "
                    f"${row['annual_budget_read']:>12,.2f} "
                    f"${row['proposed']:>12,.2f} "
                    f"${row['monthly']:>10,.2f}"
                )
                count += 1
        print()

    def read_income_statement(self) -> List[LineItem]:
        """Read line items from the Income Statement sheet."""
        print("=" * 70)
        print("STEP 1: READING FROM INCOME STATEMENT")
        print("=" * 70)
        print(f"Source file: {self.input_path}")

        source_col, col_description = self._source_column_info()

        print(f"Reading column {source_col} ({col_description})")
        print()
        _debug_breakpoint("GEN_SOURCE_COLUMN", {
            "input_path": self.input_path,
            "use_proposed": self.use_proposed,
            "source_column": source_col,
        })

        wb = load_workbook(self.input_path, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active

        self.line_items = []
        current_section = None
        in_reserve_block = False
        items_read = 0
        values_found = 0

        for row in range(self.DATA_START_ROW, ws.max_row + 1):
            col_a = ws[f'{self.COL_SECTION}{row}'].value
            col_b = ws[f'{self.COL_LABEL}{row}'].value
            annual_budget_raw = ws[f'{source_col}{row}'].value
            ytd_actual_raw = ws[f'{self.COL_YTD_ACTUAL}{row}'].value

            label = ''
            if col_b:
                label = str(col_b).strip()
            elif col_a:
                label = str(col_a).strip()

            if not label:
                continue

            annual_budget = self._parse_budget_value(annual_budget_raw)
            ytd_actual = self._parse_budget_value(ytd_actual_raw)

            # Extract account code from label (e.g. "40000 - Assessment Income" → "40000")
            account_code = ''
            account_match = re.match(r'^(\d{4,6})\s*[-–—]', label)
            if account_match:
                account_code = account_match.group(1)

            row_type = 'item'
            if 'Total' in label:
                row_type = 'total'
            elif col_a and not col_b:
                row_type = 'title'
                current_section = label
                if self._is_reserve_section_start(label):
                    in_reserve_block = True

            # Tag reserve rows with dedicated types so they flow through the pipeline.
            if in_reserve_block or self._is_reserve_label(label):
                if row_type == 'title':
                    row_type = 'reserve_title'
                elif row_type == 'total':
                    row_type = 'reserve_total'
                else:
                    row_type = 'reserve_item'

            if annual_budget != 0:
                values_found += 1

            if label == '41170 - Utility Refund (Conservice)':
                _debug_breakpoint("GEN_READ_UTILITY_REFUND", {
                    "row": row,
                    "source_column": source_col,
                    "raw_value": annual_budget_raw,
                    "parsed_value": annual_budget,
                })

            item = LineItem(
                row_num=row,
                label=label,
                annual_budget=annual_budget,
                row_type=row_type,
                section=current_section or '',
                ytd_actual=ytd_actual,
                account_code=account_code,
            )
            self.line_items.append(item)
            items_read += 1

        wb.close()

        print(f"  Read {items_read} line items")
        print(f"  Found {values_found} non-zero Annual Budget values")
        print()

        self._print_sample_items()

        return self.line_items

    def calculate_budget(self) -> List[Dict]:
        """Calculate budget rows from parsed line items."""
        print("=" * 70)
        print("STEP 2: CALCULATING BUDGET VALUES")
        print("=" * 70)
        print(f"  Growth Factor: {self.GROWTH_FACTOR:.6f} ({(self.GROWTH_FACTOR-1)*100:.2f}% increase)")
        print("  % Change: 0.0%")
        print()
        print("  Applying formulas:")
        print("    Projection = Annual_Budget * Growth_Factor")
        print("    % Diff = (Projection - Annual_Budget) / Annual_Budget")
        print("    Proposed = Annual_Budget")
        print("    Monthly = Proposed / 12")
        print()

        budget_rows = []

        for item in self.line_items:
            if item.row_type in ('item', 'reserve_item') and item.annual_budget != 0:
                row_data = self._build_item_row(item)
            elif self._is_total_row(item):
                row_data = self._build_total_row(item)
            else:
                row_data = self._build_non_numeric_row(item)

            budget_rows.append(row_data)

        self._recalculate_totals(budget_rows)
        self._print_sample_calculations(budget_rows)

        return budget_rows

    def _sum_section_items(self, budget_rows: List[Dict], total_index: int, section_name: str) -> float:
        total = 0.0
        for idx in range(total_index - 1, -1, -1):
            prev_row = budget_rows[idx]
            if prev_row['type'] in ('title', 'reserve_title') and prev_row['label'] == section_name:
                break
            if prev_row['type'] in ('item', 'reserve_item') and prev_row.get('proposed'):
                total += prev_row['proposed']
        return total

    def _set_total_row(self, row: Dict, total: float):
        row['budget_total'] = total
        self._set_month_values(row, total / 12)

    @staticmethod
    def _find_first_total_value(budget_rows: List[Dict], label: str) -> float:
        for row in budget_rows:
            if row['label'] == label and row.get('budget_total'):
                return row['budget_total']
        return 0.0

    def _recalculate_totals(self, budget_rows: List[Dict]):
        """Recalculate total rows from item values."""
        print("  Recalculating totals by summing line items...")

        total_mappings = {
            'Total Income': 'Income',
            'Total Administration Expenses': 'Administration Expenses',
            'Total Utilities': 'Utilities',
            'Total General Maintenance': 'General Maintenance',
            'Total Landscape Maintenance': 'Landscape Maintenance',
            'Total Pool/Spa Maintenance': 'Pool/Spa Maintenance',
            'Total Allocation to Reserves': 'Allocation to Reserves',
            'Total Reserve Expense': 'Reserve Expense',
            'Total Reserve Expenses (Per Reserve Study)': 'Reserve Expenses (Per Reserve Study)',
            'Total Reserve Income': 'Reserve Income',
        }

        for i, row in enumerate(budget_rows):
            if row['type'] in ('total', 'reserve_total') and row['label'] in total_mappings:
                section_name = total_mappings[row['label']]
                total = self._sum_section_items(budget_rows, i, section_name)

                if total > 0:
                    self._set_total_row(row, total)
                    print(f"    {row['label']}: Calculated ${total:,.2f} (from summing items)")
                    if row['label'] == 'Total Income':
                        _debug_breakpoint("GEN_TOTAL_INCOME_ROLLUP", {
                            "section_name": section_name,
                            "total_income": total,
                            "monthly_income": total / 12,
                        })

        total_expense = 0
        expense_totals = ['Total Administration Expenses', 'Total Utilities',
                         'Total General Maintenance', 'Total Landscape Maintenance',
                         'Total Pool/Spa Maintenance']

        for row in budget_rows:
            if row['label'] in expense_totals and row.get('budget_total'):
                total_expense += row['budget_total']

        print(f"    Total Operating Expense: Calculated ${total_expense:,.2f}")

        found_total_expense = False
        for row in budget_rows:
            if row['label'] == 'Total Operating Expense':
                self._set_total_row(row, total_expense)
                found_total_expense = True

        if not found_total_expense:
            budget_rows.append({
                'label': 'Total Operating Expense',
                'type': 'total',
                **{month: total_expense / 12 for month in self.MONTHS},
                'budget_total': total_expense,
            })

        total_income = self._find_first_total_value(budget_rows, 'Total Income')
        net_income = total_income - total_expense
        print(f"    Net Operating Income: ${total_income:,.2f} - ${total_expense:,.2f} = ${net_income:,.2f}")
        _debug_breakpoint("GEN_NET_OPERATING_INCOME", {
            "total_income": total_income,
            "total_operating_expense": total_expense,
            "net_operating_income": net_income,
        })

        found_net_income = False
        for row in budget_rows:
            if row['label'] == 'Net Operating Income' or row['label'] == 'Operating Net Total':
                row['label'] = 'Net Operating Income'
                self._set_total_row(row, net_income)
                found_net_income = True

        if not found_net_income:
            budget_rows.append({
                'label': 'Net Operating Income',
                'type': 'total',
                **{month: net_income / 12 for month in self.MONTHS},
                'budget_total': net_income,
            })

        self._net_income = net_income

        print()

    def calculate_cash_flow(self, budget_rows: List[Dict], reserve_contribution: float):
        """Append reserve, monthly cash flow, and cumulative cash flow rows."""
        print("  Calculating cash flow (Module 8 rules)...")

        net_income = getattr(self, '_net_income', 0)
        if net_income == 0:
            for row in budget_rows:
                if row['label'] == 'Net Operating Income':
                    net_income = row.get('budget_total', 0)
                    break

        # Reserve lines are excluded from this system by policy.
        excluded_reserve_contribution = reserve_contribution
        reserve_contribution = 0.0

        monthly_net_income = net_income / 12
        monthly_reserve = reserve_contribution / 12
        monthly_cash_flow = monthly_net_income - monthly_reserve
        _debug_breakpoint("GEN_CASH_FLOW", {
            "net_operating_income": net_income,
            "monthly_net_income": monthly_net_income,
            "reserve_contribution_annual": reserve_contribution,
            "excluded_reserve_contribution_annual": excluded_reserve_contribution,
            "monthly_reserve": monthly_reserve,
            "monthly_cash_flow": monthly_cash_flow,
            "annual_cash_flow": monthly_cash_flow * 12,
        })
        if excluded_reserve_contribution:
            print(
                "    Reserve contribution excluded from output rows by policy: "
                f"${excluded_reserve_contribution:,.2f}/year"
            )

        budget_rows.append({
            'label': 'Monthly Cash Flow',
            'annual_label': 'Annual Cash Flow',
            'type': 'cashflow',
            **self._month_dict(monthly_cash_flow),
            'budget_total': monthly_cash_flow * 12,
        })

        print(f"    Monthly Cash Flow: ${monthly_net_income:,.2f} - ${monthly_reserve:,.2f} = ${monthly_cash_flow:,.2f}")

        cumulative_row = {
            'label': 'Cumulative Cash Flow',
            'annual_label': 'Cumulative Cash Flow',
            'type': 'cumulative',
        }

        cumulative = 0
        for month in self.MONTHS:
            cumulative += monthly_cash_flow
            cumulative_row[month] = cumulative

        cumulative_row['budget_total'] = cumulative
        budget_rows.append(cumulative_row)

        print(f"    Cumulative Cash Flow (Dec): ${cumulative:,.2f}")
        print()

    def calculate_percentages(self, budget_rows: List[Dict]):
        """Calculate each item's share of its section total (section_share)."""
        print("  Calculating percentages...")

        section_totals = {}
        for row in budget_rows:
            if row['type'] in ('total', 'reserve_total') and row.get('budget_total'):
                section_totals[row['label']] = row['budget_total']

        current_section_total = None

        for row in budget_rows:
            if row['type'] in ('title', 'reserve_title'):
                total_name = f"Total {row['label']}"
                current_section_total = section_totals.get(total_name)

            if row['type'] in ('item', 'reserve_item') and current_section_total and row.get('budget_total'):
                row['section_share'] = row['budget_total'] / current_section_total

        print()

    def write_budget_excel(self, budget_rows: List[Dict]):
        """Write budget rows to a client-ready workbook in the reference format."""
        print("=" * 70)
        print("STEP 3: WRITING OUTPUT FILE (REFERENCE FORMAT)")
        print("=" * 70)
        print(f"Output: {self.output_path}")
        print()

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"

        # --- Styles ---
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        bold = Font(bold=True, size=11)
        normal = Font(size=11)
        currency_fmt = '$#,##0'
        pct_fmt = '0.00%'
        thin_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4'),
        )
        total_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='medium', color='999999'),
            bottom=Side(style='thin', color='D4D4D4'),
        )

        # --- Row 1: Title ---
        budget_year = self._infer_budget_year()
        hoa_label = self.hoa_name or "HOA"
        title = f"{hoa_label} — Proposed Budget"
        if budget_year:
            title += f" {budget_year}"
        ws['A1'] = title
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=14)

        # --- Row 2: Export date ---
        ws['A2'] = f"Exported: {datetime.now(timezone.utc).strftime('%B %d, %Y')}"
        ws['A2'].font = Font(size=10, color="737373")

        # --- Row 4: Column headers ---
        headers = ['Account Code', 'Line Item', 'YTD Actual', 'Annual Budget',
                   '% Change', 'Proposed Budget', 'Monthly', 'Notes']
        for i, h in enumerate(headers):
            cell = ws.cell(row=4, column=i + 1, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center' if i >= 2 else 'left')
            cell.border = thin_border

        # --- Data rows ---
        current_row = 5
        seen_labels = set()

        for row_data in budget_rows:
            label = row_data['label']
            row_type = row_data['type']

            # Skip duplicate total rows
            if row_type in ('total', 'reserve_total', 'cashflow', 'cumulative'):
                if label in seen_labels:
                    continue
                seen_labels.add(label)

            # Skip the redundant "Total Operating Expense" if we already have "Total Expense"
            if label == 'Total Operating Expense' and 'Total Expense' in seen_labels:
                continue
            if label == 'Total Expense' and 'Total Operating Expense' in seen_labels:
                continue

            # Section headers (title rows)
            if row_type in ('title', 'reserve_title'):
                section_label = label.upper()
                ws.cell(row=current_row, column=1, value='')
                cell = ws.cell(row=current_row, column=2, value=section_label)
                cell.font = bold
                cell.fill = section_fill
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).fill = section_fill
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1
                continue

            # Total / summary rows
            if row_type in ('total', 'reserve_total', 'cashflow', 'cumulative'):
                ws.cell(row=current_row, column=1, value='')
                cell_b = ws.cell(row=current_row, column=2, value=label)
                cell_b.font = bold

                budget_val = row_data.get('budget_total')
                proposed = row_data.get('proposed', budget_val)
                monthly = row_data.get('monthly', (proposed or 0) / 12 if proposed else None)
                ytd = row_data.get('ytd_actual', 0)
                annual = row_data.get('annual_budget_read', budget_val)

                if ytd:
                    cell_c = ws.cell(row=current_row, column=3, value=round(ytd))
                    cell_c.number_format = currency_fmt
                    cell_c.font = bold
                if annual:
                    cell_d = ws.cell(row=current_row, column=4, value=round(annual))
                    cell_d.number_format = currency_fmt
                    cell_d.font = bold
                if proposed is not None:
                    cell_f = ws.cell(row=current_row, column=6, value=round(proposed))
                    cell_f.number_format = currency_fmt
                    cell_f.font = bold
                if monthly is not None:
                    cell_g = ws.cell(row=current_row, column=7, value=round(monthly))
                    cell_g.number_format = currency_fmt
                    cell_g.font = bold

                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = total_border

                current_row += 1
                continue

            # Regular item rows
            account_code = row_data.get('account_code', '')
            ws.cell(row=current_row, column=1, value=account_code).font = normal

            cell_b = ws.cell(row=current_row, column=2, value=label)
            cell_b.font = normal

            ytd = row_data.get('ytd_actual', 0)
            if ytd:
                cell_c = ws.cell(row=current_row, column=3, value=round(ytd))
                cell_c.number_format = currency_fmt

            annual = row_data.get('annual_budget_read', 0)
            if annual:
                cell_d = ws.cell(row=current_row, column=4, value=round(annual))
                cell_d.number_format = currency_fmt

            pct_change = row_data.get('percent_change', 0)
            cell_e = ws.cell(row=current_row, column=5, value=(pct_change or 0) / 100.0)
            cell_e.number_format = pct_fmt

            proposed = row_data.get('proposed', row_data.get('budget_total', 0))
            if proposed:
                cell_f = ws.cell(row=current_row, column=6, value=round(proposed))
                cell_f.number_format = currency_fmt

            monthly = row_data.get('monthly', 0)
            if monthly:
                cell_g = ws.cell(row=current_row, column=7, value=round(monthly, 2))
                cell_g.number_format = currency_fmt

            for col in range(1, 9):
                ws.cell(row=current_row, column=col).border = thin_border

            current_row += 1

        # --- Column widths ---
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 25

        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        wb.save(self.output_path)

        print(f"  Budget saved to: {self.output_path}")
        print()

    def write_budget_from_template(self, budget_rows: List[Dict]):
        """Write calculated budget values into the provided template."""

        print("=" * 70)
        print("STEP 3: WRITING OUTPUT FILE (TEMPLATE MODE)")
        print("=" * 70)
        print(f"  Template: {self.template_path}")
        print(f"  Output:   {self.output_path}")
        print()

        wb = load_workbook(self.template_path)
        ws = wb.active

        budget_year = self._infer_budget_year()
        if budget_year:
            ws.cell(row=1, column=8, value=f"Operating and Cash Flow Statement for {budget_year}")

        label_to_row = {}
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if label and label.strip():
                label_to_row[label.strip()] = row

        # Reserve rows are now included in the output.

        filled = 0
        missing = []
        filled_rows = set()

        for row_data in budget_rows:
            label = row_data['label'].strip()
            template_label = self.label_aliases.get(label, label)
            excel_row = label_to_row.get(template_label)

            if excel_row is None:
                if row_data['type'] in ('item', 'reserve_item', 'total', 'reserve_total', 'cashflow', 'cumulative'):
                    missing.append(label)
                continue

            if excel_row in filled_rows:
                continue
            filled_rows.add(excel_row)

            # Keep section/title rows non-numeric in the output template.
            if row_data['type'] in ('title', 'reserve_title'):
                for col in range(2, 14):
                    ws.cell(row=excel_row, column=col).value = None
                ws.cell(row=excel_row, column=16).value = None
                ws.cell(row=excel_row, column=17).value = None
                filled += 1
                continue

            for i, month in enumerate(self.MONTHS):
                val = row_data.get(month)
                if val is not None:
                    ws.cell(row=excel_row, column=2 + i, value=round(val, 2))
                    ws.cell(row=excel_row, column=2 + i).number_format = '#,##0.00'

            pct_change = row_data.get('percent_change', 0)
            ws.cell(row=excel_row, column=16, value=(pct_change or 0) / 100.0)
            ws.cell(row=excel_row, column=16).number_format = '0.00%'

            budget_val = row_data.get('budget_total')
            if budget_val is not None:
                ws.cell(row=excel_row, column=17, value=round(budget_val, 2))
                ws.cell(row=excel_row, column=17).number_format = '#,##0.00'

            filled += 1

        # Keep key numeric columns readable for larger totals.
        self._ensure_min_column_width(ws, 'A', 40)
        for i in range(2, 14):
            self._ensure_min_column_width(ws, get_column_letter(i), 12)
        self._ensure_min_column_width(ws, 'O', 35)
        self._ensure_min_column_width(ws, 'P', 10)
        self._ensure_min_column_width(ws, 'Q', 16)

        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        wb.save(self.output_path)
        wb.close()

        print(f"  Filled {filled} rows in template")
        if missing:
            print(f"  WARNING: {len(missing)} labels not found in template: {missing[:5]}")
        print(f"  Budget saved to: {self.output_path}")
        print()

    def _infer_reserve_contribution(self) -> float:
        """Infer annual reserve contribution from source line items."""
        preferred_labels = (
            "90000 - Reserve - Allocation/Transfer",
            "Total Allocation to Reserves",
        )
        for label in preferred_labels:
            for item in self.line_items:
                if item.label == label and item.annual_budget:
                    return float(item.annual_budget)

        for item in self.line_items:
            text = item.label.lower()
            if "reserve" in text and "allocation" in text and item.annual_budget:
                return float(item.annual_budget)
        return 0.0

    def generate(self, reserve_contribution: Optional[float] = None):
        """Run the full read -> calculate -> write flow."""
        print()
        print("*" * 70)
        print("*  BUDGET GENERATION - CALCULATING FROM SOURCE DATA")
        print("*  (NOT copying values - all values are calculated)")
        print("*" * 70)
        print()

        self.read_income_statement()
        if reserve_contribution is None:
            reserve_contribution = self._infer_reserve_contribution()
            print(f"  Reserve contribution inferred from source: ${reserve_contribution:,.2f}")
            print()

        budget_rows = self.calculate_budget()
        self.calculate_cash_flow(budget_rows, reserve_contribution)
        self.calculate_percentages(budget_rows)
        if self.template_path:
            self.write_budget_from_template(budget_rows)
        else:
            self.write_budget_excel(budget_rows)

        return budget_rows


def verify_calculations(output_path: str, expected_values: Dict[str, float]):
    """
    Verify that calculated values match expected values.
    """
    print("=" * 70)
    print("STEP 4: VERIFICATION")
    print("=" * 70)
    print()

    wb = load_workbook(output_path, data_only=True)
    ws = wb.active

    print(f"{'Item':<40} {'Expected':>15} {'Calculated':>15} {'Status':>10}")
    print("-" * 82)

    all_match = True
    checked_labels = set()
    for row in range(1, ws.max_row + 1):
        label = ws[f'A{row}'].value
        if label in expected_values and label not in checked_labels:
            expected = expected_values[label]
            calculated = ws[f'Q{row}'].value or 0

            diff = abs(calculated - expected)
            match = diff < 0.02

            if not match:
                all_match = False
                _debug_breakpoint("GEN_VERIFY_MISMATCH", {
                    "label": label,
                    "expected": expected,
                    "calculated": calculated,
                    "diff": diff,
                    "output_path": output_path,
                })

            status = "✓ MATCH" if match else f"✗ DIFF: {diff:.2f}"
            print(f"{label:<40} ${expected:>12,.2f} ${calculated:>12,.2f} {status:>10}")
            checked_labels.add(label)

    wb.close()

    print("-" * 82)
    if all_match:
        print("✓ ALL VALUES VERIFIED - Calculations match expected values!")
    else:
        print("✗ SOME VALUES DO NOT MATCH")
    print()


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate budget workbook from Income Statement source."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT_PATH,
        help="Path to input Income Statement workbook.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_PATH,
        help="Path to output budget workbook.",
    )
    parser.add_argument(
        "--template",
        default=DEFAULT_TEMPLATE_PATH,
        help="Template workbook path. Set to empty string to write without template.",
    )
    parser.add_argument(
        "--growth-factor",
        type=float,
        default=None,
        help="Growth factor used for projection. If omitted, auto-detected from input workbook month coverage.",
    )
    parser.add_argument(
        "--fiscal-year-start-month",
        type=int,
        default=1,
        help="Fiscal year start month (1=Jan ... 12=Dec) used when auto-detecting growth factor.",
    )
    parser.add_argument(
        "--reserve-contribution",
        type=float,
        default=None,
        help="Annual reserve contribution. If omitted, inferred from source rows.",
    )
    parser.add_argument(
        "--source-column",
        choices=["AG", "AN"],
        default="AN",
        help="Read source values from AG or AN.",
    )
    return parser.parse_args()


def main():
    """Main entry point."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    args = parse_args()

    if args.growth_factor is not None and args.growth_factor <= 0:
        raise ValueError("--growth-factor must be greater than 0.")
    if not (1 <= args.fiscal_year_start_month <= 12):
        raise ValueError("--fiscal-year-start-month must be between 1 and 12.")

    if args.growth_factor is None:
        resolved_growth_factor, detected_months, source = infer_growth_factor_from_input(
            args.input,
            fiscal_year_start_month=args.fiscal_year_start_month,
        )
        print(
            f"Auto growth factor: {resolved_growth_factor:.6f} "
            f"(12/{detected_months}, source={source})"
        )
    else:
        resolved_growth_factor = args.growth_factor

    template_path = args.template.strip() or None
    generator = BudgetGenerator(
        args.input,
        args.output,
        use_proposed=(args.source_column == "AN"),
        growth_factor=resolved_growth_factor,
        template_path=template_path,
    )
    generator.generate(reserve_contribution=args.reserve_contribution)


if __name__ == "__main__":
    main()
