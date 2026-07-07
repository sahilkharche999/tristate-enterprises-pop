"""Durable HOA-scoped budget history service."""
from __future__ import annotations

import asyncio
import hashlib
import html
import json
import logging
import os
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..assessment_mode import (
    ASSESSMENT_MODE_VARIABLE,
    AssessmentMode,
    normalize_assessment_mode,
    package_impact_for_mode_drift,
)
from ..config import settings
from ..generate_budget import infer_growth_factor_from_input
from ..generate_budget_pipeline import BudgetPipeline
from ..models.financial_document_extraction import (
    DocumentExtractionFailure,
    ExtractedFinancialStatement,
    ExtractedFinancialStatementPage,
    ExtractedStatementLineItem,
)
from ..models.reserve_study_extraction import ExtractedReserveStudyDocument
from ..services.income_statement_parser import (
    parse_rows_with_sections,
    detect_columns,
    _parse_financial_float as _parse_financial_float_new,
    _match_section_header,
    _classify_by_account_code,
    SECTION_KINDS,
    READ_ONLY_SECTIONS,
)
from ..services.financial_document_router import choose_financial_document_route
from ..services.assessment_budget_mapping_rule_service import (
    materialize_budget_line_pool_mappings,
    select_assessment_mapping_amount,
)
from ..services.budget_line_merge_service import auto_apply_merges_on_upload
from ..services.budget_line_merge_service import finalize_applied_merges
from ..services.financial_statement_validation import (
    has_blocking_validation_issues,
    validate_extracted_statement,
)
from ..services.normalized_statement_workbook import build_normalized_statement_workbook
from ..services.pdf_vlm_extractor import extract_pdf_statement
from ..services.reserve_study_extractor import canonicalize_reserve_study_row_dicts, extract_reserve_study
from ..services.statement_period_inference import (
    extract_pdf_statement_period_hint,
    infer_growth_factor_from_statement_period,
)
from ..models.budget_history import (
    BundleFileStatus,
    BudgetDraftCompareBaselineOption,
    BudgetDraftCompareOptionsResponse,
    BudgetDraftCompareResponse,
    BudgetDraftReserveReviewRequest,
    BudgetDraftReserveReviewResponse,
    BudgetBundleUploadResponse,
    BudgetDraftChangeSummary,
    BudgetDraftCompareRow,
    BudgetReserveComponentRow,
    BudgetReserveReviewSummary,
    BudgetDraftSummary,
    BudgetDraftPayload,
    BudgetDraftSaveRequest,
    BudgetGenerateRequest,
    BudgetGenerateResponse,
    BudgetHistoryResponse,
    BudgetReserveStudyApplyResponse,
    BudgetReserveStudySaveRequest,
    BudgetNoteRecord,
    BudgetNoteSaveRequest,
    BudgetNoteSaveResponse,
    BudgetTimelineEvent,
    BudgetUploadResponse,
    ExtractionDebugInfo,
    ExtractionQualityWarning,
    BudgetVersionCompareCard,
    BudgetVersionCompareResponse,
    BudgetVersionDetail,
    BudgetVersionMetadataUpdateRequest,
    BudgetVersionMetadataUpdateResponse,
    BudgetVersionReopenResponse,
    BudgetVersionSummary,
)


def _raw_sqlite_connection(session: Session):
    raw_conn = session.connection().connection
    return (
        getattr(raw_conn, "driver_connection", None)
        or getattr(raw_conn, "connection", None)
        or raw_conn
    )
from ..services import app_settings_service, macros_service
from ..ai_implementation.db.models import (
    BUDGET_DRAFT_ACTIVE,
    BUDGET_DRAFT_GENERATED,
    BUDGET_DRAFT_SUPERSEDED,
    BUDGET_VERSION_STAGE_FINAL,
    BUDGET_VERSION_STAGE_INTERIM,
    BudgetAuditEvent,
    BudgetDraft,
    BudgetNote,
    BudgetUpload,
    BudgetVersion,
    Property,
)

# Enriched workbook output column indices (fixed positions added by the pipeline)
# These are NOT input indices — they reference AK/AL columns written by IncomeStatementEnricher.
_COL_PROJECTION_INDEX = 37
_COL_PERCENT_CHANGE_INDEX = 38

_EXPECTED_INCOME_STATEMENT_GUIDANCE = [
    (
        "Expected budget source format: upload an income statement / statement of revenues "
        "and expenses, not a monthly operating budget or cash-flow budget workbook."
    ),
    (
        "The workbook should have a line-item/account column plus financial comparison columns "
        "such as Current Period, Year To Date, and Annual Budget."
    ),
    (
        "A good example is 'Income Statement Esprit Park Aug 2025.xlsx': sheet 'Income Statement', "
        "section rows like Operating Income/Operating Expense, and an Annual Budget column with "
        "usable values for the line items."
    ),
]

SOURCE_MODE_INCOME_STATEMENT = "income_statement"
SOURCE_MODE_PROFORMA_FINAL_BUDGET = "proforma_final_budget"
_PROFORMA_GROWTH_FACTOR = 1.0
_PROFORMA_GROWTH_FACTOR_NOTE = "pro forma / final budget annual basis"
_PROFORMA_EXCEL_GEMINI_TIMEOUT_SECONDS = 240.0
_MONTH_HEADER_ALIASES = {
    "jan", "january",
    "feb", "february",
    "mar", "march",
    "apr", "april",
    "may",
    "jun", "june",
    "jul", "july",
    "aug", "august",
    "sep", "sept", "september",
    "oct", "october",
    "nov", "november",
    "dec", "december",
}


def _normalize_source_mode(source_mode: Optional[str]) -> str:
    normalized = str(source_mode or SOURCE_MODE_INCOME_STATEMENT).strip().lower()
    if normalized == SOURCE_MODE_PROFORMA_FINAL_BUDGET:
        return SOURCE_MODE_PROFORMA_FINAL_BUDGET
    return SOURCE_MODE_INCOME_STATEMENT


def _is_proforma_source_mode(source_mode: Optional[str]) -> bool:
    return _normalize_source_mode(source_mode) == SOURCE_MODE_PROFORMA_FINAL_BUDGET


def _assessment_mode_package_impacts(
    session: Session,
    *,
    hoa_id: int,
    live_assessment_mode: AssessmentMode,
) -> list[dict[str, Any]]:
    raw_conn = _raw_sqlite_connection(session)
    rows = raw_conn.execute(
        """
        SELECT id, fiscal_year, status, assessment_mode
          FROM annual_packages
         WHERE property_id = ?
         ORDER BY fiscal_year DESC, id DESC
        """,
        (hoa_id,),
    ).fetchall()
    if not rows:
        return []

    impacts: list[dict[str, Any]] = []
    latest_by_year: set[int] = set()
    for row in rows:
        package_id, fiscal_year, status, stored_mode = row
        is_latest_for_year = int(fiscal_year) not in latest_by_year
        latest_by_year.add(int(fiscal_year))
        impact, reason = package_impact_for_mode_drift(
            status=str(status),
            package_assessment_mode=stored_mode,
            live_assessment_mode=live_assessment_mode,
            is_latest_for_fiscal_year=is_latest_for_year,
        )
        if impact == "none":
            continue
        impacts.append(
            {
                "package_id": int(package_id),
                "fiscal_year": int(fiscal_year),
                "status": str(status),
                "impact": impact,
                "reason": reason,
            }
        )
    return impacts


def _update_property_assessment_mode(
    session: Session,
    *,
    hoa: Property,
    actor: dict[str, Any],
    requested_assessment_mode: Optional[str],
) -> AssessmentMode:
    next_mode = normalize_assessment_mode(
        requested_assessment_mode or getattr(hoa, "assessment_mode", None)
    )
    current_mode = normalize_assessment_mode(getattr(hoa, "assessment_mode", None))
    if current_mode == next_mode:
        hoa.assessment_mode = next_mode
        return next_mode

    hoa.assessment_mode = next_mode
    package_impacts = _assessment_mode_package_impacts(
        session,
        hoa_id=hoa.id,
        live_assessment_mode=next_mode,
    )
    _create_audit_event(
        session,
        hoa_id=hoa.id,
        actor=actor,
        event_type="assessment_mode_changed",
        summary=f"Assessment mode changed to {next_mode}",
        payload={
            "from_assessment_mode": current_mode,
            "to_assessment_mode": next_mode,
            "package_impacts": package_impacts,
        },
    )
    return next_mode


def _normalize_header_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _try_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "—"}:
            return None
        text = text.replace(",", "").replace("$", "")
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        try:
            return float(text)
        except ValueError:
            return None
    return None


_PROFORMA_EXCEL_SYSTEM_PROMPT = """You are extracting an HOA pro forma / final budget from an Excel workbook.

Return JSON only using the provided schema.

Extraction rules:
- Extract every visible detail budget line item from operating income, operating expense, reserve income, reserve contributions, and reserve expense sections.
- Use workbook cell positions, nearby headers, indentation, section labels, and account codes to infer each row.
- Map final/approved/proposed/annual/total budget synonyms to annual_budget. Prefer final or approved budget columns over proposed columns when both exist.
- If the workbook has Jan-Dec/monthly columns and an annual total/final budget column, use the annual total/final budget value for annual_budget, not one monthly value.
- Preserve account codes when visible.
- Set section_kind to income, operating, reserve_income, or reserve_expense.
- Skip subtotal, total, grand total, blank, comment, formula-check, and header-only rows.
- Do not collapse the workbook into one summary row. Return all detail rows.
- Use evidence.source_column to name the source column header you used, such as final_budget, approved_budget, proposed_budget, annual_budget, or total_budget.
"""


def _format_excel_cell_for_prompt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:.10g}"
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split()).strip()


def _extract_excel_workbook_prompt_text(
    path: str,
    *,
    max_rows_per_sheet: int = 250,
    max_cols_per_sheet: int = 80,
    max_chars: int = 120_000,
) -> str:
    """Serialize workbook cells for Gemini without doing semantic extraction."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    char_count = 0
    try:
        for worksheet in workbook.worksheets:
            sheet_header = f"Sheet: {worksheet.title}"
            lines.append(sheet_header)
            char_count += len(sheet_header) + 1
            max_row = min(int(worksheet.max_row or 0), max_rows_per_sheet)
            max_col = min(int(worksheet.max_column or 0), max_cols_per_sheet)
            if max_row <= 0 or max_col <= 0:
                lines.append("(empty)")
                char_count += len("(empty)") + 1
                continue

            included_rows = 0
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                cells: list[str] = []
                row_number = 0
                for cell in row:
                    row_number = int(getattr(cell, "row", row_number) or row_number)
                    text = _format_excel_cell_for_prompt(getattr(cell, "value", None))
                    if text:
                        cells.append(f"{cell.coordinate}={text}")
                if not cells:
                    continue
                included_rows += 1
                row_line = f"R{row_number:03d}: " + " | ".join(cells)
                lines.append(row_line)
                char_count += len(row_line) + 1
                if char_count >= max_chars:
                    lines.append("[Workbook text truncated for Gemini prompt]")
                    break
            if included_rows == 0:
                empty_note = "(no populated cells in scanned range)"
                lines.append(empty_note)
                char_count += len(empty_note) + 1
            if lines and lines[-1] == "[Workbook text truncated for Gemini prompt]":
                break
    finally:
        workbook.close()

    prompt_text = "\n".join(lines).strip()
    if not prompt_text:
        raise ValueError("Workbook did not contain readable cells for Gemini extraction.")
    return prompt_text


def _infer_proforma_category(
    *,
    current_section: str,
    label: str,
    account_code: Optional[int],
) -> tuple[str, Optional[str]]:
    lowered_label = _normalize_compare_text(label)
    if "reserve interest" in lowered_label:
        return "reserve_income", "income"
    if (
        "contribution to reserve" in lowered_label
        or "reserve contribution" in lowered_label
        or "transfer to reserve" in lowered_label
        or "allocation to reserve" in lowered_label
    ):
        return "reserve_expense", "transfer"
    if account_code is not None and account_code >= 90000:
        return "reserve_expense", "component"
    if current_section == "income":
        return "income", None
    return "operating", None


def _parse_proforma_excel_source(path: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        selected_sheet = None
        header_row_idx = None
        month_columns: list[int] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_idx in range(1, min(worksheet.max_row, 40) + 1):
                row_values = [worksheet.cell(row_idx, col).value for col in range(1, worksheet.max_column + 1)]
                found_months = [
                    col_idx
                    for col_idx, cell_value in enumerate(row_values, start=1)
                    if _normalize_header_text(cell_value) in _MONTH_HEADER_ALIASES
                ]
                if len(found_months) >= 6:
                    selected_sheet = worksheet
                    header_row_idx = row_idx
                    month_columns = found_months
                    break
            if selected_sheet is not None:
                break

        if selected_sheet is None or header_row_idx is None or not month_columns:
            raise ValueError("Could not find a Jan-Dec style month grid in this workbook.")

        first_month_col = min(month_columns)
        last_month_col = max(month_columns)

        account_code_col = 1 if first_month_col > 1 else None
        label_col = 2 if first_month_col > 2 else max(1, first_month_col - 1)
        if account_code_col is not None and label_col == account_code_col:
            label_col = account_code_col + 1

        candidate_columns: list[tuple[int, int, str]] = []
        for col_idx in range(last_month_col + 1, selected_sheet.max_column + 1):
            header_text = _normalize_header_text(selected_sheet.cell(header_row_idx, col_idx).value)
            if not header_text:
                continue
            if "final" in header_text or "approved" in header_text:
                priority = 4
                source_column = "final_budget"
            elif "proposed" in header_text:
                priority = 3
                source_column = "proposed_amount"
            elif "budget" in header_text:
                priority = 2
                source_column = "annual_budget"
            elif "total" in header_text:
                priority = 1
                source_column = "annual_budget"
            else:
                continue

            numeric_values = []
            for row_idx in range(header_row_idx + 1, min(selected_sheet.max_row, header_row_idx + 80) + 1):
                numeric_value = _try_float(selected_sheet.cell(row_idx, col_idx).value)
                if numeric_value is not None:
                    numeric_values.append(numeric_value)
            if len(numeric_values) < 3:
                continue
            tiny_ratio = sum(1 for value in numeric_values if abs(value) <= 1.5) / max(len(numeric_values), 1)
            if tiny_ratio > 0.7:
                continue
            candidate_columns.append((priority, col_idx, source_column))

        if not candidate_columns:
            raise ValueError("Could not find an annual final/proposed/budget column in this workbook.")

        candidate_columns.sort(key=lambda item: (-item[0], item[1]))
        best_priority = candidate_columns[0][0]
        top_candidates = [candidate for candidate in candidate_columns if candidate[0] == best_priority]
        if len(top_candidates) > 1:
            raise ValueError("Multiple annual-like columns conflict without a safe precedence rule.")

        _priority, annual_col_idx, annual_source_column = top_candidates[0]
        annual_col_letter = selected_sheet.cell(header_row_idx, annual_col_idx).column_letter

        line_items: list[dict[str, Any]] = []
        current_section = "operating"
        current_section_label = "Expense"

        for row_idx in range(header_row_idx + 1, selected_sheet.max_row + 1):
            label_raw = selected_sheet.cell(row_idx, label_col).value
            account_raw = selected_sheet.cell(row_idx, account_code_col).value if account_code_col is not None else None
            annual_value = _try_float(selected_sheet.cell(row_idx, annual_col_idx).value)
            month_values = [
                _try_float(selected_sheet.cell(row_idx, month_col).value)
                for month_col in month_columns
            ]
            non_empty_month_values = [value for value in month_values if value is not None]
            has_numeric_context = annual_value is not None or bool(non_empty_month_values)
            label_text = str(label_raw).strip() if label_raw is not None else ""
            account_text = str(account_raw).strip() if account_raw is not None else ""
            lowered_label = _normalize_compare_text(label_text)

            if not label_text and not account_text and not has_numeric_context:
                continue

            if (
                label_text
                and not has_numeric_context
                and not account_text
            ):
                if "income" in lowered_label:
                    current_section = "income"
                    current_section_label = label_text
                elif "expense" in lowered_label or "operating" in lowered_label or "maintenance" in lowered_label:
                    current_section = "operating"
                    current_section_label = label_text
                elif "reserve" in lowered_label and "income" in lowered_label:
                    current_section = "reserve_income"
                    current_section_label = label_text
                elif "reserve" in lowered_label:
                    current_section = "reserve_expense"
                    current_section_label = label_text
                continue

            if lowered_label.startswith("total ") or lowered_label.startswith("net ") or lowered_label.startswith("subtotal"):
                continue
            if "cash flow" in lowered_label or lowered_label.startswith("cumulative "):
                continue

            account_code = None
            if account_text:
                account_digits = "".join(character for character in account_text if character.isdigit())
                if account_digits:
                    try:
                        account_code = int(account_digits)
                    except ValueError:
                        account_code = None

            label = label_text or account_text
            if not label:
                continue

            category, reserve_group = _infer_proforma_category(
                current_section=current_section,
                label=label,
                account_code=account_code,
            )

            item: dict[str, Any] = {
                "line_item_key": str(account_code) if account_code is not None else label,
                "account_code": account_code,
                "label": label,
                "name": label,
                "category": category,
                "current_actual": non_empty_month_values[-1] if non_empty_month_values else None,
                "ytd_actual": sum(non_empty_month_values) if non_empty_month_values else None,
                "annual_budget": annual_value,
                "projection": annual_value,
                "percent_change": 0.0,
                "read_only": _effective_read_only({}, category),
                "reserve_group": reserve_group,
                "source_column": annual_source_column,
                "source_page_or_cell": f"{selected_sheet.title}!{annual_col_letter}{row_idx}",
                "raw": {
                    "section": current_section_label,
                    "sheet": selected_sheet.title,
                },
            }
            line_items.append(item)

        if len(line_items) < 2:
            raise ValueError("Not enough pro-forma line items were extracted from this workbook.")

        return line_items
    finally:
        workbook.close()


def _default_proforma_statement_month(hoa: Property) -> int:
    return int(hoa.fiscal_year_end_month or 12)


def _effective_read_only(item: dict[str, Any], category: str) -> bool:
    """Resolve the effective read-only flag for a line item.

    An explicit per-line override takes precedence over the section default.
    When override is None (absent) the category policy from READ_ONLY_SECTIONS
    applies unchanged.
    """
    override = item.get("read_only_override")
    if override is not None:
        return bool(override)
    return category in READ_ONLY_SECTIONS


def _derive_reserve_income(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Derive reserve_income contribution line amounts from reserve allocation transfers.

    Sums the annual_budget of all reserve_expense lines whose reserve_group is
    "transfer" (the operating-side contribution to the reserve fund) and sets
    that sum as the annual_budget on reserve_income lines that are not explicitly
    classified as interest (reserve_group != "income"). This keeps the two sides
    of the reserve transfer in sync without allowing them to diverge.

    Interest lines (reserve_group == "income") are left untouched.
    """
    transfer_total: float = 0.0
    for it in items:
        if (
            isinstance(it, dict)
            and it.get("category") == "reserve_expense"
            and it.get("reserve_group") == "transfer"
        ):
            transfer_total += float(it.get("annual_budget") or 0)

    if transfer_total == 0.0:
        return items

    result = []
    for it in items:
        if (
            isinstance(it, dict)
            and it.get("category") == "reserve_income"
            and it.get("reserve_group") != "income"
        ):
            result.append({**it, "annual_budget": transfer_total})
        else:
            result.append(it)
    return result


def _now_text() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _storage_root() -> Path:
    root = Path(settings.BUDGET_STORAGE_ROOT)
    root.mkdir(parents=True, exist_ok=True)
    return root


def _json_dumps(value: Any) -> str:
    return json.dumps(value, separators=(",", ":"), sort_keys=True)


def _json_loads(value: Optional[str], default: Any) -> Any:
    if not value:
        return default
    return json.loads(value)


def _assessment_mapping_category(raw_category: object) -> str:
    category = str(raw_category or "").lower()
    if category == "income":
        return "income"
    if category == "reserve_income":
        return "reserve_income"
    if category in {"reserve", "reserve_expense"}:
        return "reserve_expense"
    return "operating"


def _assessment_mapping_fund_type(category: str) -> str:
    return "reserve" if category in {"reserve_income", "reserve_expense"} else "operating"


def _line_item_to_assessment_mapping_line(item: dict[str, Any]) -> dict[str, Any]:
    label = str(item.get("label") or item.get("line_item_key") or "")
    normalized_label = " ".join(label.lower().split())
    category = _assessment_mapping_category(item.get("category"))
    account_code = item.get("account_code")
    amount, source_column_used = select_assessment_mapping_amount(item)
    return {
        "label": label,
        "normalized_label": normalized_label,
        "section": str((item.get("raw") or {}).get("section") or category),
        "category": category,
        "fund_type": _assessment_mapping_fund_type(category),
        "account_code": str(account_code) if account_code not in (None, "") else None,
        "annual_budget": item.get("annual_budget"),
        "proposed_amount": (
            item.get("proposed_amount")
            if item.get("proposed_amount") is not None
            else item.get("proposedAmount")
        ),
        "projection": item.get("projection"),
        "assessment_mapping_amount": float(amount) if amount is not None else None,
        "source_column_used": source_column_used,
        "amount": float(amount) if amount is not None else None,
        "reserve_group": item.get("reserve_group") or item.get("reserveGroup"),
        "active": not bool(item.get("inactive")),
    }


def _materialize_assessment_mappings_for_line_items(
    session: Session,
    *,
    hoa_id: int,
    line_items: list[dict[str, Any]],
) -> dict[str, int]:
    raw_conn = session.connection().connection
    property_row = raw_conn.execute(
        "SELECT default_assessment_setup_id FROM properties WHERE id = ?",
        (hoa_id,),
    ).fetchone()
    setup_id = property_row[0] if property_row else None
    if not setup_id:
        setup_row = raw_conn.execute(
            """
            SELECT id
              FROM assessment_setups
             WHERE property_id = ?
               AND status = 'approved'
             ORDER BY id DESC
             LIMIT 1
            """,
            (hoa_id,),
        ).fetchone()
        setup_id = setup_row[0] if setup_row else None
    if not setup_id:
        return {"auto_approved": 0, "manual_preserved": 0, "suggested": 0, "conflict": 0, "unmatched": 0}

    budget_lines = [
        _line_item_to_assessment_mapping_line(item)
        for item in line_items
        if isinstance(item, dict)
    ]
    return materialize_budget_line_pool_mappings(
        property_id=hoa_id,
        assessment_setup_id=int(setup_id),
        budget_lines=budget_lines,
        connection=session.connection().connection,
        commit=False,
    )


def _write_atomic_bytes(destination: Path, payload: bytes) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=destination.parent, delete=False) as handle:
        handle.write(payload)
        temp_path = Path(handle.name)
    temp_path.replace(destination)


def _relative_storage_path(*parts: object) -> str:
    return str(Path(*map(str, parts)))


def _budget_storage_path(relative_path: str) -> Path:
    return _storage_root() / relative_path


def _storage_file_available(relative_path: Optional[str]) -> bool:
    return bool(relative_path and _budget_storage_path(relative_path).exists())


def _write_temp_workbook(file_bytes: bytes, original_filename: str) -> str:
    suffix = Path(original_filename or "upload.xlsx").suffix or ".xlsx"
    os.makedirs(settings.TEMP_DIR, exist_ok=True)
    fd, path = tempfile.mkstemp(suffix=suffix, dir=settings.TEMP_DIR)
    os.close(fd)
    Path(path).write_bytes(file_bytes)
    return path


def _ensure_xlsx(temp_input_path: str) -> str:
    """Convert .xls or .pdf to .xlsx for pipeline processing. Returns the .xlsx path."""
    ext = Path(temp_input_path).suffix.lower()
    if ext == ".xls":
        from .income_statement_parser import _read_xls_rows
        from openpyxl import Workbook as _Workbook
        rows = _read_xls_rows(temp_input_path)
        wb = _Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        xlsx_path = temp_input_path.replace(".xls", ".xlsx")
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path
    elif ext == ".pdf":
        from .income_statement_parser import _read_pdf_rows
        from openpyxl import Workbook as _Workbook
        rows = _read_pdf_rows(temp_input_path)
        wb = _Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        xlsx_path = temp_input_path.rsplit(".", 1)[0] + ".xlsx"
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path
    return temp_input_path


def _draft_enriched_storage_key(hoa_id: int, draft_id: int) -> str:
    return _relative_storage_path("hoa", hoa_id, "drafts", draft_id, "enriched.xlsx")


def _extract_statement_month(fiscal_year_start_month: int, detected_months: int) -> int:
    return (fiscal_year_start_month + detected_months - 2) % 12 + 1


def _canonical_statement_from_line_items(
    line_items: list[dict[str, Any]],
    *,
    family: str,
) -> ExtractedFinancialStatement:
    canonical_items: list[ExtractedStatementLineItem] = []

    def _opt_float(value: Any) -> Optional[float]:
        return _parse_float(value) if value is not None else None

    for item in line_items:
        raw_section = str(item.get("raw", {}).get("section") or item.get("section") or item.get("category") or "")
        account_code = item.get("account_code")
        canonical_items.append(
            ExtractedStatementLineItem(
                account_code_text=None if account_code is None else str(account_code),
                label=str(item.get("label") or item.get("name") or ""),
                section_label=raw_section or None,
                section_kind=_coerce_canonical_section_kind(item.get("category")),
                # Pass through every numeric field the validator can use.
                # Previously only ytd_actual and annual_budget were forwarded,
                # which made Cummins Park-style "budgeted but no actual yet"
                # rows look zero-only and trip the validator.
                current_actual=_opt_float(item.get("current_actual")),
                current_budget=_opt_float(item.get("current_budget")),
                current_variance=_opt_float(item.get("current_variance")),
                ytd_actual=_opt_float(item.get("ytd_actual")),
                ytd_budget=_opt_float(item.get("ytd_budget")),
                ytd_variance=_opt_float(item.get("ytd_variance")),
                annual_budget=_opt_float(item.get("annual_budget")),
            )
        )
    return ExtractedFinancialStatement(
        document_family=family,
        report_type="income_statement",
        line_items=canonical_items,
        confidence=1.0,
    )


def _build_review_required_response(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    upload: BudgetUpload,
    original_filename: str,
    reason: str,
    code: str,
    warnings: Optional[list[str]] = None,
    details: Optional[dict[str, Any]] = None,
) -> BudgetUploadResponse:
    upload.enrichment_status = "failed"
    debug_info = ExtractionDebugInfo(code=code, message=reason, details=details or {})
    review_event = _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="enrichment_review_required",
        summary=f"Review required for {original_filename}",
        upload_id=upload.id,
        payload={"code": code, "reason": reason, "details": debug_info.details},
    )
    session.commit()
    return BudgetUploadResponse(
        upload_id=upload.id,
        draft=None,
        timeline_event=_serialize_timeline_event(review_event),
        warnings=warnings or [],
        review_required=True,
        review_reason=reason,
        debug_info=debug_info,
    )


def _build_income_statement_validation_feedback(
    *,
    original_filename: str,
    issues: list[dict[str, Any]],
) -> tuple[str, list[str]]:
    blocking_issues = [issue for issue in issues if issue.get("severity") == "error"]
    primary_issue = blocking_issues[0] if blocking_issues else (issues[0] if issues else {})
    primary_message = str(primary_issue.get("message") or "The statement did not match the expected layout.")
    issue_codes = {str(issue.get("code") or "") for issue in issues}
    likely_wrong_report = bool(
        issue_codes
        & {
            "missing_annual_budget_coverage",
            "missing_numeric_coverage",
            "suspicious_zero_heavy_output",
        }
    )

    reason = f"{original_filename} was not accepted as an income statement. {primary_message}"
    if likely_wrong_report:
        reason += (
            " This commonly happens when the uploaded workbook is an annual/monthly budget "
            "or cash-flow report instead of the income statement used for budget drafting."
        )

    warnings: list[str] = []
    for issue in blocking_issues or issues:
        message = str(issue.get("message") or "").strip()
        if not message:
            continue
        details = issue.get("details")
        if isinstance(details, dict) and issue.get("code") == "missing_annual_budget_coverage":
            rows = details.get("annual_budget_populated_rows")
            total = details.get("line_item_count")
            if rows is not None and total is not None:
                message = f"{message} Detected annual-budget coverage: {rows}/{total} parsed rows."
        warnings.append(message)

    warnings.extend(_EXPECTED_INCOME_STATEMENT_GUIDANCE)
    return reason, warnings


def _build_proforma_validation_feedback(
    *,
    original_filename: str,
    issues: list[dict[str, Any]],
) -> tuple[str, list[str]]:
    primary_issue = issues[0] if issues else {}
    primary_message = str(
        primary_issue.get("message")
        or "The workbook did not provide enough final annual budget coverage."
    )
    reason = (
        f"{original_filename} was not accepted as a pro forma / final budget source. "
        f"{primary_message}"
    )
    warnings = [
        "Expected a spreadsheet-export operating/cash-flow budget with Jan-Dec columns plus one annual final, proposed, or budget column.",
        "The annual final/proposed/budget column must contain enough populated values to build a safe draft.",
    ]
    for issue in issues:
        message = str(issue.get("message") or "").strip()
        if message:
            warnings.append(message)
    return reason, warnings


def _extract_pdf_statement_sync(
    path: str,
    *,
    source_mode: str = SOURCE_MODE_INCOME_STATEMENT,
) -> ExtractedFinancialStatement | DocumentExtractionFailure:
    def _extract_coro():
        try:
            return extract_pdf_statement(path, source_mode=source_mode)
        except TypeError as exc:
            if "source_mode" not in str(exc):
                raise
            return extract_pdf_statement(path)

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None

    if loop and loop.is_running():
        import concurrent.futures

        with concurrent.futures.ThreadPoolExecutor() as pool:
            return pool.submit(asyncio.run, _extract_coro()).result()

    return asyncio.run(_extract_coro())


async def _extract_proforma_excel_statement(
    path: str,
    *,
    original_filename: Optional[str] = None,
) -> ExtractedFinancialStatement | DocumentExtractionFailure:
    """Use Gemini for semantic extraction from a serialized workbook grid."""
    from ..ai_implementation.pipeline.llm_client import call_llm

    try:
        workbook_text = _extract_excel_workbook_prompt_text(path)
    except Exception as exc:
        return DocumentExtractionFailure(
            code="provider_error",
            message=f"The Excel workbook could not be prepared for Gemini extraction: {exc}",
            details={"error": str(exc)},
        )

    filename = original_filename or Path(path).name
    messages = [
        {"role": "system", "content": _PROFORMA_EXCEL_SYSTEM_PROMPT},
        {
            "role": "user",
            "content": (
                f"Filename: {filename}\n"
                "Workbook cell dump with A1 coordinates. Blank cells are omitted.\n\n"
                f"{workbook_text}\n\n"
                "Extract the pro forma / final budget line items from this workbook."
            ),
        },
    ]

    try:
        raw_result = await call_llm(
            messages,
            ExtractedFinancialStatementPage,
            temperature=0.0,
            timeout=_PROFORMA_EXCEL_GEMINI_TIMEOUT_SECONDS,
        )
        if raw_result is None:
            raise RuntimeError("Gemini extraction returned no structured result.")
        return ExtractedFinancialStatement(
            document_family="excel_budget_workbook",
            report_type="income_statement",
            statement_period=raw_result.statement_period,
            line_items=raw_result.line_items,
            totals=raw_result.totals,
            validation_issues=raw_result.validation_issues,
            confidence=raw_result.confidence,
            extraction_metadata={"extractor": "gemini_excel_text"},
        )
    except ValidationError as exc:
        return DocumentExtractionFailure(
            code="schema_validation_failed",
            message="Structured Excel extraction could not satisfy the canonical schema.",
            details={"error": str(exc)},
        )
    except Exception as exc:
        return DocumentExtractionFailure(
            code="provider_error",
            message=f"The Gemini Excel extraction provider failed: {exc}",
            details={"error": str(exc)},
        )


def _extract_proforma_excel_statement_sync(
    path: str,
    *,
    original_filename: Optional[str] = None,
) -> ExtractedFinancialStatement | DocumentExtractionFailure:
    async def _extract_coro():
        return await _extract_proforma_excel_statement(
            path,
            original_filename=original_filename,
        )

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None

    if loop and loop.is_running():
        import concurrent.futures

        with concurrent.futures.ThreadPoolExecutor() as pool:
            return pool.submit(asyncio.run, _extract_coro()).result()

    return asyncio.run(_extract_coro())


def _extract_reserve_study_sync(path: str) -> ExtractedReserveStudyDocument | DocumentExtractionFailure:
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None

    if loop and loop.is_running():
        import concurrent.futures

        with concurrent.futures.ThreadPoolExecutor() as pool:
            return pool.submit(asyncio.run, extract_reserve_study(path)).result()

    return asyncio.run(extract_reserve_study(path))


def _parse_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "—"}:
            return 0.0
        text = text.replace(",", "").replace("$", "")
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def _statement_to_budget_line_items(
    statement: ExtractedFinancialStatement,
    *,
    source_mode: str,
) -> list[dict[str, Any]]:
    line_items: list[dict[str, Any]] = []
    for item in statement.line_items:
        category = item.section_kind or "operating"
        evidence = item.evidence or {}
        source_column = evidence.get("source_column")
        if not isinstance(source_column, str) or not source_column.strip():
            source_column = (
                "final_budget"
                if _is_proforma_source_mode(source_mode)
                else "annual_budget"
            )
        reserve_group = None
        if category == "reserve_income":
            reserve_group = "income"
        elif category == "reserve_expense":
            reserve_group = "component"

        account_code = _normalize_optional_text(item.account_code_text)
        line_items.append(
            {
                "line_item_key": account_code or item.label,
                "account_code": int(account_code) if account_code and account_code.isdigit() else account_code,
                "label": item.label,
                "name": item.label,
                "category": category,
                "current_actual": item.current_actual,
                "current_budget": item.current_budget,
                "current_variance": item.current_variance,
                "ytd_actual": item.ytd_actual,
                "ytd_budget": item.ytd_budget,
                "ytd_variance": item.ytd_variance,
                "annual_budget": item.annual_budget,
                "projection": item.annual_budget,
                "percent_change": 0.0,
                "read_only": _effective_read_only({}, category),
                "reserve_group": reserve_group,
                "source_column": source_column,
                "source_page_or_cell": (
                    f"page {item.page_number}" if item.page_number is not None else None
                ),
                "raw": {
                    "section": item.section_label or category,
                    "page_number": item.page_number,
                },
            }
        )
    return line_items


def _normalize_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_compare_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _normalize_line_item_key(value: Any) -> Optional[str]:
    text = _normalize_optional_text(value)
    return text if text is not None else None


def _line_item_amount(item: Optional[dict[str, Any]]) -> float:
    if not item:
        return 0.0
    proposed_amount = item.get("proposed_amount")
    if proposed_amount is None:
        proposed_amount = item.get("proposedAmount")
    if proposed_amount is not None:
        return _parse_float(proposed_amount)

    annual_budget = _parse_float(item.get("annual_budget"))
    raw_percent_change = item.get("percent_change")
    if raw_percent_change is None:
        raw_percent_change = item.get("percentChange")
    percent_change = _parse_float(raw_percent_change)
    return annual_budget * (1 + (percent_change / 100.0))


def _line_item_category(item: Optional[dict[str, Any]]) -> str:
    if not item:
        return "operating"
    category = _normalize_optional_text(item.get("category"))
    if category:
        return category.lower()
    label = str(item.get("label") or "")
    account_code = item.get("account_code")
    section = str(item.get("raw", {}).get("section") or "")
    return _infer_category(label, _extract_account_code(str(account_code or label)), section)


def _line_item_label(item: Optional[dict[str, Any]]) -> str:
    if not item:
        return ""
    return _normalize_optional_text(item.get("label")) or _normalize_optional_text(item.get("name")) or ""


def _line_item_account_code_text(item: Optional[dict[str, Any]]) -> Optional[str]:
    if not item:
        return None
    account_code = item.get("account_code")
    if account_code is None:
        account_code = _extract_account_code(_line_item_label(item))
    return _normalize_optional_text(account_code)


def _line_item_match_tokens(item: Optional[dict[str, Any]]) -> list[str]:
    if not item:
        return []
    tokens: list[str] = []
    line_item_key = _normalize_line_item_key(item.get("line_item_key"))
    label = _normalize_compare_text(_line_item_label(item))
    account_code = _line_item_account_code_text(item)
    for token in (line_item_key, label, account_code):
        normalized = _normalize_compare_text(token)
        if normalized and normalized not in tokens:
            tokens.append(normalized)
    return tokens


def _build_line_item_index(line_items: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for item in line_items:
        for token in _line_item_match_tokens(item):
            index.setdefault(token, item)
    return index


def _find_matching_line_item(
    item: Optional[dict[str, Any]],
    *,
    index: dict[str, dict[str, Any]],
) -> Optional[dict[str, Any]]:
    for token in _line_item_match_tokens(item):
        match = index.get(token)
        if match is not None:
            return match
    return None


def _is_reserve_item(item: Optional[dict[str, Any]]) -> bool:
    if not item:
        return False
    cat = _line_item_category(item)
    return cat in ("reserve", "reserve_income", "reserve_expense")


def _reserve_group_for_item(item: Optional[dict[str, Any]]) -> Optional[str]:
    if not _is_reserve_item(item):
        return None
    explicit_group = _normalize_optional_text(
        item.get("reserve_group") if item else None
    ) or _normalize_optional_text(item.get("reserveGroup") if item else None)
    if explicit_group in {"component", "income", "transfer"}:
        return explicit_group
    normalized_label = _normalize_compare_text(_line_item_label(item))
    normalized_section = _normalize_compare_text(item.get("raw", {}).get("section") if item else "")
    if normalized_section == "reserve income" or "reserve income" in normalized_section:
        return "income"
    if "allocation to reserves" in normalized_section or normalized_section == "allocation to reserves":
        return "transfer"
    # Reserve expense items default to component
    if "reserve expense" in normalized_section:
        return "component"
    # Label-based fallbacks for pre-existing data
    if "reserve income" in normalized_label or "interest earned reserve" in normalized_label or "change in asset value" in normalized_label:
        return "income"
    if "allocation/transfer" in normalized_label or "transfer" in normalized_label:
        return "transfer"
    return "component"


def _is_reserve_overlay_excluded(label: str) -> bool:
    normalized_label = _normalize_compare_text(label)
    return (
        "reserve income" in normalized_label
        or "allocation/transfer" in normalized_label
        or "transfer" in normalized_label
    )


def _is_reserve_overlay_eligible(row: dict[str, Any]) -> bool:
    return (
        bool(row.get("is_reserve"))
        and str(row.get("reserve_group") or "") == "component"
        and _parse_float(row.get("baseline_amount")) > 0.0
    )


def _apply_reserve_inflation_overlay(
    rows: list[dict[str, Any]],
    *,
    reserve_inflation_rate: float,
) -> list[dict[str, Any]]:
    overlaid_rows: list[dict[str, Any]] = []
    for row in rows:
        updated_row = dict(row)
        eligible = _is_reserve_overlay_eligible(updated_row)
        baseline_amount = _parse_float(updated_row.get("baseline_amount"))
        updated_row["reserve_inflation_eligible"] = eligible
        updated_row["inflation_adjusted_baseline_amount"] = (
            baseline_amount * (1 + reserve_inflation_rate)
            if reserve_inflation_rate > 0.0 and eligible
            else baseline_amount
        )
        comparison_baseline_amount = _parse_float(updated_row["inflation_adjusted_baseline_amount"])
        current_amount = _parse_float(updated_row.get("current_amount"))
        updated_row["comparison_baseline_amount"] = comparison_baseline_amount
        updated_row["delta_amount"] = current_amount - comparison_baseline_amount
        updated_row["delta_percent"] = (
            None
            if comparison_baseline_amount == 0.0
            else ((current_amount - comparison_baseline_amount) / comparison_baseline_amount) * 100
        )
        updated_row["changed"] = abs(updated_row["delta_amount"]) > 1e-9
        overlaid_rows.append(updated_row)
    return overlaid_rows


def _extract_account_code(label: str) -> Optional[int]:
    head = (label or "").split("-", 1)[0].strip()
    return int(head) if head.isdigit() else None


def _coerce_canonical_section_kind(value: Any) -> Optional[str]:
    """Coerce a category value into canonical 4-value taxonomy, or None.

    Used by `_canonical_statement_from_line_items` to ensure emitted
    `section_kind` values always match SECTION_KINDS.
    """
    if value is None:
        return None
    text = str(value).strip().lower()
    return text if text in SECTION_KINDS else None


def _infer_category(label: str, account_code: Optional[int], section: str) -> str:
    """Infer canonical 4-value category from section state.

    Returns one of SECTION_KINDS: "income", "operating",
    "reserve_income", "reserve_expense".

    Delegates to the parser's `_match_section_header` (same prefix logic as
    parse_rows_with_sections) so the PDF pre-processed path stays in lockstep
    with the raw Excel path. Falls back to account-code ranges when no
    section header text matches.

    IMPORTANT: Does NOT check label keywords. Section position is authoritative.
    """
    # Direct match against canonical values (for already-classified rows)
    raw_normalized = (section or "").strip().lower()
    if raw_normalized in SECTION_KINDS:
        return raw_normalized

    # Section header prefix match (same logic as the raw-statement parser)
    matched = _match_section_header(section or "")
    if matched is not None:
        return matched

    # Fallback to account code ranges (returns one of SECTION_KINDS)
    return _classify_by_account_code(account_code)


def _line_items_to_percent_changes(line_items: list[dict[str, Any]]) -> dict[str, float]:
    changes: dict[str, float] = {}
    for item in line_items:
        if item.get("read_only") or item.get("readOnly"):
            continue
        label = macros_service._normalize_label(str(item.get("label") or item.get("name") or ""))
        if not label:
            continue
        raw_percent_change = item.get("percent_change")
        if raw_percent_change is None:
            raw_percent_change = item.get("percentChange")
        if raw_percent_change is None:
            continue
        changes[label] = _parse_float(raw_percent_change) / 100.0
    return changes


def _table_to_line_items(table: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    """Returns (line_items, warnings)."""
    warnings: list[str] = []
    headers = [str(header or f"column_{index}") for index, header in enumerate(table.get("headers", []), start=1)]

    if "Label" not in headers:
        # Raw income statement layout — use section-aware parser
        rows = table.get("rows", [])
        col_indices = detect_columns([table.get("headers", [])] + rows[:9])
        detection_tier = col_indices.pop("_detection_tier", 1)
        if detection_tier == 3:
            warnings.append(
                "Heads up \u2014 the column layout in your file looks a bit different from what we "
                "usually see. The numbers might not line up perfectly. If anything looks off, try "
                "re-uploading the standard monthly income statement from your accounting software."
            )
        # Scan all header/early rows for enrichment column labels (written by pipeline at row 5)
        for scan_row in ([table.get("headers", [])] + rows[:9]):
            for ci, cell in enumerate(scan_row or []):
                cell_text = str(cell or "").strip().lower()
                if cell_text == "projection" and "projection" not in col_indices:
                    col_indices["projection"] = ci
                elif cell_text in ("% change", "percent change") and "percent_change" not in col_indices:
                    col_indices["percent_change"] = ci
        col_indices.setdefault("projection", _COL_PROJECTION_INDEX)
        col_indices.setdefault("percent_change", _COL_PERCENT_CHANGE_INDEX)
        return parse_rows_with_sections(rows, col_indices), warnings

    # Pre-processed table with "Label" header (from enriched/generated workbook)
    line_items: list[dict[str, Any]] = []
    for row in table.get("rows", []):
        item = {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}

        # Skip section header marker rows. The normalized workbook inserts rows
        # where col A (Section) has the header text and all other cols are None.
        # These are structural markers, NOT line items, and must not produce
        # ghost "Line Item N" entries in the output.
        raw_label = item.get("Label") or item.get("Account Name") or item.get("Description")
        raw_account_code = item.get("Account Code") or item.get("Account")
        if not raw_label and not raw_account_code:
            continue

        label = raw_label or f"Line Item {len(line_items) + 1}"
        account_code = raw_account_code
        line_item_key = str(account_code or label)
        section = item.get("Section") or item.get("section") or ""
        category = _infer_category(
            str(label),
            _extract_account_code(str(account_code or label)),
            section,
        )
        normalized = {
            "line_item_key": line_item_key,
            "account_code": account_code,
            "label": label,
            "category": category,
            # Project ALL seven numeric columns from the enriched workbook so
            # downstream validation (`validate_extracted_statement`) can see
            # them. Cummins Park exposed why this matters: that PDF has rows
            # with current/ytd budgets but $0 actuals, so dropping ytd_budget
            # and current_budget made every "no actual spend yet" row look
            # zero-only and tripped `suspicious_zero_heavy_output`.
            "current_actual": item.get("Current Actual"),
            "current_budget": item.get("Current Budget"),
            "current_variance": item.get("Current Variance"),
            "ytd_actual": item.get("YTD Actual"),
            "ytd_budget": item.get("YTD Budget"),
            "ytd_variance": item.get("YTD Variance"),
            "annual_budget": item.get("Annual Budget"),
            "projection": item.get("Projection"),
            "percent_change": item.get("% Change") or item.get("Percent Change"),
            # Use _effective_read_only so per-line overrides (read_only_override)
            # take precedence over the section default. New items have no override,
            # so this is equivalent to `category in READ_ONLY_SECTIONS` on initial
            # creation; the resolver applies on re-serialization.
            "read_only": _effective_read_only(item, category),
            # Per the DRE-driven assessment engine invariant
            # (BudgetDraft.line_items.amount = annual), audit which source
            # column the annual value came from. "annual_budget" here means
            # the value was promoted from the Annual Budget column of the
            # enriched/extracted table; the engine consumes ONLY this field.
            "source_column": "annual_budget",
            "source_page_or_cell": item.get("Source Page") or item.get("Source Cell"),
            "raw": item,
        }
        line_items.append(normalized)
    return line_items, warnings


def _extract_totals_from_preview(preview: Optional[dict[str, Any]]) -> tuple[float, float, float]:
    rows = preview.get("rows", []) if preview else []
    totals = {
        "total income": 0.0,
        "total expense": 0.0,
        "net operating income": 0.0,
    }
    for row in rows:
        if len(row) < 2:
            continue
        # Find the label in any of the first few cells (handles both old 12-month
        # layout where col A has labels, and new 8-column reference format where
        # col A is account code and col B is the label).
        key = None
        for cell in row[:3]:
            candidate = str(cell or "").strip().lower()
            if candidate in totals:
                key = candidate
                break
        if key is None:
            continue
        # Find the largest absolute numeric value in the row — this is the annual
        # total regardless of which column layout is used.
        best = 0.0
        for cell in row:
            try:
                value = float(cell or 0)
                if abs(value) > abs(best):
                    best = value
            except (TypeError, ValueError):
                continue
        totals[key] = best
    return totals["total income"], totals["total expense"], totals["net operating income"]


def _actor_name(actor: dict[str, Any]) -> str:
    return str(actor.get("name") or actor.get("email") or "Unknown User")


def _serialize_timeline_event(event: BudgetAuditEvent) -> BudgetTimelineEvent:
    return BudgetTimelineEvent(
        id=event.id,
        event_type=event.event_type,
        summary=event.summary,
        actor_name=event.actor_name,
        occurred_at=event.created_at,
        related_upload_id=event.upload_id,
        related_draft_id=event.draft_id,
        related_version_id=event.version_id,
        related_note_id=event.note_id,
        file_name=None,
        version_code=None,
        payload=_json_loads(event.payload_json, None),
    )


def _serialize_draft(draft: BudgetDraft, upload: Optional[BudgetUpload] = None) -> BudgetDraftPayload:
    reserve_study_rows, _ = canonicalize_reserve_study_row_dicts(_json_loads(draft.reserve_study_rows_json, []))
    raw_line_items: list[dict[str, Any]] = _json_loads(draft.line_items_json, [])
    # Derive reserve income from the allocation transfer so they always match.
    line_items = _derive_reserve_income(raw_line_items)
    return BudgetDraftPayload(
        id=draft.id,
        status=draft.status,
        source_upload_id=draft.source_upload_id,
        reserve_study_upload_id=draft.reserve_study_upload_id,
        reopened_from_version_id=draft.reopened_from_version_id,
        line_items=line_items,
        reserve_study_status=draft.reserve_study_status or "none",
        reserve_study_rows=reserve_study_rows,
        reserve_study_warnings=_json_loads(draft.reserve_study_warnings_json, []),
        global_note=draft.global_note,
        statement_month=draft.statement_month,
        growth_factor=draft.growth_factor,
        growth_factor_note=draft.growth_factor_note,
        reserve_inflation_rate=draft.reserve_inflation_rate,
        reserve_inflation_note=draft.reserve_inflation_note,
        version_int=draft.version_int,
        updated_at=draft.updated_at,
        upload_filename=upload.original_filename if upload else None,
        enriched_file_available=_storage_file_available(draft.enriched_storage_key),
        source_mode=str(
            draft.source_mode
            or (upload.source_mode if upload else None)
            or SOURCE_MODE_INCOME_STATEMENT
        ),
        assessment_mode=normalize_assessment_mode(
            getattr(draft, "assessment_mode", None)
            or (getattr(upload, "assessment_mode", None) if upload else None)
            or ASSESSMENT_MODE_VARIABLE
        ),
    )


def _serialize_draft_summary(
    draft: BudgetDraft,
    *,
    upload: Optional[BudgetUpload] = None,
    reopened_from_version: Optional[BudgetVersion] = None,
) -> BudgetDraftSummary:
    return BudgetDraftSummary(
        id=draft.id,
        status=draft.status,
        source_upload_id=draft.source_upload_id,
        reserve_study_upload_id=draft.reserve_study_upload_id,
        source_upload_filename=upload.original_filename if upload else None,
        reopened_from_version_id=draft.reopened_from_version_id,
        reopened_from_version_code=reopened_from_version.version_code if reopened_from_version else None,
        reserve_inflation_rate=draft.reserve_inflation_rate,
        reserve_inflation_note=draft.reserve_inflation_note,
        version_int=draft.version_int,
        reserve_study_status=draft.reserve_study_status or "none",
        updated_at=draft.updated_at,
        actor_name=draft.actor_name,
        enriched_file_available=_storage_file_available(draft.enriched_storage_key),
        source_mode=str(
            draft.source_mode
            or (upload.source_mode if upload else None)
            or SOURCE_MODE_INCOME_STATEMENT
        ),
        assessment_mode=normalize_assessment_mode(
            getattr(draft, "assessment_mode", None)
            or (getattr(upload, "assessment_mode", None) if upload else None)
            or ASSESSMENT_MODE_VARIABLE
        ),
    )


def _serialize_note(note: BudgetNote) -> BudgetNoteRecord:
    return BudgetNoteRecord(
        id=note.id,
        note_scope=note.note_scope,
        line_item_key=note.line_item_key,
        title=note.title,
        body=note.body,
        created_at=note.created_at,
        created_by_name=note.created_by_name,
        upload_id=note.upload_id,
        draft_id=note.draft_id,
        version_id=note.version_id,
    )


def _serialize_version_summary(
    version: BudgetVersion,
    upload: Optional[BudgetUpload] = None,
) -> BudgetVersionSummary:
    return BudgetVersionSummary(
        id=version.id,
        version_number=version.version_number,
        version_code=version.version_code,
        stage=version.stage,
        label=version.label,
        summary_note=version.summary_note,
        total_income=version.total_income,
        total_expense=version.total_expense,
        net_operating_income=version.net_operating_income,
        growth_factor=version.growth_factor,
        growth_factor_note=version.growth_factor_note,
        reserve_inflation_rate=version.reserve_inflation_rate,
        reserve_inflation_note=version.reserve_inflation_note,
        statement_month=version.statement_month,
        created_at=version.created_at,
        created_by_name=version.created_by_name,
        source_draft_id=version.source_draft_id,
        output_storage_key=version.output_storage_key,
        source_upload_filename=upload.original_filename if upload else None,
        source_mode=str(
            version.source_mode
            or (upload.source_mode if upload else None)
            or SOURCE_MODE_INCOME_STATEMENT
        ),
        assessment_mode=normalize_assessment_mode(
            getattr(version, "assessment_mode", None)
            or (getattr(upload, "assessment_mode", None) if upload else None)
            or ASSESSMENT_MODE_VARIABLE
        ),
    )


def _serialize_version_detail(
    version: BudgetVersion,
    upload: Optional[BudgetUpload] = None,
) -> BudgetVersionDetail:
    return BudgetVersionDetail(
        **_serialize_version_summary(version, upload).model_dump(),
        source_upload_id=version.source_upload_id,
        reopened_from_version_id=version.reopened_from_version_id,
        fiscal_year_start_month=version.fiscal_year_start_month,
        fiscal_year_end_month=version.fiscal_year_end_month,
        line_items=_json_loads(version.line_items_json, []),
        budget_preview=_json_loads(version.budget_preview_json, None),
    )


def _serialize_version_compare_card(
    version: BudgetVersion,
    upload: Optional[BudgetUpload],
) -> BudgetVersionCompareCard:
    return BudgetVersionCompareCard(
        id=version.id,
        version_code=version.version_code,
        stage=version.stage,
        label=version.label,
        summary_note=version.summary_note,
        created_at=version.created_at,
        created_by_name=version.created_by_name,
        source_upload_filename=upload.original_filename if upload else None,
        total_income=version.total_income,
        total_expense=version.total_expense,
        net_operating_income=version.net_operating_income,
        growth_factor=version.growth_factor,
        growth_factor_note=version.growth_factor_note,
        statement_month=version.statement_month,
        fiscal_year_start_month=version.fiscal_year_start_month,
        fiscal_year_end_month=version.fiscal_year_end_month,
        source_mode=str(
            version.source_mode
            or (upload.source_mode if upload else None)
            or SOURCE_MODE_INCOME_STATEMENT
        ),
        assessment_mode=normalize_assessment_mode(
            getattr(version, "assessment_mode", None)
            or (getattr(upload, "assessment_mode", None) if upload else None)
            or ASSESSMENT_MODE_VARIABLE
        ),
    )


def _get_property(session: Session, hoa_id: int) -> Property:
    hoa = session.get(Property, hoa_id)
    if hoa is None:
        raise LookupError("HOA not found")
    return hoa


def _get_draft(session: Session, hoa_id: int, draft_id: int) -> BudgetDraft:
    draft = session.get(BudgetDraft, draft_id)
    if draft is None or draft.property_id != hoa_id:
        raise LookupError("Draft not found")
    return draft


def _get_editable_draft(session: Session, hoa_id: int, draft_id: int) -> BudgetDraft:
    draft = _get_draft(session, hoa_id, draft_id)
    if draft.status != BUDGET_DRAFT_ACTIVE:
        raise ValueError("Requested draft is no longer active")
    return draft


def _get_version(session: Session, hoa_id: int, version_id: int) -> BudgetVersion:
    version = session.get(BudgetVersion, version_id)
    if version is None or version.property_id != hoa_id:
        raise LookupError("Version not found")
    return version


def _get_upload(session: Session, upload_id: Optional[int]) -> Optional[BudgetUpload]:
    if upload_id is None:
        return None
    return session.get(BudgetUpload, upload_id)


def _get_reserve_study_upload(session: Session, hoa_id: int, upload_id: int) -> BudgetUpload:
    upload = session.get(BudgetUpload, upload_id)
    if (
        upload is None
        or upload.property_id != hoa_id
        or upload.document_role != "reserve_study"
    ):
        raise LookupError("Reserve study upload not found")
    return upload


def _get_income_statement_upload(session: Session, hoa_id: int, upload_id: int) -> BudgetUpload:
    upload = session.get(BudgetUpload, upload_id)
    if (
        upload is None
        or upload.property_id != hoa_id
        or upload.document_role != "budget_source"
    ):
        raise LookupError("Income statement upload not found")
    return upload


def _replace_active_draft(session: Session, hoa_id: int, timestamp: str) -> None:
    active_drafts = session.scalars(
        select(BudgetDraft).where(
            BudgetDraft.property_id == hoa_id,
            BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
        )
    ).all()
    for existing in active_drafts:
        existing.status = BUDGET_DRAFT_SUPERSEDED
        existing.updated_at = timestamp


def _create_audit_event(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    event_type: str,
    summary: str,
    upload_id: Optional[int] = None,
    draft_id: Optional[int] = None,
    version_id: Optional[int] = None,
    note_id: Optional[int] = None,
    payload: Optional[dict[str, Any]] = None,
) -> BudgetAuditEvent:
    event = BudgetAuditEvent(
        property_id=hoa_id,
        upload_id=upload_id,
        draft_id=draft_id,
        version_id=version_id,
        note_id=note_id,
        event_type=event_type,
        summary=summary,
        actor_user_id=actor["id"],
        actor_name=_actor_name(actor),
        payload_json=_json_dumps(payload) if payload is not None else None,
        created_at=_now_text(),
    )
    session.add(event)
    session.flush()
    return event


def _create_upload_record(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    original_filename: str,
    content_type: Optional[str],
    file_bytes: bytes,
    timestamp: str,
    document_role: str,
    enrichment_status: str,
    source_mode: str = SOURCE_MODE_INCOME_STATEMENT,
    assessment_mode: str = ASSESSMENT_MODE_VARIABLE,
) -> BudgetUpload:
    sha256 = hashlib.sha256(file_bytes).hexdigest()
    upload = BudgetUpload(
        property_id=hoa_id,
        document_role=document_role,
        source_mode=source_mode,
        assessment_mode=normalize_assessment_mode(assessment_mode),
        original_filename=original_filename,
        storage_key="pending",
        content_type=content_type,
        byte_size=len(file_bytes),
        sha256=sha256,
        enrichment_status=enrichment_status,
        uploaded_by_user_id=actor["id"],
        uploaded_by_name=_actor_name(actor),
        created_at=timestamp,
    )
    session.add(upload)
    session.flush()

    file_ext = Path(original_filename).suffix.lower() or ".bin"
    filename = "source" if document_role == "budget_source" else "reserve-study"
    upload.storage_key = _relative_storage_path("hoa", hoa_id, "uploads", upload.id, f"{filename}{file_ext}")
    _write_atomic_bytes(_budget_storage_path(upload.storage_key), file_bytes)
    return upload


def _bundle_status_from_budget_response(
    response: BudgetUploadResponse,
    *,
    filename: str,
) -> BundleFileStatus:
    status = "review_required" if response.review_required else "completed"
    return BundleFileStatus(
        upload_id=response.upload_id,
        filename=filename,
        status=status,
        warnings=response.warnings,
        review_reason=response.review_reason,
        debug_info=response.debug_info,
    )


def _is_applied_reserve_study_line_item(item: dict[str, Any]) -> bool:
    raw = item.get("raw")
    raw_record = raw if isinstance(raw, dict) else {}
    line_item_key = str(item.get("line_item_key") or item.get("id") or "")
    return raw_record.get("source") == "reserve_study" or line_item_key.startswith("reserve-study::")


def _reserve_study_row_due_this_budget_year(row: dict[str, Any]) -> bool:
    if row.get("row_type") == "header":
        return False
    remaining_life = row.get("remaining_life")
    if remaining_life is None:
        return False
    try:
        return float(remaining_life) <= 1.0
    except (TypeError, ValueError):
        return False


def _reserve_study_row_to_budget_line_item(row: dict[str, Any]) -> dict[str, Any]:
    row_id = str(row.get("row_id") or row.get("line_item") or "reserve-study-row")
    label = str(row.get("line_item") or "Reserve Study Component")
    replacement_cost = row.get("replacement_cost")
    amount = _parse_float(replacement_cost)
    source_page = row.get("source_page")
    return {
        "id": f"reserve-study::{row_id}",
        "category": "reserve_expense",
        "name": label,
        "label": label,
        "line_item_key": f"reserve-study::{row_id}",
        "account_code": None,
        "ytdActual": 0.0,
        "ytd_actual": 0.0,
        "annualBudget": amount,
        "annual_budget": amount,
        "percentChange": 0.0,
        "percent_change": 0.0,
        "projection": amount,
        "readOnly": True,
        "read_only": True,
        "reserve_group": "component",
        "section": "Reserve Expenses (Reserve Study)",
        "raw": {
            "section": "Reserve Expenses (Reserve Study)",
            "source": "reserve_study",
            "row_id": row_id,
            "source_page": source_page,
            "flags": row.get("flags") or [],
        },
    }


def _render_draft_snapshot_from_upload(
    upload: BudgetUpload,
    *,
    line_items: list[dict[str, Any]],
    growth_factor: Optional[float],
    growth_factor_note: Optional[str],
) -> tuple[bytes, dict[str, Any]]:
    if not upload.storage_key:
        raise LookupError("Source upload file not found")

    route = choose_financial_document_route(upload.original_filename, upload.content_type)
    use_normalized_input = (
        route.path == "pdf_vlm"
        or _is_proforma_source_mode(upload.source_mode)
    )
    if use_normalized_input:
        canonical_statement = _canonical_statement_from_line_items(
            line_items,
            family=route.family or "pdf_visual_document",
        )
        temp_input_path = build_normalized_statement_workbook(canonical_statement)
    else:
        source_path = _budget_storage_path(upload.storage_key)
        if not source_path.exists():
            raise LookupError("Source upload file not found")
        temp_input_path = _write_temp_workbook(source_path.read_bytes(), upload.original_filename)
    temp_output_dir = Path(tempfile.mkdtemp(prefix="budget_draft_snapshot_"))
    try:
        temp_input_path = _ensure_xlsx(temp_input_path)

        pdf_known_columns = {"ytd_actual": 6, "annual_budget": 9} if use_normalized_input else None
        macros_service.write_percent_changes_by_label(
            temp_input_path,
            "Income Statement",
            _line_items_to_percent_changes(line_items),
        )
        intermediate_path = str(temp_output_dir / "Income_Statement_Enriched.xlsx")
        output_path = str(temp_output_dir / "Budget_Pipeline.xlsx")
        pipeline = BudgetPipeline(
            input_path=temp_input_path,
            intermediate_path=intermediate_path,
            output_path=output_path,
            growth_factor=growth_factor,
            growth_factor_note=growth_factor_note,
            enrich_only=False,
            known_columns=pdf_known_columns,
        )
        pipeline.run()
        preview = macros_service.read_first_sheet_preview(output_path, settings.MAX_PREVIEW_ROWS)
        enriched_bytes = Path(intermediate_path).read_bytes()
        return enriched_bytes, preview
    finally:
        if os.path.exists(temp_input_path):
            os.remove(temp_input_path)
        shutil.rmtree(temp_output_dir, ignore_errors=True)


def _persist_draft_enriched_workbook(
    draft: BudgetDraft,
    *,
    enriched_bytes: bytes,
) -> str:
    storage_key = draft.enriched_storage_key or _draft_enriched_storage_key(draft.property_id, draft.id)
    _write_atomic_bytes(_budget_storage_path(storage_key), enriched_bytes)
    draft.enriched_storage_key = storage_key
    return storage_key


def _refresh_draft_snapshot_from_upload(
    session: Session,
    draft: BudgetDraft,
    upload: BudgetUpload,
) -> dict[str, Any]:
    line_items = _json_loads(draft.line_items_json, [])
    enriched_bytes, preview = _render_draft_snapshot_from_upload(
        upload,
        line_items=line_items,
        growth_factor=draft.growth_factor or upload.growth_factor,
        growth_factor_note=draft.growth_factor_note or upload.growth_factor_note,
    )
    _persist_draft_enriched_workbook(draft, enriched_bytes=enriched_bytes)
    draft.budget_preview_json = _json_dumps(preview)
    session.flush()
    return preview


def _ensure_draft_enriched_workbook(
    session: Session,
    draft: BudgetDraft,
) -> Path:
    if _storage_file_available(draft.enriched_storage_key):
        return _budget_storage_path(draft.enriched_storage_key or "")

    upload = _get_upload(session, draft.source_upload_id)
    if upload is None:
        raise LookupError("Enriched draft file unavailable")

    try:
        _refresh_draft_snapshot_from_upload(session, draft, upload)
    except LookupError as exc:
        raise LookupError("Enriched draft file unavailable") from exc
    except FileNotFoundError as exc:
        raise LookupError("Enriched draft file unavailable") from exc
    except ValueError as exc:
        raise LookupError("Enriched draft file unavailable") from exc

    if not draft.enriched_storage_key:
        raise LookupError("Enriched draft file unavailable")
    return _budget_storage_path(draft.enriched_storage_key)


def create_upload(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    original_filename: str,
    content_type: Optional[str],
    file_bytes: bytes,
    source_mode: str = SOURCE_MODE_INCOME_STATEMENT,
    assessment_mode: str = ASSESSMENT_MODE_VARIABLE,
) -> BudgetUploadResponse:
    source_mode = _normalize_source_mode(source_mode)
    hoa = _get_property(session, hoa_id)
    assessment_mode = _update_property_assessment_mode(
        session,
        hoa=hoa,
        actor=actor,
        requested_assessment_mode=assessment_mode,
    )
    timestamp = _now_text()
    sha256 = hashlib.sha256(file_bytes).hexdigest()
    pdf_result: ExtractedFinancialStatement | DocumentExtractionFailure | None = None
    upload = _create_upload_record(
        session,
        hoa_id=hoa_id,
        actor=actor,
        original_filename=original_filename,
        content_type=content_type,
        file_bytes=file_bytes,
        timestamp=timestamp,
        document_role="budget_source",
        enrichment_status="failed",
        source_mode=source_mode,
        assessment_mode=assessment_mode,
    )

    upload_received_event = _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="upload_received",
        summary=f"Uploaded {original_filename}",
        upload_id=upload.id,
        payload={
            "filename": original_filename,
            "sha256": sha256,
            "source_mode": source_mode,
            "assessment_mode": assessment_mode,
        },
    )

    route = choose_financial_document_route(original_filename, content_type)
    temp_input_path = _write_temp_workbook(file_bytes, original_filename)
    cleanup_paths = {temp_input_path}
    temp_output_dir = Path(tempfile.mkdtemp(prefix="budget_history_"))
    try:
        if _is_proforma_source_mode(source_mode):
            parse_warnings: list[str] = []
            if route.path == "pdf_vlm":
                pdf_result = _extract_pdf_statement_sync(temp_input_path, source_mode=source_mode)
                if isinstance(pdf_result, DocumentExtractionFailure):
                    return _build_review_required_response(
                        session,
                        hoa_id=hoa_id,
                        actor=actor,
                        upload=upload,
                        original_filename=original_filename,
                        reason=pdf_result.message,
                        code=pdf_result.code,
                        warnings=[pdf_result.message],
                        details=pdf_result.details,
                    )
                line_items = _statement_to_budget_line_items(pdf_result, source_mode=source_mode)
                canonical_statement = pdf_result
            else:
                normalized_path = _ensure_xlsx(temp_input_path)
                cleanup_paths.add(normalized_path)
                temp_input_path = normalized_path
                excel_result = _extract_proforma_excel_statement_sync(
                    temp_input_path,
                    original_filename=original_filename,
                )
                if isinstance(excel_result, DocumentExtractionFailure):
                    return _build_review_required_response(
                        session,
                        hoa_id=hoa_id,
                        actor=actor,
                        upload=upload,
                        original_filename=original_filename,
                        reason=excel_result.message,
                        code=excel_result.code,
                        warnings=[excel_result.message],
                        details=excel_result.details,
                    )
                line_items = _statement_to_budget_line_items(excel_result, source_mode=source_mode)
                canonical_statement = excel_result

            canonical_issues = validate_extracted_statement(canonical_statement)
            if has_blocking_validation_issues(canonical_issues):
                review_reason, validation_warnings = _build_proforma_validation_feedback(
                    original_filename=original_filename,
                    issues=canonical_issues,
                )
                return _build_review_required_response(
                    session,
                    hoa_id=hoa_id,
                    actor=actor,
                    upload=upload,
                    original_filename=original_filename,
                    reason=review_reason,
                    code="validation_failed",
                    warnings=parse_warnings + validation_warnings,
                    details={"validation_issues": canonical_issues, "source_mode": source_mode},
                )

            statement_month = _default_proforma_statement_month(hoa)
            growth_factor = _PROFORMA_GROWTH_FACTOR
            growth_factor_note = _PROFORMA_GROWTH_FACTOR_NOTE

            _replace_active_draft(session, hoa_id, timestamp)
            draft = BudgetDraft(
                property_id=hoa_id,
                source_upload_id=upload.id,
                reserve_study_upload_id=None,
                reopened_from_version_id=None,
                source_mode=source_mode,
                assessment_mode=assessment_mode,
                status=BUDGET_DRAFT_ACTIVE,
                line_items_json=_json_dumps(line_items),
                reserve_study_rows_json=_json_dumps([]),
                reserve_study_warnings_json=_json_dumps([]),
                reserve_study_status="none",
                global_note=None,
                statement_month=statement_month,
                growth_factor=growth_factor,
                growth_factor_note=growth_factor_note,
                reserve_inflation_rate=app_settings_service.get_global_reserve_inflation_rate(session),
                reserve_inflation_note=None,
                budget_preview_json=_json_dumps(None),
                created_by_user_id=actor["id"],
                updated_by_user_id=actor["id"],
                actor_name=_actor_name(actor),
                created_at=timestamp,
                updated_at=timestamp,
            )
            session.add(draft)
            session.flush()

            auto_applied_merge_count = auto_apply_merges_on_upload(
                property_id=hoa_id,
                budget_draft_id=draft.id,
                new_draft_line_items=line_items,
                db_conn=_raw_sqlite_connection(session),
            )
            if auto_applied_merge_count:
                session.expire(draft, ["line_items_json", "version_int"])
                line_items = _json_loads(draft.line_items_json, line_items)

            preview = _refresh_draft_snapshot_from_upload(session, draft, upload)
            mapping_counts = _materialize_assessment_mappings_for_line_items(
                session,
                hoa_id=hoa_id,
                line_items=line_items,
            )

            upload.enrichment_status = "completed"
            upload.line_items_json = _json_dumps(line_items)
            upload.budget_preview_json = _json_dumps(preview)
            upload.statement_month = statement_month
            upload.growth_factor = growth_factor
            upload.growth_factor_note = growth_factor_note
            upload.source_mode = source_mode
            upload.assessment_mode = assessment_mode

            _create_audit_event(
                session,
                hoa_id=hoa_id,
                actor=actor,
                event_type="enrichment_completed",
                summary=f"Enriched upload for {hoa.name}",
                upload_id=upload.id,
                draft_id=draft.id,
                payload={
                    "statement_month": statement_month,
                    "growth_factor": growth_factor,
                    "source_mode": source_mode,
                    "assessment_mode": assessment_mode,
                    "assessment_mapping_counts": mapping_counts,
                    "auto_applied_merge_count": auto_applied_merge_count,
                },
            )
            session.commit()

            quality_warning: Optional[ExtractionQualityWarning] = None
            if route.path == "pdf_vlm" and isinstance(pdf_result, ExtractedFinancialStatement):
                metadata = pdf_result.extraction_metadata or {}
                if metadata.get("used_vision_only_fallback"):
                    quality_warning = ExtractionQualityWarning(
                        code="scanned_pdf_vision_only",
                        title="Please double-check the numbers below",
                        body=(
                            "This file was a scanned image, so we had to read every "
                            "number from the page picture instead of the file's text. "
                            "That's usually accurate, but not always — please review "
                            "every line item carefully before saving."
                        ),
                        severity="warning",
                    )

            return BudgetUploadResponse(
                upload_id=upload.id,
                draft=_serialize_draft(draft, upload),
                timeline_event=_serialize_timeline_event(upload_received_event),
                warnings=parse_warnings,
                extraction_quality_warning=quality_warning,
            )

        statement_period_hint: str | None = None
        if route.path == "pdf_vlm":
            pdf_source_path = temp_input_path
            pdf_result = _extract_pdf_statement_sync(temp_input_path, source_mode=source_mode)
            if isinstance(pdf_result, DocumentExtractionFailure):
                return _build_review_required_response(
                    session,
                    hoa_id=hoa_id,
                    actor=actor,
                    upload=upload,
                    original_filename=original_filename,
                    reason=pdf_result.message,
                    code=pdf_result.code,
                    warnings=[pdf_result.message],
                    details=pdf_result.details,
                )

            # Diagnostic: log a sample of what Gemini returned BEFORE the
            # XLSX round-trip. Used to debug "missing_numeric_coverage"
            # validation failures — lets us see whether numerics were lost
            # at the model boundary or during the workbook round-trip below.
            try:
                _sample = pdf_result.line_items[:5]
                logger.info(
                    "PDF extraction sample (post-Gemini, pre-roundtrip): %d items, first 5 = %s",
                    len(pdf_result.line_items),
                    [
                        {
                            "label": getattr(item, "label", None),
                            "section_kind": getattr(item, "section_kind", None),
                            "current_actual": getattr(item, "current_actual", None),
                            "current_budget": getattr(item, "current_budget", None),
                            "ytd_actual": getattr(item, "ytd_actual", None),
                            "ytd_budget": getattr(item, "ytd_budget", None),
                            "annual_budget": getattr(item, "annual_budget", None),
                        }
                        for item in _sample
                    ],
                )
            except Exception as _log_exc:
                logger.debug("Could not log PDF extraction sample: %s", _log_exc)

            statement_period_hint = pdf_result.statement_period or extract_pdf_statement_period_hint(pdf_source_path)
            normalized_path = build_normalized_statement_workbook(pdf_result)
            cleanup_paths.add(normalized_path)
            temp_input_path = normalized_path
        else:
            normalized_path = _ensure_xlsx(temp_input_path)
            cleanup_paths.add(normalized_path)
            temp_input_path = normalized_path

        growth_factor = None
        if statement_period_hint:
            inferred = infer_growth_factor_from_statement_period(
                statement_period_hint,
                hoa.fiscal_year_start_month or 1,
            )
            if inferred is not None:
                growth_factor, detected_months, source = inferred

        if growth_factor is None:
            growth_factor, detected_months, source = infer_growth_factor_from_input(
                temp_input_path,
                fiscal_year_start_month=hoa.fiscal_year_start_month or 1,
            )
        growth_factor_note = f"auto annualization 12/{detected_months} from {source}"
        statement_month = _extract_statement_month(hoa.fiscal_year_start_month or 1, detected_months)
        intermediate_path = str(temp_output_dir / "Income_Statement_Enriched.xlsx")
        output_path = str(temp_output_dir / "Budget_Pipeline.xlsx")
        # PDF extraction: pass known column positions so enricher skips detection
        pdf_known_columns = (
            {"ytd_actual": 6, "annual_budget": 9}
            if route.path == "pdf_vlm" else None
        )
        pipeline = BudgetPipeline(
            input_path=temp_input_path,
            intermediate_path=intermediate_path,
            output_path=output_path,
            growth_factor=growth_factor,
            growth_factor_note=growth_factor_note,
            enrich_only=False,
            known_columns=pdf_known_columns,
        )
        pipeline.run()
        enriched = macros_service.read_sheet_as_table(intermediate_path, "Income Statement")
        preview = macros_service.read_first_sheet_preview(output_path, settings.MAX_PREVIEW_ROWS)
        line_items, parse_warnings = _table_to_line_items(enriched)
        # Diagnostic: log a sample of items AFTER the XLSX round-trip.
        # Compare against the post-Gemini sample logged earlier to see
        # whether numerics survived BudgetPipeline column detection and
        # _table_to_line_items.
        if route.path == "pdf_vlm":
            try:
                _sample_after = line_items[:5]
                logger.info(
                    "PDF extraction sample (post-roundtrip, pre-validation): %d items, first 5 = %s",
                    len(line_items),
                    [
                        {
                            "label": item.get("label"),
                            "category": item.get("category"),
                            "ytd_actual": item.get("ytd_actual"),
                            "annual_budget": item.get("annual_budget"),
                            "projection": item.get("projection"),
                            "raw_keys": list((item.get("raw") or {}).keys())[:12],
                        }
                        for item in _sample_after
                    ],
                )
                logger.info(
                    "Enriched workbook headers (%d total) = %s",
                    len(enriched.get("headers", [])),
                    enriched.get("headers", []),
                )
            except Exception as _log_exc:
                logger.debug("Could not log post-roundtrip sample: %s", _log_exc)
        canonical_statement = _canonical_statement_from_line_items(
            line_items,
            family=route.family or ("pdf_visual_document" if route.path == "pdf_vlm" else "known_clean_excel_workbook"),
        )
        canonical_issues = validate_extracted_statement(canonical_statement)
        if has_blocking_validation_issues(canonical_issues):
            review_reason, validation_warnings = _build_income_statement_validation_feedback(
                original_filename=original_filename,
                issues=canonical_issues,
            )
            return _build_review_required_response(
                session,
                hoa_id=hoa_id,
                actor=actor,
                upload=upload,
                original_filename=original_filename,
                reason=review_reason,
                code="validation_failed",
                warnings=parse_warnings + validation_warnings,
                details={"validation_issues": canonical_issues, "source_mode": source_mode},
            )

        _replace_active_draft(session, hoa_id, timestamp)
        draft = BudgetDraft(
            property_id=hoa_id,
            source_upload_id=upload.id,
            reserve_study_upload_id=None,
            reopened_from_version_id=None,
            source_mode=source_mode,
            assessment_mode=assessment_mode,
            status=BUDGET_DRAFT_ACTIVE,
            line_items_json=_json_dumps(line_items),
            reserve_study_rows_json=_json_dumps([]),
            reserve_study_warnings_json=_json_dumps([]),
            reserve_study_status="none",
            global_note=None,
            statement_month=statement_month,
            growth_factor=growth_factor,
            growth_factor_note=growth_factor_note,
            reserve_inflation_rate=app_settings_service.get_global_reserve_inflation_rate(session),
            reserve_inflation_note=None,
            budget_preview_json=_json_dumps(preview),
            created_by_user_id=actor["id"],
            updated_by_user_id=actor["id"],
            actor_name=_actor_name(actor),
            created_at=timestamp,
            updated_at=timestamp,
        )
        session.add(draft)
        session.flush()
        auto_applied_merge_count = auto_apply_merges_on_upload(
            property_id=hoa_id,
            budget_draft_id=draft.id,
            new_draft_line_items=line_items,
            db_conn=_raw_sqlite_connection(session),
        )
        if auto_applied_merge_count:
            session.expire(draft, ["line_items_json", "version_int"])
            line_items = _json_loads(draft.line_items_json, line_items)
            preview = _refresh_draft_snapshot_from_upload(session, draft, upload)
        else:
            _persist_draft_enriched_workbook(
                draft,
                enriched_bytes=Path(intermediate_path).read_bytes(),
            )
        mapping_counts = _materialize_assessment_mappings_for_line_items(
            session,
            hoa_id=hoa_id,
            line_items=line_items,
        )

        upload.enrichment_status = "completed"
        upload.line_items_json = _json_dumps(line_items)
        upload.budget_preview_json = _json_dumps(preview)
        upload.statement_month = statement_month
        upload.growth_factor = growth_factor
        upload.growth_factor_note = growth_factor_note
        upload.source_mode = source_mode
        upload.assessment_mode = assessment_mode

        _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="enrichment_completed",
            summary=f"Enriched upload for {hoa.name}",
            upload_id=upload.id,
            draft_id=draft.id,
            payload={
                "statement_month": statement_month,
                "growth_factor": growth_factor,
                "source_mode": source_mode,
                "assessment_mode": assessment_mode,
                "assessment_mapping_counts": mapping_counts,
                "auto_applied_merge_count": auto_applied_merge_count,
            },
        )
        session.commit()
        # If the extractor took a degraded path (e.g. scanned-PDF vision-only
        # fallback), surface a one-shot quality warning so the frontend can
        # show a dismissible dialog. The text is written for non-technical
        # users and explains why they should double-check the numbers.
        quality_warning: Optional[ExtractionQualityWarning] = None
        if route.path == "pdf_vlm" and isinstance(pdf_result, ExtractedFinancialStatement):
            metadata = pdf_result.extraction_metadata or {}
            if metadata.get("used_vision_only_fallback"):
                quality_warning = ExtractionQualityWarning(
                    code="scanned_pdf_vision_only",
                    title="Please double-check the numbers below",
                    body=(
                        "This file was a scanned image, so we had to read every "
                        "number from the page picture instead of the file's text. "
                        "That's usually accurate, but not always — please review "
                        "every line item carefully (especially actuals and budgets) "
                        "before saving. If you find mistakes, try re-uploading the "
                        "original Excel file instead."
                    ),
                    severity="warning",
                )
        return BudgetUploadResponse(
            upload_id=upload.id,
            draft=_serialize_draft(draft, upload),
            timeline_event=_serialize_timeline_event(upload_received_event),
            warnings=parse_warnings,
            extraction_quality_warning=quality_warning,
        )
    except Exception as exc:
        upload.enrichment_status = "failed"
        _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="enrichment_failed",
            summary=f"Enrichment failed for {original_filename}",
            upload_id=upload.id,
            payload={"error": str(exc)},
        )
        session.commit()
        raise
    finally:
        for cleanup_path in cleanup_paths:
            if cleanup_path and os.path.exists(cleanup_path):
                os.remove(cleanup_path)
        shutil.rmtree(temp_output_dir, ignore_errors=True)


def _persist_reserve_study_to_draft(
    session: Session,
    *,
    hoa_id: int,
    draft_id: Optional[int],
    actor: dict[str, Any],
    reserve_filename: str,
    reserve_content_type: Optional[str],
    reserve_file_bytes: bytes,
) -> tuple[Optional["BudgetUpload"], BundleFileStatus]:
    """Upload, extract, and persist a reserve study PDF onto a draft.

    Returns (upload_record, status). On hard extraction failure, prior reserve
    rows on the draft are preserved — only the upload_id and warnings are
    updated. Pass draft_id=None to record the upload without touching any draft.
    """
    reserve_route = choose_financial_document_route(reserve_filename, reserve_content_type)
    reserve_is_pdf = (
        Path(reserve_filename).suffix.lower() == ".pdf"
        or "pdf" in (reserve_content_type or "").lower()
    )
    if not (reserve_route.is_supported and reserve_is_pdf):
        unsupported_reason = "Unsupported reserve study file type. Upload a reserve study PDF."
        return None, BundleFileStatus(
            filename=reserve_filename,
            status="failed",
            review_reason=unsupported_reason,
            debug_info=ExtractionDebugInfo(
                code="unsupported_file_type",
                message=unsupported_reason,
                details={"content_type": reserve_content_type},
            ),
        )

    timestamp = _now_text()
    reserve_upload = _create_upload_record(
        session,
        hoa_id=hoa_id,
        actor=actor,
        original_filename=reserve_filename,
        content_type=reserve_content_type,
        file_bytes=reserve_file_bytes,
        timestamp=timestamp,
        document_role="reserve_study",
        enrichment_status="completed",
    )

    try:
        reserve_result = _extract_reserve_study_sync(
            _budget_storage_path(reserve_upload.storage_key).as_posix()
        )
    except Exception as exc:
        reserve_result = DocumentExtractionFailure(
            code="reserve_provider_error",
            message=f"Reserve study extraction could not complete automatically: {exc}",
            details={"error": str(exc)},
        )

    if isinstance(reserve_result, DocumentExtractionFailure):
        reserve_status = BundleFileStatus(
            upload_id=reserve_upload.id,
            filename=reserve_filename,
            status="review_required",
            warnings=[reserve_result.message],
            review_reason=reserve_result.message,
            debug_info=ExtractionDebugInfo(
                code=reserve_result.code,
                message=reserve_result.message,
                details=reserve_result.details,
            ),
        )
        if draft_id is not None:
            draft_row = _get_editable_draft(session, hoa_id, draft_id)
            draft_row.reserve_study_upload_id = reserve_upload.id
            draft_row.reserve_study_status = "review_required"
            # Preserve prior rows on hard failure — do not wipe them.
            draft_row.reserve_study_warnings_json = _json_dumps([reserve_result.message])
            draft_row.updated_by_user_id = actor["id"]
            draft_row.actor_name = _actor_name(actor)
            draft_row.updated_at = _now_text()
            session.commit()
    else:
        has_review_flags = bool(reserve_result.warnings) or any(
            row.flags for row in reserve_result.rows
        )
        persisted_status = "review_required" if has_review_flags else "completed"
        reserve_status = BundleFileStatus(
            upload_id=reserve_upload.id,
            filename=reserve_filename,
            status=persisted_status,
            warnings=reserve_result.warnings,
            review_reason=(
                "Reserve study rows need review before applying to the budget."
                if has_review_flags
                else None
            ),
        )
        if draft_id is not None:
            draft_row = _get_editable_draft(session, hoa_id, draft_id)
            draft_row.reserve_study_upload_id = reserve_upload.id
            draft_row.reserve_study_status = persisted_status
            draft_row.reserve_study_rows_json = _json_dumps(
                [row.model_dump() for row in reserve_result.rows]
            )
            draft_row.reserve_study_warnings_json = _json_dumps(reserve_result.warnings)
            draft_row.updated_by_user_id = actor["id"]
            draft_row.actor_name = _actor_name(actor)
            draft_row.updated_at = _now_text()
            extracted_date = getattr(reserve_result, "study_date", None)
            if extracted_date:
                from ..services import hoa_settings_service as _hoa_settings_service
                settings_row = _hoa_settings_service.get_or_create(session, hoa_id=hoa_id)
                settings_row.reserve_study_date = str(extracted_date)
            session.commit()

    return reserve_upload, reserve_status


def replace_reserve_study(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
    actor: dict[str, Any],
    reserve_filename: str,
    reserve_content_type: Optional[str],
    reserve_file_bytes: bytes,
) -> BudgetDraftPayload:
    """Replace the reserve study on an existing draft with a new PDF.

    The new study is extracted and persisted, but NOT automatically applied
    to budget line items. The operator applies separately via the existing
    apply endpoint. On hard extraction failure prior reserve rows are preserved.
    """
    _get_property(session, hoa_id)
    draft_row = _get_editable_draft(session, hoa_id, draft_id)
    _, _ = _persist_reserve_study_to_draft(
        session,
        hoa_id=hoa_id,
        draft_id=draft_id,
        actor=actor,
        reserve_filename=reserve_filename,
        reserve_content_type=reserve_content_type,
        reserve_file_bytes=reserve_file_bytes,
    )
    # Reload after commit inside helper
    session.refresh(draft_row)
    return _serialize_draft(draft_row, _get_upload(session, draft_row.source_upload_id))


def create_upload_bundle(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    budget_filename: str,
    budget_content_type: Optional[str],
    budget_file_bytes: bytes,
    reserve_filename: str,
    reserve_content_type: Optional[str],
    reserve_file_bytes: bytes,
    source_mode: str = SOURCE_MODE_INCOME_STATEMENT,
    assessment_mode: str = ASSESSMENT_MODE_VARIABLE,
) -> BudgetBundleUploadResponse:
    source_mode = _normalize_source_mode(source_mode)
    hoa = _get_property(session, hoa_id)
    assessment_mode = _update_property_assessment_mode(
        session,
        hoa=hoa,
        actor=actor,
        requested_assessment_mode=assessment_mode,
    )

    budget_route = choose_financial_document_route(budget_filename, budget_content_type)
    if budget_route.is_supported:
        budget_response = create_upload(
            session,
            hoa_id=hoa_id,
            actor=actor,
            original_filename=budget_filename,
            content_type=budget_content_type,
            file_bytes=budget_file_bytes,
            source_mode=source_mode,
            assessment_mode=assessment_mode,
        )
        draft = budget_response.draft
        budget_status = _bundle_status_from_budget_response(budget_response, filename=budget_filename)
    else:
        unsupported_reason = (
            "Unsupported budget file type. Upload an Excel workbook or PDF pro forma / final budget."
            if _is_proforma_source_mode(source_mode)
            else "Unsupported budget file type. Upload an Excel workbook or PDF income statement."
        )
        draft = None
        budget_status = BundleFileStatus(
            filename=budget_filename,
            status="failed",
            review_reason=unsupported_reason,
            debug_info=ExtractionDebugInfo(
                code="unsupported_file_type",
                message=unsupported_reason,
                details={"content_type": budget_content_type, "source_mode": source_mode},
            ),
        )

    reserve_upload, reserve_status = _persist_reserve_study_to_draft(
        session,
        hoa_id=hoa_id,
        draft_id=draft.id if draft is not None else None,
        actor=actor,
        reserve_filename=reserve_filename,
        reserve_content_type=reserve_content_type,
        reserve_file_bytes=reserve_file_bytes,
    )
    if draft is not None and reserve_upload is not None:
        draft_row = _get_editable_draft(session, hoa_id, draft.id)
        draft = _serialize_draft(draft_row, _get_upload(session, draft_row.source_upload_id))

    return BudgetBundleUploadResponse(
        draft=draft,
        budget_source=budget_status,
        reserve_study=reserve_status,
        can_continue_with_budget_only=draft is not None and reserve_status.status == "failed",
        can_continue_with_reserve_study_only=draft is None and reserve_upload is not None and budget_status.status == "failed",
    )


def get_active_draft(session: Session, hoa_id: int) -> BudgetDraftPayload:
    draft = session.scalars(
        select(BudgetDraft).where(
            BudgetDraft.property_id == hoa_id,
            BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
        ).order_by(BudgetDraft.updated_at.desc())
    ).first()
    if draft is None:
        raise LookupError("Active draft not found")
    return _serialize_draft(draft, _get_upload(session, draft.source_upload_id))


def get_requested_draft(session: Session, hoa_id: int, draft_id: int) -> BudgetDraftPayload:
    draft = _get_editable_draft(session, hoa_id, draft_id)
    return _serialize_draft(draft, _get_upload(session, draft.source_upload_id))


def get_draft_compare_options(
    session: Session,
    *,
    hoa_id: int,
) -> BudgetDraftCompareOptionsResponse:
    _get_property(session, hoa_id)
    draft = _get_editable_draft(
        session,
        hoa_id,
        session.scalars(
            select(BudgetDraft.id).where(
                BudgetDraft.property_id == hoa_id,
                BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
            ).order_by(BudgetDraft.updated_at.desc(), BudgetDraft.id.desc())
        ).first() or 0,
    )
    versions = session.scalars(
        select(BudgetVersion)
        .where(BudgetVersion.property_id == hoa_id)
        .order_by(BudgetVersion.created_at.desc(), BudgetVersion.id.desc())
    ).all()

    recommended_version = next(
        (version for version in versions if version.stage == BUDGET_VERSION_STAGE_FINAL),
        None,
    )
    recommended_reason = "Latest Final" if recommended_version is not None else None
    if recommended_version is None and versions:
        recommended_version = versions[0]
        recommended_reason = "Latest generated (no Final version yet)"

    recommended_version_id = recommended_version.id if recommended_version else None
    return BudgetDraftCompareOptionsResponse(
        draft_id=draft.id,
        recommended_baseline_version_id=recommended_version_id,
        recommended_baseline_reason=recommended_reason,
        baseline_versions=[
            BudgetDraftCompareBaselineOption(
                id=version.id,
                version_code=version.version_code,
                stage=version.stage,
                label=version.label,
                created_at=version.created_at,
                summary_note=version.summary_note,
                is_recommended=version.id == recommended_version_id,
            )
            for version in versions
        ],
        reserve_inflation_rate=app_settings_service.get_global_reserve_inflation_rate(session),
        reserve_inflation_note=None,
    )


def save_draft(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    payload: BudgetDraftSaveRequest,
) -> tuple[BudgetDraftPayload, Optional[BudgetTimelineEvent]]:
    _get_property(session, hoa_id)
    draft = _get_editable_draft(session, hoa_id, payload.draft_id)
    upload = _get_upload(session, draft.source_upload_id)
    if upload is None:
        raise LookupError("Source upload not found")
    existing_snapshot = {
        "line_items": _json_loads(draft.line_items_json, []),
        "global_note": draft.global_note,
        "statement_month": draft.statement_month,
        "growth_factor": draft.growth_factor,
        "growth_factor_note": draft.growth_factor_note,
    }
    incoming_snapshot = {
        "line_items": payload.line_items,
        "global_note": payload.global_note,
        "statement_month": payload.statement_month,
        "growth_factor": payload.growth_factor,
        "growth_factor_note": payload.growth_factor_note,
    }

    # Recompute read_only from read_only_override before persisting, so the
    # stored flag always reflects the current override state.
    normalized_items = [
        {**it, "read_only": _effective_read_only(it, str(it.get("category", "")))}
        if isinstance(it, dict) else it
        for it in payload.line_items
    ]
    draft.line_items_json = _json_dumps(normalized_items)
    draft.global_note = payload.global_note
    draft.statement_month = payload.statement_month
    draft.growth_factor = payload.growth_factor
    draft.growth_factor_note = payload.growth_factor_note
    # Resolve the reserve inflation rate in priority order:
    #   1. operator-supplied payload value (explicit override on save)
    #   2. HOA-level ``properties.reserve_inflation_rate`` when non-zero
    #   3. app-level global default
    # Keeps the HOA-level rate as the implicit save-time choice without
    # forcing the operator to re-type it on every PATCH.
    _hoa_rate = getattr(_get_property(session, hoa_id), "reserve_inflation_rate", None)
    if payload.reserve_inflation_rate is not None:
        draft.reserve_inflation_rate = payload.reserve_inflation_rate
    elif _hoa_rate is not None and _hoa_rate > 0.0:
        draft.reserve_inflation_rate = _hoa_rate
    else:
        draft.reserve_inflation_rate = (
            app_settings_service.get_global_reserve_inflation_rate(session)
        )
    draft.reserve_inflation_note = payload.reserve_inflation_note
    draft.updated_by_user_id = actor["id"]
    draft.actor_name = _actor_name(actor)
    draft.updated_at = _now_text()
    _refresh_draft_snapshot_from_upload(session, draft, upload)
    _materialize_assessment_mappings_for_line_items(
        session,
        hoa_id=hoa_id,
        line_items=_json_loads(draft.line_items_json, payload.line_items),
    )
    session.flush()

    timeline_event = None
    if existing_snapshot != incoming_snapshot:
        event = _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="manual_overrides_saved",
            summary="Saved meaningful draft overrides",
            upload_id=draft.source_upload_id,
            draft_id=draft.id,
            payload={"line_item_count": len(payload.line_items)},
        )
        timeline_event = _serialize_timeline_event(event)
    session.commit()
    return _serialize_draft(draft, upload), timeline_event


def save_reserve_study_rows(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
    actor: dict[str, Any],
    payload: BudgetReserveStudySaveRequest,
) -> BudgetDraftPayload:
    draft = _get_editable_draft(session, hoa_id, draft_id)
    upload = _get_upload(session, draft.source_upload_id)
    normalized_rows, _ = canonicalize_reserve_study_row_dicts(payload.rows)
    has_review_flags = bool(payload.warnings) or any(
        isinstance(row.get("flags"), list) and len(row.get("flags") or []) > 0
        for row in normalized_rows
    )
    draft.reserve_study_rows_json = _json_dumps(normalized_rows)
    draft.reserve_study_warnings_json = _json_dumps(payload.warnings)
    draft.reserve_study_status = "review_required" if has_review_flags else "completed"
    draft.updated_by_user_id = actor["id"]
    draft.actor_name = _actor_name(actor)
    draft.updated_at = _now_text()
    session.commit()
    return _serialize_draft(draft, upload)


def apply_reserve_study_to_budget(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
    actor: dict[str, Any],
) -> BudgetReserveStudyApplyResponse:
    draft = _get_editable_draft(session, hoa_id, draft_id)
    upload = _get_upload(session, draft.source_upload_id)
    reserve_rows, _ = canonicalize_reserve_study_row_dicts(_json_loads(draft.reserve_study_rows_json, []))
    due_rows = [
        row for row in reserve_rows
        if isinstance(row, dict) and _reserve_study_row_due_this_budget_year(row)
    ]
    if not due_rows:
        return BudgetReserveStudyApplyResponse(
            draft=_serialize_draft(draft, upload),
            applied_count=0,
            message="No reserve study components are due this budget year.",
        )

    current_line_items = _json_loads(draft.line_items_json, [])
    preserved_line_items = [
        item for item in current_line_items
        if isinstance(item, dict) and not _is_applied_reserve_study_line_item(item)
    ]
    applied_line_items = [_reserve_study_row_to_budget_line_item(row) for row in due_rows]
    draft.line_items_json = _json_dumps([*preserved_line_items, *applied_line_items])
    _materialize_assessment_mappings_for_line_items(
        session,
        hoa_id=hoa_id,
        line_items=[*preserved_line_items, *applied_line_items],
    )
    draft.updated_by_user_id = actor["id"]
    draft.actor_name = _actor_name(actor)
    draft.updated_at = _now_text()
    session.commit()
    return BudgetReserveStudyApplyResponse(
        draft=_serialize_draft(draft, upload),
        applied_count=len(applied_line_items),
        message=(
            "Applied reserve study rows to the budget."
            if applied_line_items
            else "No reserve study components are due this budget year."
        ),
    )


def _latest_line_item_notes(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
) -> dict[str, BudgetNote]:
    notes = session.scalars(
        select(BudgetNote)
        .where(
            BudgetNote.property_id == hoa_id,
            BudgetNote.draft_id == draft_id,
            BudgetNote.line_item_key.is_not(None),
        )
        .order_by(BudgetNote.created_at.desc(), BudgetNote.id.desc())
    ).all()
    latest_by_key: dict[str, BudgetNote] = {}
    for note in notes:
        normalized_key = _normalize_line_item_key(note.line_item_key)
        if normalized_key and normalized_key not in latest_by_key:
            latest_by_key[normalized_key] = note
    return latest_by_key


def _build_draft_compare_summary(rows: list[dict[str, Any]]) -> BudgetDraftChangeSummary:
    baseline_amount = round(sum(_parse_float(row.get("baseline_amount")) for row in rows), 2)
    current_amount = round(sum(_parse_float(row.get("current_amount")) for row in rows), 2)
    delta_amount = round(sum(_parse_float(row.get("delta_amount")) for row in rows), 2)
    delta_percent = None if baseline_amount == 0 else (delta_amount / baseline_amount) * 100
    return BudgetDraftChangeSummary(
        baseline_amount=baseline_amount,
        current_amount=current_amount,
        delta_amount=delta_amount,
        delta_percent=delta_percent,
    )


def _build_reserve_review_summary(rows: list[dict[str, Any]]) -> BudgetReserveReviewSummary:
    baseline_amount = round(sum(_parse_float(row.get("baseline_amount")) for row in rows), 2)
    inflation_adjusted_amount = round(
        sum(_parse_float(row.get("inflation_adjusted_baseline_amount")) for row in rows),
        2,
    )
    impact_amount = round(inflation_adjusted_amount - baseline_amount, 2)
    return BudgetReserveReviewSummary(
        baseline_amount=baseline_amount,
        inflation_adjusted_amount=inflation_adjusted_amount,
        impact_amount=impact_amount,
        eligible_component_count=len(rows),
    )


def _build_draft_reserve_review_rows(
    line_items: list[dict[str, Any]],
    *,
    reserve_inflation_rate: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in line_items:
        if _reserve_group_for_item(item) != "component":
            continue
        baseline_amount = _line_item_amount(item)
        inflation_adjusted_amount = (
            baseline_amount * (1 + reserve_inflation_rate)
            if reserve_inflation_rate > 0.0
            else baseline_amount
        )
        rows.append(
            {
                "line_item_key": _normalize_line_item_key(item.get("line_item_key"))
                or _line_item_account_code_text(item)
                or _line_item_label(item),
                "label": _line_item_label(item),
                "baseline_amount": round(baseline_amount, 2),
                "inflation_adjusted_baseline_amount": round(inflation_adjusted_amount, 2),
            }
        )

    rows.sort(
        key=lambda row: (
            -(
                _parse_float(row.get("inflation_adjusted_baseline_amount"))
                - _parse_float(row.get("baseline_amount"))
            ),
            _normalize_compare_text(row.get("label")),
        )
    )
    return rows


def compare_draft_to_version(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
    baseline_version_id: int,
    changed_only: bool,
) -> BudgetDraftCompareResponse:
    draft = _get_editable_draft(session, hoa_id, draft_id)
    baseline_version = _get_version(session, hoa_id, baseline_version_id)
    if baseline_version_id == draft.id:
        raise ValueError("Baseline version must be distinct from the active draft")

    draft_line_items = _json_loads(draft.line_items_json, [])
    baseline_line_items = _json_loads(baseline_version.line_items_json, [])
    baseline_index = _build_line_item_index(baseline_line_items)
    draft_index = _build_line_item_index(draft_line_items)
    note_by_key = _latest_line_item_notes(session, hoa_id=hoa_id, draft_id=draft.id)

    ordered_pairs: list[tuple[Optional[dict[str, Any]], Optional[dict[str, Any]]]] = []
    seen_baseline_ids: set[int] = set()

    for draft_item in draft_line_items:
        baseline_item = _find_matching_line_item(draft_item, index=baseline_index)
        if baseline_item is not None:
            seen_baseline_ids.add(id(baseline_item))
        ordered_pairs.append((draft_item, baseline_item))

    for baseline_item in baseline_line_items:
        if id(baseline_item) in seen_baseline_ids:
            continue
        ordered_pairs.append((_find_matching_line_item(baseline_item, index=draft_index), baseline_item))

    raw_rows: list[dict[str, Any]] = []
    for draft_item, baseline_item in ordered_pairs:
        item = draft_item or baseline_item
        if item is None:
            continue
        label = _line_item_label(item)
        baseline_amount = _line_item_amount(baseline_item)
        current_amount = _line_item_amount(draft_item)
        delta_amount = current_amount - baseline_amount
        line_item_key = (
            _normalize_line_item_key(item.get("line_item_key"))
            or _line_item_account_code_text(item)
            or label
        )
        note = note_by_key.get(line_item_key)
        raw_rows.append(
            {
                "line_item_key": line_item_key,
                "label": label,
                "category": _line_item_category(item),
                "reserve_group": _reserve_group_for_item(item),
                "is_reserve": _is_reserve_item(item),
                "reserve_inflation_eligible": False,
                "baseline_amount": baseline_amount,
                "inflation_adjusted_baseline_amount": baseline_amount,
                "current_amount": current_amount,
                "delta_amount": delta_amount,
                "delta_percent": None if baseline_amount == 0 else (delta_amount / baseline_amount) * 100,
                "changed": abs(delta_amount) > 1e-9,
                "note_title": note.title if note else None,
                "note_body": note.body if note else None,
            }
        )

    draft_compare_all_rows = [row for row in raw_rows if not row.get("is_reserve")]
    draft_compare_rows = (
        [row for row in draft_compare_all_rows if row["changed"]]
        if changed_only
        else draft_compare_all_rows
    )

    return BudgetDraftCompareResponse(
        draft_id=draft.id,
        baseline_version_id=baseline_version.id,
        changed_only=changed_only,
        draft_compare_summary=_build_draft_compare_summary(draft_compare_all_rows),
        draft_compare_rows=[BudgetDraftCompareRow(**row) for row in draft_compare_rows],
        reserve_review_summary=_build_reserve_review_summary([]),
        reserve_component_rows=[],
    )


def review_draft_reserves(
    session: Session,
    *,
    hoa_id: int,
    draft_id: int,
    payload: BudgetDraftReserveReviewRequest,
) -> BudgetDraftReserveReviewResponse:
    draft = _get_editable_draft(session, hoa_id, draft_id)
    working_line_items = payload.line_items or _json_loads(draft.line_items_json, [])
    reserve_inflation_rate = _parse_float(payload.reserve_inflation_rate)
    reserve_component_rows = _build_draft_reserve_review_rows(
        working_line_items,
        reserve_inflation_rate=reserve_inflation_rate,
    )
    return BudgetDraftReserveReviewResponse(
        draft_id=draft.id,
        reserve_inflation_rate=reserve_inflation_rate,
        reserve_review_summary=_build_reserve_review_summary(reserve_component_rows),
        reserve_component_rows=[
            BudgetReserveComponentRow(
                line_item_key=str(row["line_item_key"]),
                label=str(row["label"]),
                baseline_amount=round(_parse_float(row.get("baseline_amount")), 2),
                inflation_adjusted_amount=round(
                    _parse_float(row.get("inflation_adjusted_baseline_amount")),
                    2,
                ),
                impact_amount=round(
                    _parse_float(row.get("inflation_adjusted_baseline_amount"))
                    - _parse_float(row.get("baseline_amount")),
                    2,
                ),
            )
            for row in reserve_component_rows
        ],
    )


def save_note(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    payload: BudgetNoteSaveRequest,
) -> BudgetNoteSaveResponse:
    draft = _get_editable_draft(session, hoa_id, payload.draft_id)
    version = _get_version(session, hoa_id, payload.version_id) if payload.version_id else None
    note = BudgetNote(
        property_id=hoa_id,
        upload_id=draft.source_upload_id,
        draft_id=draft.id,
        version_id=version.id if version else None,
        note_scope=payload.note_scope,
        line_item_key=payload.line_item_key,
        title=payload.title,
        body=payload.body,
        created_by_user_id=actor["id"],
        created_by_name=_actor_name(actor),
        actor_name=_actor_name(actor),
        created_at=_now_text(),
    )
    session.add(note)
    session.flush()
    event = _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="note_saved",
        summary=f"Saved note '{payload.title}'",
        upload_id=note.upload_id,
        draft_id=note.draft_id,
        version_id=note.version_id,
        note_id=note.id,
        payload={"note_scope": payload.note_scope, "line_item_key": payload.line_item_key},
    )
    session.commit()
    return BudgetNoteSaveResponse(
        note=_serialize_note(note),
        timeline_event=_serialize_timeline_event(event),
    )


def create_budget_version(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    payload: BudgetGenerateRequest,
) -> BudgetGenerateResponse:
    hoa = _get_property(session, hoa_id)
    draft = _get_editable_draft(session, hoa_id, payload.draft_id)
    upload = _get_upload(session, draft.source_upload_id)
    if upload is None:
        raise LookupError("Source upload not found")

    timestamp = _now_text()
    line_items = payload.line_items or _json_loads(draft.line_items_json, [])
    global_note = payload.global_note if payload.global_note is not None else draft.global_note
    mapping_counts = _materialize_assessment_mappings_for_line_items(
        session,
        hoa_id=hoa_id,
        line_items=line_items,
    )

    route = choose_financial_document_route(upload.original_filename, upload.content_type)
    if route.path == "pdf_vlm" or _is_proforma_source_mode(upload.source_mode):
        source_path = _ensure_draft_enriched_workbook(session, draft)
        temp_input_path = _write_temp_workbook(source_path.read_bytes(), "enriched-income-statement.xlsx")
    else:
        source_path = _budget_storage_path(upload.storage_key)
        temp_input_path = _write_temp_workbook(source_path.read_bytes(), upload.original_filename)
    temp_output_dir = Path(tempfile.mkdtemp(prefix="budget_version_"))
    try:
        temp_input_path = _ensure_xlsx(temp_input_path)

        pdf_known_columns = (
            {"ytd_actual": 6, "annual_budget": 9}
            if route.path == "pdf_vlm" or _is_proforma_source_mode(upload.source_mode)
            else None
        )
        macros_service.write_percent_changes_by_label(
            temp_input_path,
            "Income Statement",
            _line_items_to_percent_changes(line_items),
        )
        intermediate_path = str(temp_output_dir / "Income_Statement_Enriched.xlsx")
        output_path = str(temp_output_dir / "Budget_Pipeline.xlsx")
        pipeline = BudgetPipeline(
            input_path=temp_input_path,
            intermediate_path=intermediate_path,
            output_path=output_path,
            growth_factor=draft.growth_factor or upload.growth_factor,
            growth_factor_note=draft.growth_factor_note or upload.growth_factor_note,
            enrich_only=False,
            hoa_name=hoa.name or '',
            known_columns=pdf_known_columns,
        )
        pipeline.run()
        enriched_bytes = Path(intermediate_path).read_bytes()
        preview = macros_service.read_first_sheet_preview(output_path, settings.MAX_PREVIEW_ROWS)
        total_income, total_expense, net_operating_income = _extract_totals_from_preview(preview)

        next_version_number = (
            session.scalar(
                select(func.coalesce(func.max(BudgetVersion.version_number), 0)).where(
                    BudgetVersion.property_id == hoa_id
                )
            )
            or 0
        ) + 1
        version = BudgetVersion(
            property_id=hoa_id,
            source_upload_id=upload.id,
            source_draft_id=draft.id,
            reopened_from_version_id=draft.reopened_from_version_id,
            source_mode=str(draft.source_mode or upload.source_mode or SOURCE_MODE_INCOME_STATEMENT),
            assessment_mode=normalize_assessment_mode(
                getattr(draft, "assessment_mode", None)
                or getattr(upload, "assessment_mode", None)
                or ASSESSMENT_MODE_VARIABLE
            ),
            storage_key=None,
            output_storage_key=None,
            version_number=next_version_number,
            version_code=f"V{next_version_number}",
            stage=BUDGET_VERSION_STAGE_INTERIM,
            label=None,
            summary_note=global_note,
            line_items_json=_json_dumps(line_items),
            budget_preview_json=_json_dumps(preview),
            total_income=total_income,
            total_expense=total_expense,
            net_operating_income=net_operating_income,
            growth_factor=draft.growth_factor or upload.growth_factor,
            growth_factor_note=draft.growth_factor_note or upload.growth_factor_note,
            reserve_inflation_rate=draft.reserve_inflation_rate or 0.0,
            reserve_inflation_note=draft.reserve_inflation_note,
            statement_month=draft.statement_month or upload.statement_month,
            fiscal_year_start_month=hoa.fiscal_year_start_month or 1,
            fiscal_year_end_month=hoa.fiscal_year_end_month or 12,
            created_by_user_id=actor["id"],
            created_by_name=_actor_name(actor),
            actor_name=_actor_name(actor),
            created_at=timestamp,
        )
        session.add(version)
        session.flush()

        output_storage_key = _relative_storage_path("hoa", hoa_id, "versions", version.id, "budget.xlsx")
        _write_atomic_bytes(_budget_storage_path(output_storage_key), Path(output_path).read_bytes())
        version.output_storage_key = output_storage_key

        draft.line_items_json = _json_dumps(line_items)
        draft.global_note = global_note
        draft.status = BUDGET_DRAFT_GENERATED
        _persist_draft_enriched_workbook(draft, enriched_bytes=enriched_bytes)
        draft.updated_by_user_id = actor["id"]
        draft.updated_at = timestamp
        draft.budget_preview_json = _json_dumps(preview)
        session.flush()

        event = _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="budget_generated",
            summary=f"Generated {version.version_code}",
            upload_id=upload.id,
            draft_id=draft.id,
        version_id=version.id,
        payload={
            "stage": version.stage,
            "version_code": version.version_code,
            "source_mode": str(version.source_mode or upload.source_mode or SOURCE_MODE_INCOME_STATEMENT),
            "assessment_mode": normalize_assessment_mode(
                getattr(version, "assessment_mode", None)
                or getattr(upload, "assessment_mode", None)
                or ASSESSMENT_MODE_VARIABLE
            ),
            "assessment_mapping_counts": mapping_counts,
        },
    )
        session.commit()
        return BudgetGenerateResponse(
            draft=_serialize_draft(draft, upload),
            version=_serialize_version_detail(version, upload),
            timeline_event=_serialize_timeline_event(event),
        )
    finally:
        if os.path.exists(temp_input_path):
            os.remove(temp_input_path)
        shutil.rmtree(temp_output_dir, ignore_errors=True)


def get_history(session: Session, hoa_id: int) -> BudgetHistoryResponse:
    _get_property(session, hoa_id)
    active_draft = session.scalars(
        select(BudgetDraft).where(
            BudgetDraft.property_id == hoa_id,
            BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
        ).order_by(BudgetDraft.updated_at.desc())
    ).first()
    drafts = session.scalars(
        select(BudgetDraft).where(BudgetDraft.property_id == hoa_id).order_by(BudgetDraft.updated_at.desc(), BudgetDraft.id.desc())
    ).all()
    notes = session.scalars(
        select(BudgetNote).where(BudgetNote.property_id == hoa_id).order_by(BudgetNote.created_at.desc(), BudgetNote.id.desc())
    ).all()
    versions = session.scalars(
        select(BudgetVersion).where(BudgetVersion.property_id == hoa_id).order_by(BudgetVersion.created_at.desc(), BudgetVersion.id.desc())
    ).all()
    timeline = session.scalars(
        select(BudgetAuditEvent).where(BudgetAuditEvent.property_id == hoa_id).order_by(BudgetAuditEvent.created_at.desc(), BudgetAuditEvent.id.desc())
    ).all()
    upload = _get_upload(session, active_draft.source_upload_id) if active_draft else None
    return BudgetHistoryResponse(
        active_draft=_serialize_draft(active_draft, upload) if active_draft else None,
        drafts=[
            _serialize_draft_summary(
                draft,
                upload=_get_upload(session, draft.source_upload_id),
                reopened_from_version=session.get(BudgetVersion, draft.reopened_from_version_id)
                if draft.reopened_from_version_id
                else None,
            )
            for draft in drafts
        ],
        timeline=[_serialize_timeline_event(event) for event in timeline],
        versions=[
            _serialize_version_summary(version, _get_upload(session, version.source_upload_id))
            for version in versions
        ],
        notes=[_serialize_note(note) for note in notes],
    )


def get_version_detail(session: Session, hoa_id: int, version_id: int) -> BudgetVersionDetail:
    version = _get_version(session, hoa_id, version_id)
    return _serialize_version_detail(version, _get_upload(session, version.source_upload_id))


def compare_versions(
    session: Session,
    *,
    hoa_id: int,
    left_version_id: int,
    right_version_id: int,
) -> BudgetVersionCompareResponse:
    if left_version_id == right_version_id:
        raise ValueError("Compare requires two distinct version ids")

    left_version = session.get(BudgetVersion, left_version_id)
    right_version = session.get(BudgetVersion, right_version_id)
    if left_version is None or right_version is None:
        raise LookupError("Version not found")
    if left_version.property_id != hoa_id or right_version.property_id != hoa_id:
        raise ValueError("Compared versions must belong to the selected HOA")

    return BudgetVersionCompareResponse(
        versions=[
            _serialize_version_compare_card(left_version, _get_upload(session, left_version.source_upload_id)),
            _serialize_version_compare_card(right_version, _get_upload(session, right_version.source_upload_id)),
        ]
    )


def reopen_version_as_draft(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    version_id: int,
) -> BudgetVersionReopenResponse:
    hoa = _get_property(session, hoa_id)
    version = _get_version(session, hoa_id, version_id)
    upload = _get_upload(session, version.source_upload_id)
    if upload is None:
        raise LookupError("Source upload not found")
    timestamp = _now_text()
    previous_active_draft = session.scalars(
        select(BudgetDraft).where(
            BudgetDraft.property_id == hoa_id,
            BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
        ).order_by(BudgetDraft.updated_at.desc())
    ).first()
    previous_active_draft_id = previous_active_draft.id if previous_active_draft else None
    _replace_active_draft(session, hoa_id, timestamp)
    source_draft = session.get(BudgetDraft, version.source_draft_id) if version.source_draft_id else None

    draft = BudgetDraft(
        property_id=hoa_id,
        source_upload_id=version.source_upload_id,
        reserve_study_upload_id=source_draft.reserve_study_upload_id if source_draft else None,
        reopened_from_version_id=version.id,
        source_mode=str(version.source_mode or upload.source_mode or SOURCE_MODE_INCOME_STATEMENT),
        assessment_mode=normalize_assessment_mode(
            getattr(version, "assessment_mode", None)
            or getattr(upload, "assessment_mode", None)
            or ASSESSMENT_MODE_VARIABLE
        ),
        status=BUDGET_DRAFT_ACTIVE,
        line_items_json=version.line_items_json,
        reserve_study_rows_json=source_draft.reserve_study_rows_json if source_draft else _json_dumps([]),
        reserve_study_warnings_json=source_draft.reserve_study_warnings_json if source_draft else _json_dumps([]),
        reserve_study_status=(source_draft.reserve_study_status if source_draft else None) or "none",
        global_note=version.summary_note,
        statement_month=version.statement_month,
        growth_factor=version.growth_factor,
        growth_factor_note=version.growth_factor_note,
        reserve_inflation_rate=version.reserve_inflation_rate or 0.0,
        reserve_inflation_note=version.reserve_inflation_note,
        budget_preview_json=version.budget_preview_json,
        created_by_user_id=actor["id"],
        updated_by_user_id=actor["id"],
        actor_name=_actor_name(actor),
        created_at=timestamp,
        updated_at=timestamp,
    )
    session.add(draft)
    session.flush()

    route = choose_financial_document_route(upload.original_filename, upload.content_type)
    copied_enriched_snapshot = False
    if route.path == "pdf_vlm" and source_draft is not None:
        source_enriched_key = source_draft.enriched_storage_key
        if _storage_file_available(source_enriched_key):
            source_enriched_path = _budget_storage_path(source_enriched_key or "")
            _persist_draft_enriched_workbook(
                draft,
                enriched_bytes=source_enriched_path.read_bytes(),
            )
            copied_enriched_snapshot = True

    if not copied_enriched_snapshot:
        _refresh_draft_snapshot_from_upload(session, draft, upload)
    _materialize_assessment_mappings_for_line_items(
        session,
        hoa_id=hoa_id,
        line_items=_json_loads(draft.line_items_json, []),
    )

    event = _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="version_reopened",
        summary=f"Reopened {version.version_code} as a new draft",
        upload_id=version.source_upload_id,
        draft_id=draft.id,
        version_id=version.id,
        payload={
            "source_version_id": version.id,
            "source_version_code": version.version_code,
            "source_mode": str(version.source_mode or upload.source_mode or SOURCE_MODE_INCOME_STATEMENT),
            "assessment_mode": normalize_assessment_mode(
                getattr(version, "assessment_mode", None)
                or getattr(upload, "assessment_mode", None)
                or ASSESSMENT_MODE_VARIABLE
            ),
            "old_draft_id": previous_active_draft_id,
            "new_draft_id": draft.id,
        },
    )
    session.commit()
    return BudgetVersionReopenResponse(
        draft=_serialize_draft(draft, upload),
        timeline_event=_serialize_timeline_event(event),
    )


def update_version_metadata(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    version_id: int,
    payload: BudgetVersionMetadataUpdateRequest,
) -> BudgetVersionMetadataUpdateResponse:
    version = _get_version(session, hoa_id, version_id)
    timeline_events: list[BudgetTimelineEvent] = []
    changes: dict[str, Any] = {}
    current_stage = version.stage

    if payload.stage is not None and payload.stage != version.stage:
        version.stage = payload.stage
        changes["stage"] = payload.stage
    if payload.label != version.label:
        version.label = payload.label
        changes["label"] = payload.label
    if payload.summary_note != version.summary_note:
        version.summary_note = payload.summary_note
        changes["summary_note"] = payload.summary_note

    if changes:
        metadata_event = _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="version_metadata_updated",
            summary=f"Updated metadata for {version.version_code}",
            upload_id=version.source_upload_id,
            version_id=version.id,
            payload={"version_code": version.version_code, "changes": changes},
        )
        timeline_events.append(_serialize_timeline_event(metadata_event))

    if current_stage != BUDGET_VERSION_STAGE_FINAL and version.stage == BUDGET_VERSION_STAGE_FINAL:
        finalized_merge_count = 0
        if version.source_draft_id is not None:
            finalized_merge_count = finalize_applied_merges(
                property_id=hoa_id,
                budget_draft_id=version.source_draft_id,
                db_conn=_raw_sqlite_connection(session),
            )
        final_event = _create_audit_event(
            session,
            hoa_id=hoa_id,
            actor=actor,
            event_type="version_marked_final",
            summary=f"Marked {version.version_code} as Final",
            upload_id=version.source_upload_id,
            version_id=version.id,
            payload={
                "version_code": version.version_code,
                "finalized_merge_count": finalized_merge_count,
            },
        )
        timeline_events.append(_serialize_timeline_event(final_event))

    session.commit()
    return BudgetVersionMetadataUpdateResponse(
        version=_serialize_version_detail(version, _get_upload(session, version.source_upload_id)),
        timeline_events=timeline_events,
    )


def record_version_download(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    version_id: int,
) -> tuple[Path, str]:
    version = _get_version(session, hoa_id, version_id)
    if not version.output_storage_key:
        raise LookupError("Version file not found")

    file_path = _budget_storage_path(version.output_storage_key)
    if not file_path.exists():
        raise LookupError("Version file not found")

    _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="file_downloaded",
        summary=f"Downloaded workbook for {version.version_code}",
        upload_id=version.source_upload_id,
        version_id=version.id,
        payload={"version_code": version.version_code, "storage_key": version.output_storage_key},
    )
    session.commit()
    return file_path, f"{version.version_code}-budget.xlsx"


def get_reserve_study_upload_file(
    session: Session,
    *,
    hoa_id: int,
    upload_id: int,
) -> tuple[Path, str]:
    upload = _get_reserve_study_upload(session, hoa_id, upload_id)
    file_path = _budget_storage_path(upload.storage_key)
    if not file_path.exists():
        raise LookupError("Reserve study file not found on disk")
    return file_path, upload.original_filename


def get_income_statement_upload_file(
    session: Session,
    *,
    hoa_id: int,
    upload_id: int,
) -> tuple[Path, str, str]:
    upload = _get_income_statement_upload(session, hoa_id, upload_id)
    file_path = _budget_storage_path(upload.storage_key)
    if not file_path.exists():
        raise LookupError("Income statement file not found on disk")
    if upload.content_type:
        media_type = upload.content_type
    elif choose_financial_document_route(upload.original_filename, upload.content_type).path == "pdf_vlm":
        media_type = "application/pdf"
    else:
        media_type = "application/octet-stream"
    return file_path, upload.original_filename, media_type


def get_income_statement_upload_as_html(
    session: Session,
    *,
    hoa_id: int,
    upload_id: int,
) -> tuple[str, str]:
    """Render every sheet of an Excel-family income-statement upload as plain HTML.

    No header-row detection, no attempt to identify a single "relevant" sheet — this
    is a visual reference for the compare view, not a parser. `.xls` (which openpyxl
    cannot open) is normalized to `.xlsx` first via the same `_ensure_xlsx` conversion
    the upload pipeline already uses.
    """
    from openpyxl import load_workbook

    upload = _get_income_statement_upload(session, hoa_id, upload_id)
    file_path = _budget_storage_path(upload.storage_key)
    if not file_path.exists():
        raise LookupError("Income statement file not found on disk")

    temp_path: Optional[str] = None
    workbook_path = str(file_path)
    try:
        if Path(upload.original_filename).suffix.lower() == ".xls":
            temp_path = _write_temp_workbook(file_path.read_bytes(), upload.original_filename)
            workbook_path = _ensure_xlsx(temp_path)

        workbook = load_workbook(workbook_path, data_only=True)
        try:
            sheet_sections = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows_html = []
                for row in worksheet.iter_rows(values_only=True):
                    cells_html = "".join(
                        f"<td>{html.escape(str(value))}</td>" if value is not None else "<td></td>"
                        for value in row
                    )
                    rows_html.append(f"<tr>{cells_html}</tr>")
                sheet_sections.append(
                    f"<h2>{html.escape(sheet_name)}</h2><table>{''.join(rows_html)}</table>"
                )
        finally:
            workbook.close()
    finally:
        if temp_path is not None:
            Path(temp_path).unlink(missing_ok=True)
        if workbook_path != str(file_path) and workbook_path != temp_path:
            Path(workbook_path).unlink(missing_ok=True)

    document = (
        "<html><head><meta charset=\"utf-8\">"
        "<style>table { border-collapse: collapse; } td { border: 1px solid #ddd; "
        "padding: 4px 8px; font-family: sans-serif; font-size: 13px; white-space: nowrap; } "
        "h2 { font-family: sans-serif; font-size: 14px; }</style>"
        f"</head><body>{''.join(sheet_sections)}</body></html>"
    )
    return document, upload.original_filename


def record_draft_enriched_download(
    session: Session,
    *,
    hoa_id: int,
    actor: dict[str, Any],
    draft_id: int,
) -> tuple[Path, str]:
    draft = _get_draft(session, hoa_id, draft_id)
    file_path = _ensure_draft_enriched_workbook(session, draft)

    _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="draft_enriched_downloaded",
        summary=f"Downloaded enriched workbook for Draft {draft.id}",
        upload_id=draft.source_upload_id,
        draft_id=draft.id,
        version_id=draft.reopened_from_version_id,
        payload={"draft_id": draft.id, "status": draft.status},
    )
    session.commit()
    return file_path, f"draft-{draft.id}-enriched.xlsx"


def delete_active_draft(
    session: Session,
    hoa_id: int,
    actor: dict[str, Any],
) -> dict[str, int]:
    draft = session.scalars(
        select(BudgetDraft).where(
            BudgetDraft.property_id == hoa_id,
            BudgetDraft.status == BUDGET_DRAFT_ACTIVE,
        ).order_by(BudgetDraft.updated_at.desc())
    ).first()
    if draft is None:
        raise LookupError("No active draft")

    has_version = session.scalar(
        select(BudgetVersion.id).where(
            BudgetVersion.source_draft_id == draft.id,
        ).limit(1)
    )
    if has_version is not None:
        raise ValueError("cannot_delete_draft_with_versions")

    draft_id = draft.id
    source_upload_id = draft.source_upload_id
    enriched_key = draft.enriched_storage_key

    _create_audit_event(
        session,
        hoa_id=hoa_id,
        actor=actor,
        event_type="draft_deleted",
        summary=f"Deleted active Draft {draft_id}",
        upload_id=source_upload_id,
        payload={"draft_id": draft_id, "source_upload_id": source_upload_id},
    )

    session.delete(draft)
    session.commit()

    if enriched_key:
        enriched_path = _budget_storage_path(enriched_key)
        if enriched_path.exists():
            enriched_path.unlink(missing_ok=True)

    return {"deleted_draft_id": draft_id}
