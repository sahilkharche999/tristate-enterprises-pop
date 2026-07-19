"""
income_statement_parser.py
==========================

Centralized income statement parser module.

Provides:
- Section state machine: classifies line items by their section position,
  not by label keywords (fixes "reserve" label misclassification).
- Column auto-detection: 3-tier strategy:
    1. Alias matching (multi-row header aware)
    2. LLM zero-shot mapping (if fewer than 2 aliases matched)
    3. Hardcoded fallback indices (if LLM also fails)
- Multi-format Excel reading: .xlsx (openpyxl), .xls (xlrd)
- Financial float parsing: handles $, commas, parentheses negatives, dashes
- Row-level validation: deferred (YTD vs Annual comparison is cross-period, needs YTD Budget extraction)

Public API:
    detect_columns(rows) -> dict[str, int]
    parse_rows_with_sections(rows, col_indices) -> list[dict]
    _match_section_header(label) -> Optional[str]
    _parse_financial_float(value) -> float
    _read_xls_rows / _read_xlsx_rows (format readers used by budget_history_service)

PDF income statements use the Vision (pdf_vlm) path + normalized workbook, not this module.
"""
import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Literal, NamedTuple, Optional, get_args

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Canonical section taxonomy — single source of truth for section classification.
# ---------------------------------------------------------------------------
# Imported by: budget_history_service, normalized_statement_workbook,
# financial_document_extraction (Pydantic model), and taxonomy invariant tests.
#
# Every layer of the pipeline (Gemini extraction, parser state machine,
# `_infer_category`, frontend LineItem) MUST use these exact strings.
SectionKind = Literal["income", "operating", "reserve_income", "reserve_expense"]
SECTION_KINDS: tuple[str, ...] = get_args(SectionKind)

# Sections whose items are displayed as read-only (not editable, not in totals).
# Changing this single constant changes read-only behavior everywhere.
READ_ONLY_SECTIONS: frozenset[str] = frozenset({"reserve_income", "reserve_expense"})

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Section state machine: maps lowercase section header prefixes to category states.
# Reserve is split into reserve_income and reserve_expense so the frontend
# can show them as separate groups with correct totals.
_SECTION_TRANSITIONS: dict[str, str] = {
    "operating income": "income",
    "operating expense": "operating",
    "reserve income": "reserve_income",
    "reserve expense": "reserve_expense",
    "reserve expenses": "reserve_expense",
    "reserve contribution": "reserve_income",
}

# Column header aliases for auto-detection (case-insensitive, stripped)
_HEADER_ALIASES: dict[str, list[str]] = {
    "ytd_actual": ["year to date actual", "ytd actual", "ytd actuals", "year-to-date actual", "ytd"],
    "annual_budget": ["annual budget", "annual budgeted", "total budget"],
    "variance": ["variance", "difference", "$ variance"],
}

# Group-level header terms that span multiple columns
# These appear in a parent row above the detail row (e.g., "Year To Date" above "Actual | Budget | Variance")
_GROUP_HEADERS: dict[str, str] = {
    "year to date": "ytd",
    "annual budget": "annual_budget",
    "current period": "current_period",
}

# Hardcoded fallback column indices (0-based) — used when all detection tiers fail
_FALLBACK_COLUMNS: dict[str, int] = {
    "ytd_actual": 19,
    "annual_budget": 32,
    "variance": 26,
}

# Parentheses negative number pattern: (1,234.56) -> -1234.56
_PAREN_PATTERN = re.compile(r"^\(([0-9,\.]+)\)$")

# Account code extraction: first segment before "-" if it is digits
_ACCOUNT_CODE_PATTERN = re.compile(r"^(\d{4,6})\s*-")

# Reserve study sub-section keywords (for read_only flagging)
_RESERVE_STUDY_SUBHEADERS = {
    "reserve expenses (per reserve study)",
    "reserve expenses per reserve study",
}

_LLM_REQUIRED_COLUMNS = ("ytd_actual", "annual_budget", "variance")
_LLM_PROMPT_MAX_COLUMNS = 40
_LLM_LEFT_SHIFT_LIMIT = 5


class IncomeStatementMissingAnnualColumn(ValueError):
    """Raised when the income-statement extractor cannot identify an
    "Annual Budget" column.

    Per the DRE-driven assessment engine invariant
    (``BudgetDraft.line_items.amount`` is ALWAYS annual; promotion picks
    ONLY the Annual Budget column), the parser MUST refuse to promote
    line items when the Annual Budget column is unresolvable. Otherwise
    a hardcoded fallback column index can silently promote YTD or
    current-period values, which the engine would then treat as annual
    dues and produce wrong assessments.

    Operator sees an actionable error; promotion never happens.
    """

    def __init__(self, *, detected_columns: Optional[dict] = None) -> None:
        self.detected_columns = detected_columns or {}
        super().__init__(
            "Income statement extractor could not identify an 'Annual Budget' "
            "column. The DRE-driven assessment engine requires the annual "
            "amount; no fallback is applied. Detected columns: "
            f"{ {k: v for k, v in self.detected_columns.items() if not k.startswith('_')} }"
        )
_LLM_NUMERIC_CONFIDENCE_RATIO = 0.6


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _normalize(text: str) -> str:
    """NFC unicode, collapse whitespace, strip, lowercase."""
    text = unicodedata.normalize("NFC", str(text))
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _cell_text(row: list, col: int) -> str:
    """Return stripped string for cell at col, or '' if out of range or None."""
    if 0 <= col < len(row) and row[col] is not None:
        return str(row[col]).strip()
    return ""


def _safe_get(row: list, col: int) -> Any:
    """Return cell value at col or None if out of range."""
    return row[col] if 0 <= col < len(row) else None


def _is_numeric_cell(value: Any) -> bool:
    """Check if a cell value is numeric (int, float, or parseable string)."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip().replace(",", "").replace("$", "")
    if text in ("", "-", "--", "None"):
        return False
    try:
        float(text.strip("()"))
        return True
    except (ValueError, TypeError):
        return False


def _max_visible_prompt_width(header_rows: list, sample_rows: list) -> int:
    """Return the max visible width LLM can reference from the compressed prompt."""
    widths = [
        min(len(row), _LLM_PROMPT_MAX_COLUMNS)
        for row in [*header_rows, *sample_rows]
        if row is not None
    ]
    return max(widths, default=0)


def _build_line_item_sample_rows(rows: list, limit: int = 5) -> list:
    """Collect the first few true line-item rows for deterministic numeric validation."""
    sample_rows = []
    for row in rows:
        col_a = _cell_text(row, 0)
        col_b = _cell_text(row, 1)
        if col_a and not col_b and _match_section_header(col_a) is not None:
            continue

        label = col_b or col_a
        if not label or _normalize(label).startswith("total "):
            continue
        if _extract_account_code(label) is None:
            continue

        sample_rows.append(row)
        if len(sample_rows) >= limit:
            break

    return sample_rows


def _reject_llm_mapping(reason: str, mapping: Optional[dict], detail: str = "") -> None:
    """Log a rejected LLM mapping with a stable reason code."""
    suffix = f" ({detail})" if detail else ""
    logger.warning("Rejecting LLM column mapping [%s]: %s%s", reason, mapping, suffix)


def _sanitize_llm_column_map(col_map: Optional[dict], max_prompt_width: int) -> Optional[dict]:
    """Reject malformed or impossible LLM column suggestions before validation."""
    if not col_map or max_prompt_width <= 0:
        _reject_llm_mapping("out_of_range", col_map, "no visible prompt columns")
        return None

    sanitized: dict[str, int] = {}
    for key in _LLM_REQUIRED_COLUMNS:
        if key not in col_map or type(col_map[key]) is not int:
            _reject_llm_mapping("invalid_type", col_map, f"{key} must be int")
            return None

        idx = col_map[key]
        if idx < 0:
            _reject_llm_mapping("negative", col_map, f"{key}={idx}")
            return None
        if idx >= max_prompt_width:
            _reject_llm_mapping("out_of_range", col_map, f"{key}={idx}, width={max_prompt_width}")
            return None
        sanitized[key] = idx

    if len(set(sanitized.values())) != len(sanitized):
        _reject_llm_mapping("duplicate", sanitized)
        return None

    return sanitized


def _numeric_ratio_for_column(data_rows: list, col_idx: int) -> float:
    """Return the fraction of sample rows with numeric values in a candidate column."""
    if not data_rows:
        return 0.0
    numeric_count = sum(1 for row in data_rows if _is_numeric_cell(_safe_get(row, col_idx)))
    return numeric_count / len(data_rows)


def _repair_llm_column_left(col_idx: int, data_rows: list) -> Optional[int]:
    """Accept the proposed column or shift left to the first high-confidence numeric column."""
    for offset in range(_LLM_LEFT_SHIFT_LIMIT + 1):
        candidate = col_idx - offset
        if candidate < 0:
            break
        if _numeric_ratio_for_column(data_rows, candidate) >= _LLM_NUMERIC_CONFIDENCE_RATIO:
            return candidate
    return None


def _validate_llm_columns_against_data(col_map: Optional[dict], data_rows: list) -> Optional[dict]:
    """Require LLM-proposed columns to prove themselves on actual line-item rows."""
    if not col_map:
        return None
    if not data_rows:
        _reject_llm_mapping("low_confidence", col_map, "no line-item sample rows")
        return None

    validated: dict[str, int] = {}
    for key, col_idx in col_map.items():
        repaired = _repair_llm_column_left(col_idx, data_rows)
        if repaired is None:
            ratio = _numeric_ratio_for_column(data_rows, col_idx)
            _reject_llm_mapping("low_confidence", col_map, f"{key}={col_idx}, ratio={ratio:.2f}")
            return None
        if repaired != col_idx:
            logger.info("LLM column %s adjusted LEFT from %d to %d", key, col_idx, repaired)
        validated[key] = repaired

    if len(set(validated.values())) != len(validated):
        _reject_llm_mapping("duplicate", validated, "after left-shift repair")
        return None

    return validated


# ---------------------------------------------------------------------------
# Financial float parser
# ---------------------------------------------------------------------------

def _parse_financial_float(value: Any) -> float:
    """Backward-compatible float wrapper over :func:`parse_financial_cell`.

    Empty / dash / unparseable all collapse to ``0.0`` here so existing
    arithmetic call sites (display columns, coverage heuristics) keep their
    numeric behavior. Callers that must distinguish "no value" or
    "unparseable" from a real zero — notably the promoted annual column —
    call :func:`parse_financial_cell` directly and act on ``kind``.

    Negative forms ("(1,234)", "1,234-", "-1234") now parse WITH sign; that
    is a strict correctness improvement over the previous silent ``0``.
    """
    cell = parse_financial_cell(value)
    return cell.value if cell.value is not None else 0.0


class ParsedCell(NamedTuple):
    """Outcome of parsing one financial cell (C5).

    ``kind`` distinguishes states the legacy ``-> float`` parsers collapsed
    into ``0.0``:

    - ``ok``: a real numeric value is in ``value``.
    - ``empty``: blank / None \u2014 "no value present".
    - ``dash``: only a dash / em-dash \u2014 "no value stated" (explicit
      not-applicable, distinct from a true ``0``).
    - ``unparseable``: text we could not turn into a number (OCR noise,
      European decimals, malformed input). ``value`` is ``None``; this MUST
      surface for operator review rather than silently become ``0``.

    ``value`` is ``None`` for every kind except ``ok``. Zero is only ever
    produced by a cell that actually states zero.
    """

    value: Optional[float]
    kind: str
    raw: str


def parse_financial_cell(value: Any) -> ParsedCell:
    """Shared numeric normalizer for every financial-cell parse path (C5).

    Normalization order is fixed so sign survives currency/thousands
    formatting: strip ``$``, commas, and spaces FIRST, then detect the
    negative forms. Fixes the class of defects where ``($1,234)`` and
    ``1,234-`` silently became ``0`` (sign AND magnitude lost).

    Never returns ``0`` as a fallback \u2014 unreadable input yields
    ``kind='unparseable'`` with ``value=None``.
    """
    if value is None:
        return ParsedCell(None, "empty", "")
    # bool is an int subclass \u2014 never a financial amount.
    if isinstance(value, bool):
        return ParsedCell(None, "unparseable", str(value))
    if isinstance(value, (int, float)):
        return ParsedCell(float(value), "ok", str(value))

    text = str(value).strip()
    if not text:
        return ParsedCell(None, "empty", str(value))
    if text in {"-", "\u2014", "\u2013"}:
        return ParsedCell(None, "dash", text)

    # Strip currency, thousands separators, and spaces BEFORE detecting
    # negative forms, so a currency symbol inside parens ("($1,234)") no
    # longer defeats the negative match.
    cleaned = (
        text.replace("$", "")
        .replace(",", "")
        .replace(" ", "")
        .replace("\u00a0", "")
        .strip()
    )

    negative = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        negative = True
        cleaned = cleaned[1:-1].strip()
    elif cleaned.endswith("-"):  # trailing-minus accounting negative
        negative = True
        cleaned = cleaned[:-1].strip()
    elif cleaned.startswith("-"):
        negative = True
        cleaned = cleaned[1:].strip()
    elif cleaned.startswith("+"):
        cleaned = cleaned[1:].strip()

    if not cleaned:
        return ParsedCell(None, "unparseable", text)
    try:
        num = float(cleaned)
    except ValueError:
        return ParsedCell(None, "unparseable", text)
    return ParsedCell(-num if negative else num, "ok", text)


# ---------------------------------------------------------------------------
# Section state machine
# ---------------------------------------------------------------------------

def _match_section_header(label: str) -> Optional[str]:
    """Return the section category for a recognized section header, or None.

    Uses case-insensitive starts-with matching — consistent with the HOA
    accounting software that all 81 HOAs use.

    Args:
        label: Cell text from column A of the row.

    Returns:
        "income", "operating", or "reserve" if recognized; None otherwise.
    """
    if not label or not label.strip():
        return None
    normalized = _normalize(label)
    for prefix, state in _SECTION_TRANSITIONS.items():
        if normalized.startswith(prefix):
            return state
    return None


def _classify_by_account_code(account_code: Optional[int]) -> str:
    """Fallback classification when no section headers detected.

    Account code ranges:
      40000-49999 = income (or reserve_income if in reserve context)
      50000-89999 = operating
      90000+      = reserve_expense
    """
    if account_code is None:
        return "operating"
    if 40000 <= account_code <= 49999:
        return "income"
    if 50000 <= account_code <= 89999:
        return "operating"
    if account_code >= 90000:
        return "reserve_expense"
    return "operating"


def _extract_account_code(label: str) -> Optional[int]:
    """Extract leading 4-6 digit account code from label, or None."""
    if not label:
        return None
    m = _ACCOUNT_CODE_PATTERN.match(label.strip())
    if m:
        return int(m.group(1))
    # Try raw leading digits
    head = label.split("-", 1)[0].strip()
    if head.isdigit():
        return int(head)
    return None


# ---------------------------------------------------------------------------
# Column auto-detection
# ---------------------------------------------------------------------------

def _llm_column_fallback(header_rows: list, sample_rows: list) -> Optional[dict]:
    """Tier 2: LLM fallback for zero-shot column detection.

    Sends header rows + up to 3 sample data rows to LLM for zero-shot
    schema mapping. Uses sheet compression to minimize tokens.

    Args:
        header_rows: The first <=10 rows of the sheet (header area).
        sample_rows: A few data rows (up to 3) to help the LLM understand the layout.

    Returns:
        Dict with keys {ytd_actual, annual_budget, variance} (0-based indices),
        or None if LLM fails.
    """
    import asyncio
    from pydantic import BaseModel

    class ColumnMapping(BaseModel):
        ytd_actual: int
        annual_budget: int
        variance: int

    # Compress: take header rows + up to 3 sample data rows
    compressed = []
    for row in header_rows:
        compressed.append([str(cell) if cell is not None else "" for cell in row[:_LLM_PROMPT_MAX_COLUMNS]])
    for row in sample_rows[:3]:
        compressed.append([str(cell) if cell is not None else "" for cell in row[:_LLM_PROMPT_MAX_COLUMNS]])

    # Format as a readable table for the LLM
    table_text = ""
    for i, row in enumerate(compressed):
        non_empty = [(j, v) for j, v in enumerate(row) if v.strip()]
        if non_empty:
            table_text += f"Row {i}: " + ", ".join(f"col{j}={v}" for j, v in non_empty) + "\n"

    if not table_text.strip():
        return None

    messages = [
        {
            "role": "system",
            "content": (
                "You are a financial spreadsheet analyst. Given rows from an income statement, "
                "identify the 0-based column indices for: ytd_actual (Year-To-Date Actual), "
                "annual_budget (Annual Budget), and variance. Return only non-negative 0-based "
                "indices that reference visible colN entries in the provided rows. Never return "
                "negative placeholders like -1, never reuse the same index for multiple fields, "
                "and never guess hidden columns. "
                'Return JSON: {"ytd_actual": <int>, "annual_budget": <int>, "variance": <int>}'
            ),
        },
        {
            "role": "user",
            "content": (
                f"Here are the header rows and first 3 data rows from an income statement:\n\n"
                f"{table_text}\n\nIdentify the 0-based column indices. Use only visible colN "
                f"entries from the rows above, and do not use negative values or duplicates."
            ),
        },
    ]

    try:
        from ..ai_implementation.pipeline.llm_client import call_llm

        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            loop = None

        if loop and loop.is_running():
            import concurrent.futures

            def _run_in_new_loop():
                new_loop = asyncio.new_event_loop()
                try:
                    return new_loop.run_until_complete(
                        call_llm(messages, ColumnMapping, temperature=0.0, timeout=15.0)
                    )
                finally:
                    new_loop.close()

            with concurrent.futures.ThreadPoolExecutor() as pool:
                result = pool.submit(_run_in_new_loop).result()
        else:
            result = asyncio.run(call_llm(messages, ColumnMapping, temperature=0.0, timeout=15.0))

        if result is not None:
            mapping = {
                "ytd_actual": result.ytd_actual,
                "annual_budget": result.annual_budget,
                "variance": result.variance,
            }
            logger.info("LLM column detection succeeded: %s", mapping)
            return mapping

    except Exception as e:
        logger.warning("LLM column fallback failed: %s", e)

    return None


def _validate_columns_against_data(col_map: dict, data_rows: list) -> dict:
    """Validate that matched columns have numeric data. If not, scan LEFT.

    Merged cells can cause header text to sit at the RIGHT edge of the merge
    while data lives at the LEFT edge. This function checks each column and
    adjusts leftward if needed.
    """
    if not data_rows or not col_map:
        return col_map
    validated = {}
    for key, col_idx in col_map.items():
        has_data = any(
            _is_numeric_cell(_safe_get(r, col_idx))
            for r in data_rows
        )
        if has_data:
            validated[key] = col_idx
        else:
            found = False
            for offset in range(1, 6):
                candidate = col_idx - offset
                if candidate < 0:
                    break
                count = sum(1 for r in data_rows if _is_numeric_cell(_safe_get(r, candidate)))
                if count >= max(len(data_rows) // 2, 1):
                    logger.info("Column %s: position %d has no data, adjusted LEFT to %d", key, col_idx, candidate)
                    validated[key] = candidate
                    found = True
                    break
            if not found:
                validated[key] = col_idx
    return validated


def detect_columns(rows: list) -> dict:
    """Scan rows 0-9 for column header row(s). Return 0-based column indices.

    3-tier detection strategy:
    1. Alias matching against known header terms (multi-row header aware).
       - For multi-row headers like Esprit Park (group row + detail row),
         find group spans first, then locate "Actual"/"Variance" within groups.
    2. If fewer than 2 required columns found -> LLM zero-shot mapping
       with headers + 3 sample rows (per CONTEXT.md locked decision).
    3. If LLM also fails -> hardcoded fallback indices.

    Returns _FALLBACK_COLUMNS if all tiers fail.
    """
    scan_rows = rows[:10] if len(rows) >= 10 else rows

    matched: dict[str, int] = {}

    # -- Tier 1a: Group-header pass (multi-row headers like Esprit Park) --
    # Find rows that contain group labels ("Year To Date", "Annual Budget", "Current Period")
    # Record which columns they span.
    group_spans: dict[str, tuple[int, int]] = {}  # group_key -> (start_col, end_col)

    for row_idx, row in enumerate(scan_rows):
        group_positions: list[tuple[int, str]] = []
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_norm = _normalize(str(cell))
            for grp_prefix, grp_key in _GROUP_HEADERS.items():
                if cell_norm.startswith(grp_prefix):
                    group_positions.append((col_idx, grp_key))

        if not group_positions:
            continue

        # Assign spans: from this col to the next group's start col (or end of row)
        group_positions.sort(key=lambda x: x[0])
        for i, (col_start, grp_key) in enumerate(group_positions):
            col_end = group_positions[i + 1][0] - 1 if i + 1 < len(group_positions) else len(row) - 1
            group_spans[grp_key] = (col_start, col_end)

    # -- Tier 1b: Detail-row alias pass --
    # If group spans were found, look for "Actual"/"Variance"/"Budget" in the detail row
    # and constrain to appropriate spans.
    for row_idx, row in enumerate(scan_rows):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_norm = _normalize(str(cell))

            for canonical_key, aliases in _HEADER_ALIASES.items():
                if canonical_key in matched:
                    continue
                for alias in aliases:
                    if cell_norm == alias or cell_norm.startswith(alias):
                        matched[canonical_key] = col_idx
                        break

            # Special handling for bare "Actual" and "Budget" with group context
            if "ytd_actual" not in matched and cell_norm == "actual":
                ytd_span = group_spans.get("ytd")
                if ytd_span:
                    start, end = ytd_span
                    if start <= col_idx <= end:
                        matched["ytd_actual"] = col_idx
                elif not group_spans:
                    # No group context — accept the first "Actual" we find
                    matched.setdefault("ytd_actual", col_idx)

            if "annual_budget" not in matched and cell_norm in ("actual", "budget"):
                ab_span = group_spans.get("annual_budget")
                if ab_span:
                    start, end = ab_span
                    if start <= col_idx <= end:
                        matched["annual_budget"] = col_idx

            if "variance" not in matched and cell_norm in ("variance",):
                ytd_span = group_spans.get("ytd")
                if ytd_span:
                    start, end = ytd_span
                    if start <= col_idx <= end:
                        matched["variance"] = col_idx
                elif not group_spans:
                    matched.setdefault("variance", col_idx)

    # -- Tier 1c: Data-validation pass --
    data_rows = [r for r in rows[5:15] if len(r) > 2 and r[1]]  # rows with col B labels
    if data_rows and matched:
        matched = _validate_columns_against_data(matched, data_rows)

    # Also fix variance: prefer YTD variance over Current Period variance
    if "variance" in matched and group_spans.get("ytd"):
        ytd_start, ytd_end = group_spans["ytd"]
        if not (ytd_start - 3 <= matched["variance"] <= ytd_end):
            for row_idx, row in enumerate(scan_rows):
                for col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_norm = _normalize(str(cell))
                    if cell_norm == "variance" and ytd_start - 3 <= col_idx <= ytd_end:
                        matched["variance"] = col_idx
                        break

    if len(matched) >= 2:
        real_keys = [k for k in matched.keys() if not k.startswith("_")]
        for key in _FALLBACK_COLUMNS:
            matched.setdefault(key, _FALLBACK_COLUMNS[key])
        matched["_detection_tier"] = 1
        matched["_real_matched_keys"] = real_keys
        logger.debug("Column detection tier 1 (alias): %s", matched)
        return matched

    # -- Tier 2: LLM fallback --
    header_rows = scan_rows
    sample_rows = rows[10:15] if len(rows) > 10 else rows[5:]
    line_item_sample_rows = _build_line_item_sample_rows(rows)
    max_prompt_width = _max_visible_prompt_width(header_rows, sample_rows)
    llm_result = _llm_column_fallback(header_rows, sample_rows)
    llm_result = _sanitize_llm_column_map(llm_result, max_prompt_width)
    llm_result = _validate_llm_columns_against_data(llm_result, line_item_sample_rows)
    if llm_result is not None and len(llm_result) >= 2:
        real_keys = [k for k in llm_result.keys() if not k.startswith("_")]
        for key in _FALLBACK_COLUMNS:
            llm_result.setdefault(key, _FALLBACK_COLUMNS[key])
        llm_result["_detection_tier"] = 2
        llm_result["_real_matched_keys"] = real_keys
        logger.info("Column detection tier 2 (LLM): %s", llm_result)
        return llm_result

    # -- Tier 3: Hardcoded fallback --
    logger.warning("Column detection: all tiers failed, using hardcoded fallback %s", _FALLBACK_COLUMNS)
    result = dict(_FALLBACK_COLUMNS)
    result["_detection_tier"] = 3
    result["_real_matched_keys"] = []
    return result


# ---------------------------------------------------------------------------
# Row parsing with section state machine
# ---------------------------------------------------------------------------

def parse_rows_with_sections(
    rows: list,
    col_indices: dict,
    capture: Optional[dict] = None,
) -> list:
    """Walk rows top-to-bottom, track section state, classify each line item.

    Classification is determined entirely by section position, NOT by label text.
    This is the core fix: "90000 - Reserve - Allocation/Transfer" under
    Operating Expense stays category="operating" and read_only=False.

    Args:
        rows: Normalized rows (list of lists), 0-based columns.
        col_indices: 0-based column indices from detect_columns().
        capture: Optional out-dict (C5/C6). When provided, two lists are
            appended to in place:

            - ``capture["stated_totals"]``: one entry per recognized
              "Total …" row — ``{section, label, ytd_actual, annual_budget}``
              — the document's own stated totals, captured instead of
              discarded so the subtotal cross-check can run downstream.
            - ``capture["review_questions"]``: one entry per line item whose
              PROMOTED (annual-budget) cell was unparseable —
              ``{label, account_code, raw_text, column}`` — so the caller
              can raise a blocking review issue instead of silently
              promoting a fake ``0``.

            When ``capture`` is None behavior is unchanged except that an
            unparseable annual cell now yields ``annual_budget=None``
            (previously a silent ``0.0``).

    Returns:
        List of line item dicts with keys:
        line_item_key, account_code, category, label, ytd_actual,
        annual_budget, projection, percent_change, read_only,
        section, validation_warning, raw
    """
    real_matched_keys = col_indices.get("_real_matched_keys") if col_indices else None
    # Pop metadata key so it doesn't interfere with column lookups
    col_indices = {k: v for k, v in col_indices.items() if not k.startswith("_")}

    def _capture_total_row(row: list, label: str, section: str) -> None:
        if capture is None:
            return
        ytd_idx = col_indices.get("ytd_actual", _FALLBACK_COLUMNS["ytd_actual"])
        annual_idx = col_indices.get("annual_budget", _FALLBACK_COLUMNS["annual_budget"])
        ytd_cell = parse_financial_cell(_safe_get(row, ytd_idx))
        annual_cell = parse_financial_cell(_safe_get(row, annual_idx))
        capture.setdefault("stated_totals", []).append(
            {
                "section": section,
                "label": label.strip(),
                "ytd_actual": ytd_cell.value,
                "annual_budget": annual_cell.value,
            }
        )

    current_section = "operating"  # safe default per spec
    in_reserve_study_block = False
    section_transition_count = 0
    line_items = []

    for row in rows:
        col_a = _cell_text(row, 0)
        col_b = _cell_text(row, 1)

        # Section / sub-section header: col A has text, col B is empty
        if col_a and not col_b:
            label_norm = _normalize(col_a)

            # "Total X" section rows: captured as the document's stated
            # totals (C6) — still excluded from line items.
            if label_norm.startswith("total "):
                _capture_total_row(row, col_a, current_section)
                continue

            # Check for a major section transition
            new_section = _match_section_header(col_a)
            if new_section is not None:
                current_section = new_section
                in_reserve_study_block = False  # reset on major section change
                section_transition_count += 1

            # Track reserve study sub-section (triggers read_only=True)
            if label_norm.startswith("reserve expenses"):
                in_reserve_study_block = True

            continue  # section/sub-section row — not a data row

        # Line item: col B has text (or col A with no col B alternative)
        label = col_b if col_b else col_a
        if not label:
            continue

        # "Total X" data rows: captured as stated totals (C6), not line items.
        if _normalize(label).startswith("total "):
            _capture_total_row(row, label, current_section)
            continue

        account_code = _extract_account_code(label)
        is_read_only = current_section in READ_ONLY_SECTIONS

        ytd_idx = col_indices.get("ytd_actual", _FALLBACK_COLUMNS["ytd_actual"])
        annual_idx = col_indices.get("annual_budget", _FALLBACK_COLUMNS["annual_budget"])
        variance_idx = col_indices.get("variance", _FALLBACK_COLUMNS["variance"])
        projection_idx = col_indices.get("projection", 37)
        pct_change_idx = col_indices.get("percent_change", 38)

        ytd_actual = _parse_financial_float(_safe_get(row, ytd_idx))
        # The annual column feeds BudgetLineInput.amount (the legally-binding
        # assessment basis), so it uses the tagged parse (C5): a dash/empty
        # cell is "no value stated" (None, non-blocking) and an unparseable
        # cell is None PLUS a review question — never a silent 0.
        annual_cell = parse_financial_cell(_safe_get(row, annual_idx))
        annual_budget = annual_cell.value
        if annual_cell.kind == "unparseable" and capture is not None:
            capture.setdefault("review_questions", []).append(
                {
                    "label": label.strip(),
                    "account_code": account_code,
                    "raw_text": annual_cell.raw,
                    "column": "annual_budget",
                }
            )
        variance = _parse_financial_float(_safe_get(row, variance_idx))

        # Per the DRE-driven assessment engine invariant
        # (BudgetDraft.line_items.amount = annual), every line records
        # which source column its annual_budget came from. 'annual_budget'
        # = real match against the source's Annual Budget column;
        # 'annual_budget (fallback)' = column-detection fallback was used
        # (operator should treat as suspect).
        annual_real = (
            "annual_budget" in real_matched_keys
            if real_matched_keys is not None
            else False
        )
        source_column = "annual_budget" if annual_real else "annual_budget (fallback)"

        line_items.append(
            {
                "line_item_key": str(account_code if account_code is not None else label),
                "account_code": account_code,
                "category": current_section,
                "label": label.strip(),
                "ytd_actual": ytd_actual,
                "annual_budget": annual_budget,
                "projection": _parse_financial_float(_safe_get(row, projection_idx)),
                "percent_change": _parse_financial_float(_safe_get(row, pct_change_idx)),
                "read_only": is_read_only,
                "section": current_section,
                "source_column": source_column,
                "source_page_or_cell": None,  # Excel doesn't carry per-cell location here
                "raw": {
                    "section": current_section,
                    "label": col_b.strip() if col_b else None,
                },
            }
        )

    # Fallback: if no section headers were detected, reclassify by account code ranges
    if section_transition_count == 0 and line_items:
        for item in line_items:
            item["category"] = _classify_by_account_code(item["account_code"])
            item["section"] = item["category"]

    return line_items


def parse_rows_with_sections_strict(
    rows: list,
    col_indices: dict,
) -> list:
    """Strict-mode wrapper around ``parse_rows_with_sections``.

    Refuses to proceed when the column detector did NOT find a real
    "Annual Budget" column match (i.e. ``annual_budget`` was filled by
    the hardcoded fallback rather than alias-matched or LLM-suggested).

    Use this entry point when promoting parsed rows to
    ``BudgetDraft.line_items`` for the DRE-driven assessment engine.
    The legacy ``parse_rows_with_sections`` keeps lenient behavior for
    review-screen display where YTD/variance are useful even when the
    Annual Budget column is missing.

    Raises:
        IncomeStatementMissingAnnualColumn: when ``annual_budget`` is
            not in ``col_indices['_real_matched_keys']``.
    """
    real_matched_keys = col_indices.get("_real_matched_keys", [])
    if "annual_budget" not in (real_matched_keys or []):
        raise IncomeStatementMissingAnnualColumn(detected_columns=col_indices)
    return parse_rows_with_sections(rows, col_indices)


# ---------------------------------------------------------------------------
# File format readers
# ---------------------------------------------------------------------------

def _read_xlsx_rows(path: str, sheet_name: str = "Income Statement") -> list:
    """Read .xlsx file into normalized list of lists.

    Uses openpyxl data_only=True to read cached values (not formulas).
    Returns rows as list of lists with None for blank cells.

    H13: hidden rows and hidden columns are MASKED (blanked to None) rather
    than deleted, so hidden subtotal/scratch rows can't become phantom line
    items and a hidden column can't shift positional column detection — while
    original row/column indices stay aligned for any position-keyed consumer
    (e.g. the Excel source preview). Exclusions are logged; a sheet that is
    mostly hidden is logged prominently so an accidental hide is visible.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        hidden_cols = {
            c
            for c in range(1, ws.max_column + 1)
            if get_column_letter(c) in ws.column_dimensions
            and ws.column_dimensions[get_column_letter(c)].hidden
        }
        hidden_rows_seen = 0
        nonempty_rows = 0
        rows = []
        for r in range(1, ws.max_row + 1):
            raw = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            row_has_data = any(v is not None and str(v).strip() != "" for v in raw)
            row_hidden = (
                r in ws.row_dimensions and ws.row_dimensions[r].hidden
            )
            if row_has_data:
                nonempty_rows += 1
            if row_hidden:
                if row_has_data:
                    hidden_rows_seen += 1
                # Blank the whole hidden row (preserves index alignment).
                rows.append([None] * len(raw))
                continue
            # Mask hidden columns within a visible row.
            row = [
                None if (c + 1) in hidden_cols else v
                for c, v in enumerate(raw)
            ]
            rows.append(row)

        if hidden_rows_seen or hidden_cols:
            logger.warning(
                "H13: excluded %d hidden data row(s) and %d hidden column(s) "
                "from %s while parsing", hidden_rows_seen, len(hidden_cols),
                Path(path).name,
            )
        # A mostly-hidden sheet is almost always an accidental hide — surface
        # it prominently rather than importing a near-empty statement.
        if nonempty_rows and hidden_rows_seen > nonempty_rows / 2:
            logger.warning(
                "H13: %d of %d non-empty rows in %s are HIDDEN — the parsed "
                "statement may be nearly empty; operator should unhide and "
                "re-upload.", hidden_rows_seen, nonempty_rows, Path(path).name,
            )
        return rows
    finally:
        wb.close()


def _read_xls_rows(path: str, sheet_name: str = "Income Statement") -> list:
    """Read .xls (Excel 97-2003) file into normalized list of lists.

    Uses xlrd. Normalizes empty string "" to None to match openpyxl convention.
    """
    import xlrd

    # H13: formatting_info=True is what exposes xlrd's rowinfo_map/colinfo_map
    # hidden flags. When it's unavailable (some .xls variants), we cannot see
    # hidden state — proceed as before and note the limitation.
    have_formatting = True
    try:
        wb = xlrd.open_workbook(path, formatting_info=True)
    except Exception:
        wb = xlrd.open_workbook(path)
        have_formatting = False
    try:
        ws = wb.sheet_by_name(sheet_name)
    except xlrd.XLRDError:
        ws = wb.sheet_by_index(0)

    # Hidden rows/columns from the format info (empty sets when unavailable).
    hidden_rows = set()
    hidden_cols = set()
    if have_formatting:
        for r, info in (getattr(ws, "rowinfo_map", {}) or {}).items():
            if getattr(info, "hidden", 0):
                hidden_rows.add(r)
        for c, info in (getattr(ws, "colinfo_map", {}) or {}).items():
            if getattr(info, "hidden", 0):
                hidden_cols.add(c)
    elif Path(path).suffix.lower() == ".xls":
        logger.warning(
            "H13: could not read hidden-cell formatting for %s; hidden rows/"
            "columns (if any) were NOT excluded.", Path(path).name,
        )

    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            # Mask hidden rows/columns to None (same as the xlsx path).
            if r in hidden_rows or c in hidden_cols:
                row.append(None)
                continue
            v = ws.cell_value(r, c)
            # Normalize blank cells to None (openpyxl convention)
            row.append(None if v == "" else v)
        rows.append(row)

    if hidden_rows or hidden_cols:
        logger.warning(
            "H13: excluded %d hidden row(s) and %d hidden column(s) from %s",
            len(hidden_rows), len(hidden_cols), Path(path).name,
        )

    # Propagate merged cell values to the leftmost column of each merge.
    # xlrd stores merged_cells as (row_lo, row_hi, col_lo, col_hi) ranges.
    # The cell value may live at any position in the merge, so copy it to col_lo
    # if col_lo is currently None.
    for rlo, rhi, clo, chi in getattr(ws, "merged_cells", []):
        # Find the value anywhere in this merge range
        val = None
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                if r < len(rows) and c < len(rows[r]) and rows[r][c] is not None:
                    val = rows[r][c]
                    break
            if val is not None:
                break
        # Place the value at (rlo, clo) — the top-left of the merge
        if val is not None and rlo < len(rows) and clo < len(rows[rlo]):
            rows[rlo][clo] = val

    return rows
