"""Workbook helpers used by the budget history pipeline.

Excel-tool macros HTTP endpoints were removed (ponytail Tier B). This module
keeps only the deterministic sheet read / percent-change helpers that live
budget generate and enrich still call.
"""
import logging
import re
import time
import unicodedata
from typing import Any, Dict

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

_AM_COL = column_index_from_string("AM")  # column 39, pre-computed once

logger = logging.getLogger(__name__)


def _normalize_label(raw: str) -> str:
    """Normalize a label for matching: NFC unicode, collapse whitespace, strip."""
    text = unicodedata.normalize("NFC", str(raw))
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def read_sheet_as_table(path: str, sheet: str, header_row: int = 1) -> Dict[str, Any]:
    """Read a worksheet into headers and rows (skip empty rows).

    Args:
        path: path to workbook (.xlsx)
        sheet: worksheet name
        header_row: 1-based row index containing headers (default 1)

    Returns:
        Dict with keys: sheet, headers, rows
    """
    start = time.perf_counter()
    logger.info("read_sheet_as_table start: path=%s sheet=%s header_row=%s", path, sheet, header_row)
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        max_col = ws.max_column
        headers = []
        if header_row <= ws.max_row:
            for c in range(1, max_col + 1):
                headers.append(ws.cell(row=header_row, column=c).value)
        else:
            headers = [None] * max_col

        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(v is not None for v in row):
                rows.append(row)
        result = {"sheet": sheet, "headers": headers, "rows": rows}
        duration = time.perf_counter() - start
        logger.info("read_sheet_as_table done: path=%s sheet=%s rows=%d headers=%d duration=%.3fs", path, sheet, len(rows), len(headers), duration)
        return result
    finally:
        wb.close()

def read_first_sheet_preview(path: str, max_rows: int) -> Dict[str, Any]:
    """Read the first worksheet up to max_rows, returning {sheet, headers, rows}."""
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        max_col = ws.max_column
        headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        rows = []
        for r in range(2, min(ws.max_row, max_rows) + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if any(v is not None for v in row):
                rows.append(row)
        return {"sheet": ws.title, "headers": headers, "rows": rows}
    finally:
        wb.close()

def write_percent_changes_by_label(path: str, sheet: str, changes: Dict[str, float], pct_change_col: int = None) -> int:
    """Write decimal percent change values to the % Change column for rows whose col B label matches.

    Args:
        path: path to workbook (.xlsx)
        sheet: worksheet name (typically "Income Statement")
        changes: mapping of label → decimal percent change (e.g. {"Insurance": 0.085})
        pct_change_col: 1-based column index to write to (default: auto-detect or fallback to AM=39)

    Returns:
        Number of labels that were matched and written.
    """
    if not changes:
        return 0

    normalized_changes: Dict[str, float] = {}
    norm_to_original: Dict[str, str] = {}
    for key, value in changes.items():
        norm_key = _normalize_label(key)
        normalized_changes[norm_key] = value
        norm_to_original[norm_key] = key

    wb = load_workbook(path)
    matched_keys: set = set()
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active

        # Determine the % Change column position
        target_col = pct_change_col or _AM_COL
        if pct_change_col is None:
            # Auto-detect: scan first 10 rows for "% Change" header
            for r in range(1, min(11, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    if str(val or "").strip().lower() in ("% change", "percent change"):
                        target_col = c
                        break
                if target_col != _AM_COL:
                    break

        for r in range(1, ws.max_row + 1):
            # Legacy statements keep the label in column B. Normalized PDF
            # workbooks keep Account Code in B and Label in C. Check both so
            # manual percent changes survive the PDF -> XLSX round-trip.
            for label_col in (2, 3):
                label = ws.cell(row=r, column=label_col).value
                if label is None:
                    continue
                norm_label = _normalize_label(str(label))
                if norm_label in normalized_changes:
                    ws.cell(row=r, column=target_col, value=normalized_changes[norm_label])
                    matched_keys.add(norm_label)
                    break
        wb.save(path)
    finally:
        wb.close()

    matched = len(matched_keys)
    total_requested = len(normalized_changes)
    if matched < total_requested:
        unmatched = [norm_to_original[nk] for nk in normalized_changes if nk not in matched_keys]
        logger.warning(
            "write_percent_changes_by_label: matched %d/%d labels. Unmatched: %r",
            matched, total_requested, unmatched,
        )
    else:
        logger.info("write_percent_changes_by_label: all %d labels matched.", total_requested)

    return matched

