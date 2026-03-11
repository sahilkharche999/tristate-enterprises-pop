"""HTTP router exposing macro-like endpoints."""
from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Response
from ..services import macros_service
from ..models.schemas import TableResponse, SumResponse, DupsResponse, SimpleResponse
from ..config import settings
import os
import tempfile
import shutil

router = APIRouter()


def _tmp_save_path(suffix: str = '.xlsx') -> str:
    os.makedirs(settings.TEMP_DIR, exist_ok=True)
    fd, path = tempfile.mkstemp(suffix=suffix, dir=settings.TEMP_DIR)
    os.close(fd)
    return path


def save_upload_tmp(upload_file: UploadFile) -> str:
    suffix = os.path.splitext(upload_file.filename or "")[1] or ".xlsx"
    path = _tmp_save_path(suffix)
    with open(path, "wb") as out:
        content = upload_file.file.read()
        out.write(content)
    return path


@router.post("/macros/payment-search", response_model=TableResponse)
async def payment_search(file: UploadFile = File(...), sheet: str = Form('Sheet1')):
    path = save_upload_tmp(file)
    try:
        res = macros_service.payment_search_format(path, sheet)
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/sum-and-format", response_model=SumResponse)
async def sum_and_format(
    file: UploadFile = File(...), sheet: str = Form(...), start_cell: str = Form(...), row_count: int = Form(...), target_col: str = Form('H')
):
    path = save_upload_tmp(file)
    try:
        res = macros_service.sum_and_format(path, sheet, start_cell, row_count, target_col)
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/ach-sum", response_model=SumResponse)
async def ach_sum(file: UploadFile = File(...), sheet: str = Form(...), start_cell: str = Form(...), find_text: str = Form('ACH Draft')):
    path = save_upload_tmp(file)
    try:
        res = macros_service.ach_sum(path, sheet, start_cell, find_text)
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/one-cell", response_model=SimpleResponse)
async def one_cell(file: UploadFile = File(...), sheet: str = Form(...), start_cell: str = Form(...)):
    path = save_upload_tmp(file)
    try:
        r = macros_service.one_cell(path, sheet, start_cell)
        return {"message": f"Copied value {r.get('value')}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/find-dups", response_model=DupsResponse)
async def find_dups(file: UploadFile = File(...), sheet: str = Form(...), lookup_col: str = Form('F')):
    path = save_upload_tmp(file)
    try:
        r = macros_service.find_dups(path, sheet, lookup_col)
        return r
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/cumulative-sum")
async def cumulative_sum(
    file: UploadFile = File(...), sheet: str = Form(...), start_row: int = Form(...), start_col: int = Form(...), end_row: int = Form(...)
):
    path = save_upload_tmp(file)
    try:
        r = macros_service.cumulative_sum(path, sheet, start_row, start_col, end_row)
        return r
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/remove-protection")
async def remove_protection(file: UploadFile = File(...)):
    path = save_upload_tmp(file)
    try:
        data = macros_service.remove_protection_return_bytes(path)
        return Response(content=data, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/run-pipeline")
async def run_pipeline(file: UploadFile = File(...), sheet: str = Form('Income Statement')):
    path = save_upload_tmp(file)
    try:
        r = macros_service.run_all_macros_pipeline(path, sheet)
        return r
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


@router.post("/macros/generate-budget")
async def generate_budget(
    file: UploadFile = File(...),
    growth_factor: float | None = Form(None),
    fiscal_year_start_month: int = Form(1),
    reserve_contribution: float | None = Form(None),
    template_file: UploadFile | None = File(None),
    am_seed_file: UploadFile | None = File(None),
    aliases_csv: UploadFile | None = File(None),
    enrich_only: bool = Form(False),
):
    """Run the full budget pipeline reusing existing Python code in the repo.

    This endpoint saves the uploaded files to temporary paths, calls the
    `generate_budget_pipeline.BudgetPipeline` runner, and returns JSON containing
    the enriched intermediate sheet and a small preview of the final budget workbook.
    """
    # Save uploaded input file
    input_path = save_upload_tmp(file)
    tempdir = tempfile.mkdtemp(prefix="budget_pipeline_")
    try:
        intermediate_path = os.path.join(tempdir, "Income_Statement_Enriched.xlsx")
        output_path = os.path.join(tempdir, "Budget_Pipeline.xlsx")

        # Save optional ancillary uploads
        template_path = None
        am_seed_path = None
        aliases_path = None
        if template_file is not None:
            template_path = save_upload_tmp(template_file)
        if am_seed_file is not None:
            am_seed_path = save_upload_tmp(am_seed_file)
        if aliases_csv is not None:
            aliases_path = save_upload_tmp(aliases_csv)

        # Ensure repo root is on sys.path so we can import pipeline modules
        import sys
        from pathlib import Path

        # Allow overriding repo root via settings; otherwise walk upward to find modules
        repo_root = settings.REPO_ROOT
        if not repo_root:
            curr = Path(__file__).resolve()
            repo_root = None
            for parent in curr.parents:
                if (parent / "generate_budget_pipeline.py").exists() and (parent / "generate_budget.py").exists():
                    repo_root = str(parent)
                    break
        if repo_root is None:
            raise HTTPException(status_code=500, detail="Could not locate generate_budget modules in repository ancestors")
        if repo_root not in sys.path:
            sys.path.insert(0, repo_root)

        # Import pipeline classes
        try:
            from generate_budget_pipeline import BudgetPipeline
            from generate_budget import infer_growth_factor_from_input
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to import budget pipeline: {e}")

        # Determine growth factor if not provided
        resolved_growth_factor = growth_factor
        growth_factor_note = "configured value"
        if resolved_growth_factor is None:
            try:
                resolved_growth_factor, detected_months, source = infer_growth_factor_from_input(
                    input_path, fiscal_year_start_month
                )
                growth_factor_note = f"auto annualization 12/{detected_months} from {source}"
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Could not infer growth factor: {e}")

        # Prepare pipeline and run
        pipeline = BudgetPipeline(
            input_path=input_path,
            intermediate_path=intermediate_path,
            output_path=output_path,
            growth_factor=resolved_growth_factor,
            reserve_contribution=reserve_contribution,
            template_path=template_path,
            growth_factor_note=growth_factor_note,
            am_seed_workbook=am_seed_path,
            aliases_path=aliases_path,
            enrich_only=enrich_only,
        )

        try:
            pipeline.run()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Budget pipeline failed: {e}")

        # Read enriched intermediate and budget preview
        enriched = macros_service.read_sheet_as_table(intermediate_path, "Income Statement")

        # Read budget preview from generated output: first worksheet
        from openpyxl import load_workbook
        wb = load_workbook(output_path, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            max_col = ws.max_column
            headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            rows = []
            for r in range(2, min(ws.max_row, settings.MAX_PREVIEW_ROWS) + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                if any(v is not None for v in row):
                    rows.append(row)
            budget_preview = {"sheet": ws.title, "headers": headers, "rows": rows}
        finally:
            wb.close()

        # Compose response
        resp = {
            "growth_factor": resolved_growth_factor,
            "growth_factor_note": growth_factor_note,
            "enriched": enriched,
            "budget_preview": budget_preview,
        }
        return resp
    finally:
        # Cleanup temporary files
        try:
            os.remove(input_path)
        except Exception:
            pass
        try:
            if template_path:
                os.remove(template_path)
        except Exception:
            pass
        try:
            if am_seed_path:
                os.remove(am_seed_path)
        except Exception:
            pass
        try:
            if aliases_path:
                os.remove(aliases_path)
        except Exception:
            pass
        shutil.rmtree(tempdir, ignore_errors=True)
