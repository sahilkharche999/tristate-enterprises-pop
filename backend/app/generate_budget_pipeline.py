"""Pipeline runner for enrichment (AK:AZ) and budget generation."""

import argparse
import builtins
import csv
import os
import re
import shutil
from typing import Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .generate_budget import (
    BudgetGenerator,
    infer_growth_factor_from_input,
)


QUIET_MODE = True


def _print_quiet(*args, **kwargs):
    if not QUIET_MODE:
        builtins.print(*args, **kwargs)


print = _print_quiet

DEFAULT_INPUT_PATH = "data/2024_Income_Statement.xlsx"
DEFAULT_INTERMEDIATE_PATH = "output/Income_Statement_Enriched.xlsx"
DEFAULT_OUTPUT_PATH = "output/Budget_Pipeline.xlsx"
DEFAULT_OUTPUT_DIR = "output_final"
DEFAULT_TEMPLATE_PATH = "data/Budget_Template_Reusable.xlsx"
DEFAULT_ALIASES_CSV = ""
DEFAULT_AM_SEED_WORKBOOK = ""
DEFAULT_INTERMEDIATE_FILENAME = "Income_Statement_Enriched.xlsx"
DEFAULT_OUTPUT_FILENAME = "Budget_Pipeline.xlsx"


def resolve_output_paths(
    output_dir: Optional[str],
    intermediate_path: str,
    output_path: str,
):
    """Resolve intermediate/final paths, optionally under one output directory."""
    if not output_dir:
        return intermediate_path, output_path
    return (
        os.path.join(output_dir, DEFAULT_INTERMEDIATE_FILENAME),
        os.path.join(output_dir, DEFAULT_OUTPUT_FILENAME),
    )


def _debug_enabled_for(tag: str) -> bool:
    """Return True when debug breakpoints are enabled for the given tag."""
    enabled = os.getenv("BUDGET_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    if not enabled:
        return False

    tags_raw = os.getenv("BUDGET_DEBUG_TAGS", "ALL").strip()
    if not tags_raw or tags_raw.upper() == "ALL":
        return True

    selected_tags = {t.strip() for t in tags_raw.split(",") if t.strip()}
    return tag in selected_tags


def _debug_breakpoint(tag: str, context: Optional[dict] = None):
    """
    Conditional breakpoint helper.

    Use with:
      BUDGET_DEBUG=1 BUDGET_DEBUG_TAGS=ALL
      BUDGET_DEBUG=1 BUDGET_DEBUG_TAGS=PIPELINE_UTILITY_REFUND_ROW
    """
    if not _debug_enabled_for(tag):
        return

    print(f"[DEBUG:{tag}] breakpoint hit")
    if context:
        for key, value in context.items():
            print(f"  {key}={value}")
    breakpoint()


def _load_label_aliases(path: Optional[str]) -> Dict[str, str]:
    """
    Load optional label alias CSV with columns:
      - source_label
      - template_label
    """
    if not path:
        return {}
    if not os.path.exists(path):
        raise FileNotFoundError(f"Alias file not found: {path}")

    aliases: Dict[str, str] = {}
    with open(path, newline="", encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        if not reader.fieldnames:
            return aliases
        for idx, raw in enumerate(reader, start=2):
            src = (raw.get("source_label") or "").strip()
            dst = (raw.get("template_label") or "").strip()
            if not src or not dst:
                raise ValueError(
                    f"Invalid alias row {idx} in {path}: source_label/template_label required"
                )
            aliases[src] = dst
    return aliases


class IncomeStatementEnricher:
    """
    Step 1: Copy the raw Income Statement and add calculated columns AK through AZ.

    Replicates VBA Module 6 logic:
    - AL4 = Growth Factor (configurable, e.g. 12/8=1.5 for 8 months, 12/12=1.0 for 12 months)
    - AL (Projection) = Column T (YTD Actual) × Growth Factor
    - AK (% Diff) = (Projection - Annual Budget) / Annual Budget
    - AM (% Change) = per-item budget adjustment (default 0%)
    - AN (Proposed) = Annual Budget × (1 + % Change)
    - AO-AZ (Monthly) = Proposed / 12
    """

    MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    COL_A = 1
    COL_B = 2
    COL_T = 20   # YTD Actual
    COL_AG = 33  # Annual Budget
    COL_AK = 37  # % Diff
    COL_AL = 38  # Projection
    COL_AM = 39  # % Change
    COL_AN = 40  # Proposed
    COL_AO = 41  # Jan (first monthly column)
    RESERVE_SECTION_START_TITLES = {
        "reserve income",
        "reserve expense",
        "reserve expenses (per reserve study)",
    }

    def __init__(
        self,
        input_path: str,
        output_path: str,
        growth_factor: float = None,
        growth_factor_note: Optional[str] = None,
        am_seed_workbook: Optional[str] = None,
        sheet_name: str = "Income Statement",
    ):
        self.input_path = input_path
        self.output_path = output_path
        if growth_factor is None:
            raise ValueError(
                "Growth factor is required. Pass --growth-factor or use auto-detection from input."
            )
        self.GROWTH_FACTOR = growth_factor
        self.growth_factor_note = growth_factor_note or "configured value"
        self.am_seed_workbook = am_seed_workbook
        self.sheet_name = sheet_name
        self.am_seed_by_row: Dict[int, float] = {}
        self.an_seed_by_row: Dict[int, float] = {}

    @staticmethod
    def _parse_percent_change(raw_value) -> Optional[float]:
        """Parse % Change as decimal (e.g., 1.3, 0.05, -0.4, '130%')."""
        if raw_value is None:
            return None
        if isinstance(raw_value, (int, float)):
            return float(raw_value)
        text = str(raw_value).strip()
        if not text:
            return None
        if text.endswith("%"):
            text = text[:-1].strip()
            try:
                return float(text) / 100.0
            except ValueError:
                return None
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _safe_float(value) -> float:
        """Convert cell value to float. Handles '-', None, strings."""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '').replace('$', '')
            if cleaned in ('-', '', '--'):
                return 0.0
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    @staticmethod
    def _evaluate_column_formulas(ws_formula, col, value_maps):
        """
        Evaluate formulas in a column by resolving cell references dynamically.

        Handles standard patterns found in VBA-generated Income Statements:
        - Hardcoded values: 8000, =4083.33*12
        - Cell math: =AG9*(1+AM9), =AG38*(1+AM38)+AN16
        - SUM ranges: =SUM($AN$9:$AN$18)
        - SUM single: =SUM($AN$60)
        - Additions: =AN71+AN66+AN61+...
        - Differences: =AN19-AN72

        Args:
            ws_formula: Worksheet opened with data_only=False (contains formulas).
            col: Column number to evaluate (e.g. 40 for AN).
            value_maps: Dict of {column_letter: {row: float}} for reference columns.
                        The target column (e.g. 'AN') is auto-populated as rows are
                        evaluated in order, allowing cross-references.

        Returns:
            Dict[int, float] of {row: evaluated_value}.
        """
        cell_ref_pat = re.compile(r'\$?([A-Z]{1,2})\$?(\d+)')
        sum_range_pat = re.compile(
            r'SUM\(\$?([A-Z]{1,2})\$?(\d+):\$?([A-Z]{1,2})\$?(\d+)\)'
        )
        sum_single_pat = re.compile(r'SUM\(\$?([A-Z]{1,2})\$?(\d+)\)')

        from openpyxl.utils import get_column_letter
        target_col_letter = get_column_letter(col)
        if target_col_letter not in value_maps:
            value_maps[target_col_letter] = {}
        target = value_maps[target_col_letter]

        for row in range(6, ws_formula.max_row + 1):
            raw = ws_formula.cell(row=row, column=col).value

            if raw is None:
                target[row] = None
                continue

            if isinstance(raw, (int, float)):
                target[row] = float(raw)
                continue

            formula = str(raw)
            if not formula.startswith('='):
                target[row] = 0.0
                continue

            expr = formula[1:]

            m = sum_range_pat.match(expr)
            if m:
                start_r, end_r = int(m.group(2)), int(m.group(4))
                col_letter = m.group(1)
                src = value_maps.get(col_letter, target)
                target[row] = sum(src.get(r, 0) for r in range(start_r, end_r + 1))
                continue

            m = sum_single_pat.match(expr)
            if m:
                col_letter = m.group(1)
                src = value_maps.get(col_letter, target)
                target[row] = src.get(int(m.group(2)), 0)
                continue

            def _resolve(m):
                col_letter = m.group(1)
                r = int(m.group(2))
                src = value_maps.get(col_letter, {})
                return str(src.get(r, 0))

            evaluated = cell_ref_pat.sub(_resolve, expr)
            try:
                target[row] = eval(evaluated)  # noqa: S307 — trusted workbook formulas
            except Exception:
                target[row] = 0.0

        return target

    def _load_am_seed_values(self):
        """
        Optionally load seed values from another workbook's Income Statement.

        Reads AG, AM, and evaluates AN formulas from the seed workbook.
        The AN column often contains formulas (=AG*(1+AM)) with no cached values
        in openpyxl.  Some formulas are non-standard (cross-references, hardcoded
        monthly amounts).  This method evaluates them all dynamically.

        The evaluated AN values become seeds: the enricher back-calculates
        %Change as (target_AN / raw_AG) - 1 for each item row.
        """
        self.am_seed_by_row = {}
        self.an_seed_by_row = {}
        if not self.am_seed_workbook:
            return
        if not os.path.exists(self.am_seed_workbook):
            raise FileNotFoundError(f"AM seed workbook not found: {self.am_seed_workbook}")

        wb_data = load_workbook(self.am_seed_workbook, data_only=True)
        if "Income Statement" not in wb_data.sheetnames:
            wb_data.close()
            raise ValueError(f"'Income Statement' sheet not found in {self.am_seed_workbook}")
        ws_data = wb_data["Income Statement"]

        ag_vals = {}
        am_vals = {}
        for row in range(6, ws_data.max_row + 1):
            ag_vals[row] = self._safe_float(ws_data.cell(row=row, column=self.COL_AG).value)
            parsed_am = self._parse_percent_change(ws_data.cell(row=row, column=self.COL_AM).value)
            if parsed_am is not None:
                am_vals[row] = parsed_am
                self.am_seed_by_row[row] = parsed_am
        wb_data.close()

        wb_formula = load_workbook(self.am_seed_workbook, data_only=False)
        ws_formula = wb_formula["Income Statement"]
        value_maps = {'AG': ag_vals, 'AM': am_vals}
        evaluated_an = self._evaluate_column_formulas(ws_formula, self.COL_AN, value_maps)
        wb_formula.close()

        for row, val in evaluated_an.items():
            if val is None:
                if ag_vals.get(row, 0) != 0:
                    self.an_seed_by_row[row] = 0.0
            elif val != 0:
                self.an_seed_by_row[row] = val

    def _classify_row(self, ws, row: int) -> str:
        """Classify a row as 'item', 'total', 'section', or 'empty'."""
        col_a = ws.cell(row=row, column=self.COL_A).value
        col_b = ws.cell(row=row, column=self.COL_B).value

        label = ''
        if col_b:
            label = str(col_b).strip()
        elif col_a:
            label = str(col_a).strip()

        if not label:
            return 'empty'
        if 'Total' in label:
            return 'total'
        if col_a and not col_b:
            return 'section'
        return 'item'

    def _find_section_items(self, ws, total_row: int) -> List[int]:
        """Scan upward from a total row to find all item rows in that section."""
        items = []
        for r in range(total_row - 1, 0, -1):
            row_type = self._classify_row(ws, r)
            if row_type == 'section':
                break
            if row_type == 'total':
                break
            if row_type == 'item':
                items.append(r)
        return items

    def _get_label(self, ws, row: int) -> str:
        """Get the label for a row."""
        col_b = ws.cell(row=row, column=self.COL_B).value
        col_a = ws.cell(row=row, column=self.COL_A).value
        if col_b:
            return str(col_b).strip()
        if col_a:
            return str(col_a).strip()
        return ''

    @classmethod
    def _is_reserve_section_start(cls, label: str) -> bool:
        return (label or "").strip().lower() in cls.RESERVE_SECTION_START_TITLES

    @staticmethod
    def _is_reserve_label(label: str) -> bool:
        return "reserve" in (label or "").strip().lower()

    def add_headers(self, ws):
        """Write column headers (row 5) and growth factor (AL4)."""
        ws.cell(row=4, column=self.COL_AL, value=self.GROWTH_FACTOR)
        ws.cell(row=4, column=self.COL_AL).font = Font(bold=True)

        headers = {
            self.COL_AK: '% Diff',
            self.COL_AL: 'Projection',
            self.COL_AM: '% Change',
            self.COL_AN: 'Proposed',
        }
        for col, label in headers.items():
            ws.cell(row=5, column=col, value=label)
            ws.cell(row=5, column=col).font = Font(bold=True)

        for i, month in enumerate(self.MONTHS):
            ws.cell(row=5, column=self.COL_AO + i, value=month)
            ws.cell(row=5, column=self.COL_AO + i).font = Font(bold=True)

        print("    Added headers: AK-AZ (row 5) and growth factor in AL4")

    @staticmethod
    def _ensure_min_column_width(ws, col_idx: int, min_width: float):
        letter = get_column_letter(col_idx)
        existing = ws.column_dimensions[letter].width
        if existing is None or existing < min_width:
            ws.column_dimensions[letter].width = min_width

    def _apply_output_column_widths(self, ws):
        # Prevent Excel displaying ####### for larger totals in AK:AZ.
        self._ensure_min_column_width(ws, self.COL_AK, 10)
        self._ensure_min_column_width(ws, self.COL_AL, 14)
        self._ensure_min_column_width(ws, self.COL_AM, 10)
        self._ensure_min_column_width(ws, self.COL_AN, 16)
        for col in self._month_columns():
            self._ensure_min_column_width(ws, col, 14)

    def _month_columns(self) -> range:
        return range(self.COL_AO, self.COL_AO + 12)

    def _calculated_value_columns(self) -> List[int]:
        return [self.COL_AL, self.COL_AN, *self._month_columns()]

    def _all_output_columns(self) -> List[int]:
        return [self.COL_AK, self.COL_AL, self.COL_AM, self.COL_AN, *self._month_columns()]

    @staticmethod
    def _write_number_cell(ws, row: int, col: int, value, fmt: str, bold: bool = False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)

    def process_line_items(self, ws) -> List[int]:
        """Calculate and write AK-AZ for all line item rows. Returns list of total rows found."""
        total_rows = []
        items_processed = 0
        in_reserve_block = False

        for row in range(6, ws.max_row + 1):
            row_type = self._classify_row(ws, row)
            label = self._get_label(ws, row)

            if row_type == 'section' and self._is_reserve_section_start(label):
                in_reserve_block = True

            # Reserve rows now receive enrichment calculations like operating items.

            if row_type == 'total':
                total_rows.append(row)
                continue

            if row_type != 'item':
                continue

            ytd_actual = self._safe_float(ws.cell(row=row, column=self.COL_T).value)
            annual_budget = self._safe_float(ws.cell(row=row, column=self.COL_AG).value)

            projection = ytd_actual * self.GROWTH_FACTOR

            if annual_budget != 0:
                pct_diff = (projection - annual_budget) / annual_budget
            else:
                pct_diff = 0.0

            existing_am_raw = ws.cell(row=row, column=self.COL_AM).value
            existing_am = self._parse_percent_change(existing_am_raw)
            if row in self.an_seed_by_row and annual_budget != 0:
                pct_change = (self.an_seed_by_row[row] / annual_budget) - 1.0
            elif row in self.am_seed_by_row:
                pct_change = self.am_seed_by_row[row]
            else:
                pct_change = existing_am if existing_am is not None else 0.0

            proposed = annual_budget * (1 + pct_change)
            monthly = proposed / 12 if proposed != 0 else 0.0

            if label == '41170 - Utility Refund (Conservice)':
                _debug_breakpoint("PIPELINE_UTILITY_REFUND_ROW", {
                    "row": row,
                    "label": label,
                    "ytd_actual_col_T": ytd_actual,
                    "annual_budget_col_AG": annual_budget,
                    "existing_am_cell": existing_am_raw,
                    "percent_change_col_AM": pct_change,
                    "calculated_proposed": proposed,
                    "monthly_value": monthly,
                })

            self._write_number_cell(ws, row, self.COL_AK, pct_diff, '0.00%')
            self._write_number_cell(ws, row, self.COL_AL, round(projection, 2), '#,##0.00')
            self._write_number_cell(ws, row, self.COL_AM, pct_change, '0.00%')
            self._write_number_cell(ws, row, self.COL_AN, round(proposed, 2), '#,##0.00')
            for col in self._month_columns():
                self._write_number_cell(ws, row, col, round(monthly, 2), '#,##0.00')

            items_processed += 1

        print(f"    Processed {items_processed} line item rows")
        return total_rows

    def process_total_rows(self, ws, total_rows: List[int]):
        """Calculate totals by summing section items for columns AL, AN, AO-AZ."""
        special_totals = {}

        for total_row in total_rows:
            label = self._get_label(ws, total_row)
            item_rows = self._find_section_items(ws, total_row)

            if not item_rows:
                special_totals[total_row] = label
                continue

            sum_projection = sum(self._safe_float(ws.cell(row=r, column=self.COL_AL).value) for r in item_rows)
            sum_proposed = sum(self._safe_float(ws.cell(row=r, column=self.COL_AN).value) for r in item_rows)

            self._write_number_cell(ws, total_row, self.COL_AL, round(sum_projection, 2), '#,##0.00', bold=True)
            self._write_number_cell(ws, total_row, self.COL_AN, round(sum_proposed, 2), '#,##0.00', bold=True)

            for col in self._month_columns():
                month_sum = sum(self._safe_float(ws.cell(row=r, column=col).value) for r in item_rows)
                self._write_number_cell(ws, total_row, col, round(month_sum, 2), '#,##0.00', bold=True)

            print(f"    {label}: ${sum_proposed:,.2f} (summed from {len(item_rows)} items)")
            if label == 'Total Income' and total_row == 19:
                row_breakdown = {
                    self._get_label(ws, r): self._safe_float(ws.cell(row=r, column=self.COL_AN).value)
                    for r in sorted(item_rows)
                }
                _debug_breakpoint("PIPELINE_TOTAL_INCOME_ROLLUP", {
                    "total_row": total_row,
                    "sum_proposed_col_AN": sum_proposed,
                    "monthly_sum": round(sum_proposed / 12, 2),
                    "item_rows": sorted(item_rows),
                    "row_values_col_AN": row_breakdown,
                })

        self._process_special_totals(ws, special_totals, total_rows)

    def _process_special_totals(self, ws, special_totals: dict, all_total_rows: List[int]):
        """Handle higher-level totals like Total Expense and Net Totals.
        Uses proximity: always finds the nearest matching total ABOVE each target row."""
        for total_row in sorted(special_totals.keys()):
            label = special_totals[total_row]
            computed = False

            if label == 'Total Expense':
                sub_totals = self._find_sub_total_rows_above(ws, total_row, all_total_rows)
                if sub_totals:
                    self._sum_rows(ws, total_row, sub_totals)
                    computed = True

            elif 'Net Total' in label and 'Operating' in label:
                income_row = self._find_nearest_above(ws, total_row, 'Total Income', all_total_rows)
                expense_row = self._find_nearest_above(ws, total_row, 'Total Expense', all_total_rows)
                if income_row and expense_row:
                    self._write_difference(ws, total_row, income_row, expense_row, all_total_rows)
                    computed = True

            elif 'Net Total' in label and 'Reserve' in label:
                income_row = self._find_nearest_above(ws, total_row, 'Total Income', all_total_rows)
                expense_row = self._find_nearest_above(ws, total_row, 'Total Expense', all_total_rows)
                if income_row and expense_row:
                    self._write_difference(ws, total_row, income_row, expense_row, all_total_rows)
                    computed = True

            elif label == 'Net Total':
                op_net = self._find_nearest_above(ws, total_row, 'Operating Net Total', all_total_rows)
                res_net = self._find_nearest_above(ws, total_row, 'Reserve Net Total', all_total_rows)
                if op_net and res_net:
                    self._write_sum_of_two(ws, total_row, op_net, res_net, all_total_rows)
                    computed = True

            if not computed and label in ('Total Income', 'Total Expense'):
                source = self._find_nearest_above(ws, total_row, label, all_total_rows)
                if source:
                    self._copy_total_row(ws, total_row, source)
                    computed = True

            if computed:
                print(f"    {label} (row {total_row}): ${self._safe_float(ws.cell(row=total_row, column=self.COL_AN).value):,.2f}")

    def _find_nearest_above(self, ws, target_row: int, label: str, all_total_rows: List[int]) -> Optional[int]:
        """Find the nearest total row ABOVE target_row with the given label."""
        best = None
        for row in all_total_rows:
            if row >= target_row:
                continue
            if self._get_label(ws, row) == label:
                if best is None or row > best:
                    best = row
        return best

    def _find_sub_total_rows_above(self, ws, total_expense_row: int, all_total_rows: List[int]) -> List[int]:
        """Find all sub-category total rows between a section start and the Total Expense row.
        Scans upward from Total Expense until hitting a non-total/non-item row pattern."""
        sub_totals = []
        for row in all_total_rows:
            if row >= total_expense_row:
                continue
            label = self._get_label(ws, row)
            if 'Total' in label and label != 'Total Income' and label != 'Total Expense':
                sub_totals.append(row)

        boundary = 0
        for row in all_total_rows:
            if row >= total_expense_row:
                break
            label = self._get_label(ws, row)
            if label == 'Total Income':
                boundary = row

        return [r for r in sub_totals if r > boundary]

    def _sum_rows(self, ws, target_row: int, source_rows: List[int]):
        """Sum values from source rows into the target row for all calculated columns."""
        for col in self._calculated_value_columns():
            total = sum(self._safe_float(ws.cell(row=r, column=col).value) for r in source_rows)
            self._write_number_cell(ws, target_row, col, round(total, 2), '#,##0.00', bold=True)

    def _write_difference(self, ws, target_row: int, income_row: int, expense_row: int, all_total_rows: List[int]):
        """Write income - expense for each calculated column."""
        for col in self._calculated_value_columns():
            inc_val = self._safe_float(ws.cell(row=income_row, column=col).value)
            exp_val = self._safe_float(ws.cell(row=expense_row, column=col).value)
            self._write_number_cell(ws, target_row, col, round(inc_val - exp_val, 2), '#,##0.00', bold=True)

    def _write_sum_of_two(self, ws, target_row: int, row_a: int, row_b: int, all_total_rows: List[int]):
        """Write sum of two rows for each calculated column."""
        for col in self._calculated_value_columns():
            val_a = self._safe_float(ws.cell(row=row_a, column=col).value)
            val_b = self._safe_float(ws.cell(row=row_b, column=col).value)
            self._write_number_cell(ws, target_row, col, round(val_a + val_b, 2), '#,##0.00', bold=True)

    def _copy_total_row(self, ws, target_row: int, source_row: int):
        """Copy calculated column values from one total row to another (for duplicates)."""
        for col in self._all_output_columns():
            val = ws.cell(row=source_row, column=col).value
            ws.cell(row=target_row, column=col, value=val)
            ws.cell(row=target_row, column=col).number_format = ws.cell(row=source_row, column=col).number_format
            if ws.cell(row=source_row, column=col).font.bold:
                ws.cell(row=target_row, column=col).font = Font(bold=True)

    def enrich(self) -> str:
        """
        Main method: Copy input file, add calculated columns AK-AZ, save.
        Returns the path to the enriched intermediate file.
        """
        print("=" * 70)
        print("STEP 1: CREATE INTERMEDIATE FILE (Input + Calculated Columns)")
        print("=" * 70)
        print(f"  Input:  {self.input_path}")
        print(f"  Output: {self.output_path}")
        print()

        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        shutil.copy2(self.input_path, self.output_path)
        print("  Copied input file to intermediate location")

        wb = load_workbook(self.output_path)
        ws = wb[self.sheet_name]

        print(f"  Sheet: '{self.sheet_name}' ({ws.max_row} rows, {ws.max_column} cols)")
        print(f"  Growth Factor: {self.GROWTH_FACTOR:.6f} ({self.growth_factor_note})")
        print("  Default % Change: 0.00% (used when AM is blank)")
        self._load_am_seed_values()
        if self.am_seed_workbook:
            print(
                f"  AM seed workbook: {self.am_seed_workbook} "
                f"(AM rows={len(self.am_seed_by_row)}, AN rows={len(self.an_seed_by_row)})"
            )
        print()

        self.add_headers(ws)
        self._apply_output_column_widths(ws)
        _debug_breakpoint("PIPELINE_GROWTH_FACTOR", {
            "sheet": "Income Statement",
            "configured_growth_factor": self.GROWTH_FACTOR,
            "cell_AL4": ws.cell(row=4, column=self.COL_AL).value,
        })

        print("\n  Processing line items (VBA Module 6 formulas):")
        total_rows = self.process_line_items(ws)

        print("\n  Processing total rows (VBA Budget_SumUpAll):")
        self.process_total_rows(ws, total_rows)

        wb.save(self.output_path)
        wb.close()

        print(f"\n  Intermediate file saved: {self.output_path}")
        print(f"  Columns added: AK (% Diff), AL (Projection), AM (% Change), AN (Proposed), AO-AZ (Monthly)")
        print()

        return self.output_path


class BudgetPipeline:
    """Orchestrates the 3-file pipeline: Input → Intermediate → Final Output."""

    def __init__(self, input_path: str, intermediate_path: str, output_path: str,
                 growth_factor: float = None, reserve_contribution: Optional[float] = None,
                 template_path: str = None,
                 growth_factor_note: Optional[str] = None,
                 am_seed_workbook: Optional[str] = None,
                 aliases_path: Optional[str] = None,
                 enrich_only: bool = False,
                 hoa_name: str = ''):
        self.input_path = input_path
        self.intermediate_path = intermediate_path
        self.output_path = output_path
        self.growth_factor = growth_factor
        self.growth_factor_note = growth_factor_note or "configured value"
        self.reserve_contribution = reserve_contribution
        self.template_path = template_path
        self.am_seed_workbook = am_seed_workbook
        self.aliases_path = aliases_path
        self.enrich_only = enrich_only
        self.hoa_name = hoa_name

    def run(self):
        """Execute the full pipeline."""
        self._print_header()
        self._run_enrichment_step()

        if self.enrich_only:
            self._print_enrichment_only_summary()
            return

        self._run_budget_step()
        self._print_summary()

    def _print_header(self):
        print()
        print("*" * 70)
        print("*  BUDGET PIPELINE: Input → Intermediate → Final Output")
        print("*" * 70)
        print()

    def _run_enrichment_step(self):
        enricher = IncomeStatementEnricher(
            self.input_path,
            self.intermediate_path,
            growth_factor=self.growth_factor,
            growth_factor_note=self.growth_factor_note,
            am_seed_workbook=self.am_seed_workbook,
        )
        enricher.enrich()

        intermediate_wb = load_workbook(self.intermediate_path, data_only=True)
        intermediate_ws = intermediate_wb["Income Statement"]
        _debug_breakpoint("PIPELINE_INTERMEDIATE_CHECK", {
            "intermediate_path": self.intermediate_path,
            "AL4_growth_factor": intermediate_ws["AL4"].value,
            "row16_AG_annual_budget": intermediate_ws["AG16"].value,
            "row16_AM_percent_change": intermediate_ws["AM16"].value,
            "row16_AN_proposed": intermediate_ws["AN16"].value,
            "row19_AN_total_income": intermediate_ws["AN19"].value,
        })
        intermediate_wb.close()

    def _load_aliases(self) -> Dict[str, str]:
        if not self.aliases_path:
            return {}
        aliases = _load_label_aliases(self.aliases_path)
        print(f"  Loaded template label aliases: {self.aliases_path} ({len(aliases)})")
        print()
        return aliases

    def _run_budget_step(self):
        print("=" * 70)
        print("STEP 2: GENERATE FINAL BUDGET")
        print("=" * 70)
        print(f"  Reading from: {self.intermediate_path} (column AN)")
        print(f"  Writing to:   {self.output_path}")
        print()

        generator = BudgetGenerator(
            self.intermediate_path,
            self.output_path,
            use_proposed=True,
            growth_factor=self.growth_factor,
            template_path=self.template_path,
            label_aliases=self._load_aliases(),
            hoa_name=self.hoa_name,
        )
        generator.generate(reserve_contribution=self.reserve_contribution)

    def _print_summary(self):
        """Print summary of all 3 files."""
        print("=" * 70)
        print("PIPELINE COMPLETE - 3 FILES")
        print("=" * 70)

        files = [
            ("Input (unchanged)", self.input_path),
            ("Intermediate (enriched)", self.intermediate_path),
            ("Final Output (Budget)", self.output_path),
        ]

        for label, path in files:
            exists = os.path.exists(path)
            size = os.path.getsize(path) if exists else 0
            status = f"{size:,} bytes" if exists else "MISSING"
            print(f"  {label}")
            print(f"    {path} ({status})")
            print()

        print("  Pipeline flow:")
        print(f"    {self.input_path}")
        print(f"      ↓  Step 1: Add calculated columns (VBA Module 6 logic)")
        print(f"    {self.intermediate_path}")
        print(f"      ↓  Step 2: Generate Budget (read column AN from enriched)")
        print(f"    {self.output_path}")
        print()

        print("  Note: Final output reads Proposed values from column AN of the enriched file.")
        print("  Strict macro mode: % Change is read only from AM (blank -> 0%).")
        if self.am_seed_workbook:
            print(f"  AM values seeded from workbook: {self.am_seed_workbook}")
        if self.aliases_path:
            print(f"  Template label aliases loaded from: {self.aliases_path}")
        print()

    def _print_enrichment_only_summary(self):
        """Print summary for enrich-only mode."""
        print("=" * 70)
        print("PIPELINE COMPLETE - ENRICHMENT ONLY")
        print("=" * 70)
        print("  Input (unchanged)")
        print(
            f"    {self.input_path} "
            f"({os.path.getsize(self.input_path):,} bytes)"
        )
        print()
        print("  Intermediate (enriched)")
        print(
            f"    {self.intermediate_path} "
            f"({os.path.getsize(self.intermediate_path):,} bytes)"
        )
        print()
        print("  Pipeline flow:")
        print(f"    {self.input_path}")
        print("      ↓  Step 1: Add calculated columns (VBA Module 6 logic)")
        print(f"    {self.intermediate_path}")
        print()
        print("  --enrich-only was set. Step 2 (Budget generation) was skipped.")
        print("  Strict macro mode: % Change is read only from AM (blank -> 0%).")
        if self.am_seed_workbook:
            print(f"  AM values seeded from workbook: {self.am_seed_workbook}")
        print()


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run budget pipeline (Income Statement -> Enriched -> Budget)."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT_PATH,
        help="Input Income Statement workbook path.",
    )
    parser.add_argument(
        "--intermediate",
        default=DEFAULT_INTERMEDIATE_PATH,
        help="Output path for enriched intermediate workbook.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_PATH,
        help="Output path for final budget workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help="Optional output directory. If set, writes Income_Statement_Enriched.xlsx and Budget_Pipeline.xlsx there.",
    )
    parser.add_argument(
        "--template",
        default=DEFAULT_TEMPLATE_PATH,
        help="Template workbook path for final budget layout.",
    )
    parser.add_argument(
        "--growth-factor",
        type=float,
        default=None,
        help="Growth factor used for projection. If omitted, auto-detected from input workbook month coverage.",
    )
    parser.add_argument(
        "--fiscal-year-start-month",
        type=int,
        default=1,
        help="Fiscal year start month (1=Jan ... 12=Dec) used when auto-detecting growth factor.",
    )
    parser.add_argument(
        "--reserve-contribution",
        type=float,
        default=None,
        help="Annual reserve contribution for cash flow calculation. If omitted, inferred from source.",
    )
    parser.add_argument(
        "--am-seed-workbook",
        default=DEFAULT_AM_SEED_WORKBOOK,
        help="Optional workbook path whose Income Statement!AM values seed %% Change by row.",
    )
    parser.add_argument(
        "--aliases-csv",
        default=DEFAULT_ALIASES_CSV,
        help="Optional CSV with columns: source_label,template_label for template row label differences.",
    )
    parser.add_argument(
        "--enrich-only",
        action="store_true",
        help="Run only Step 1 and save Income_Statement_Enriched.xlsx; skip final budget generation.",
    )
    return parser.parse_args()


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    args = parse_args()

    if args.growth_factor is not None and args.growth_factor <= 0:
        raise ValueError("--growth-factor must be greater than 0.")
    if not (1 <= args.fiscal_year_start_month <= 12):
        raise ValueError("--fiscal-year-start-month must be between 1 and 12.")

    if args.growth_factor is None:
        resolved_growth_factor, detected_months, source = infer_growth_factor_from_input(
            args.input,
            fiscal_year_start_month=args.fiscal_year_start_month,
        )
        growth_factor_note = (
            f"auto annualization 12/{detected_months} from {source}"
        )
    else:
        resolved_growth_factor = args.growth_factor
        approx_months = 12.0 / resolved_growth_factor
        growth_factor_note = (
            f"user override ({resolved_growth_factor:.6f}, ~{approx_months:.2f} months)"
        )

    intermediate_path, output_path = resolve_output_paths(
        args.output_dir,
        args.intermediate,
        args.output,
    )

    pipeline = BudgetPipeline(
        input_path=args.input,
        intermediate_path=intermediate_path,
        output_path=output_path,
        growth_factor=resolved_growth_factor,
        growth_factor_note=growth_factor_note,
        reserve_contribution=args.reserve_contribution,
        template_path=args.template,
        am_seed_workbook=args.am_seed_workbook,
        aliases_path=args.aliases_csv,
        enrich_only=args.enrich_only,
    )
    pipeline.run()


if __name__ == "__main__":
    main()
