"""Seed the AI pipeline SQLite database from client data files.

Run this once before starting the server:
    python -m app.ai_implementation.seed.seed_database

Data sources (all in seed/data/ folder):
    - ai_agent_casebase.json   → 18 casebase cases
    - 401 - 2025 Income Statement FINAL.xlsx  → ~17 historical cases (non-zero AM col)
    - 2024_Income_Statement.xlsx              → ~40 historical cases (Proposed != Budget)
    - ai_agent_sop_notes.txt   → 6 SOP rules
"""
import json
import logging
import sqlite3
import sys
from pathlib import Path
from typing import Any, Optional

import openpyxl

from ..database import get_db, init_db, write_lock
from ..pipeline.feature_engineering import (
    encode_account_hierarchy,
    get_seasonality_index,
    compute_adjusted_features,
)

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

DATA_DIR = Path(__file__).parent / "data"
PROPERTY_NAME = "Esprit Park Owners Association"
PRIMARY_PORTFOLIO_HOA_CODE = "1"

PORTFOLIO_SEED: list[dict[str, Any]] = [
    {
        "hoa_code": "1",
        "name": "Esprit park",
        "legacy_names": [PROPERTY_NAME],
        "units": 120,
        "tax_id": "12-3456789",
        "fiscal_year_start_month": 4,
        "fiscal_year_end_month": 3,
        "portfolio_year": 2025,
        "city": "San Francisco",
        "workflow_status": "Not Started",
    },
    {
        "hoa_code": "2",
        "name": "July Heights HOA",
        "units": 85,
        "tax_id": "98-7654321",
        "fiscal_year_start_month": 7,
        "fiscal_year_end_month": 6,
        "portfolio_year": 2025,
        "city": "Oakland",
        "workflow_status": "In Progress",
    },
    {
        "hoa_code": "3",
        "name": "October Ridge HOA",
        "units": 200,
        "tax_id": "45-6789012",
        "fiscal_year_start_month": 10,
        "fiscal_year_end_month": 9,
        "portfolio_year": 2024,
        "city": "Berkeley",
        "workflow_status": "Completed",
    },
    {
        "hoa_code": "4",
        "name": "131 Missouri",
        "units": 64,
        "tax_id": "78-9012345",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2025,
        "city": "San Francisco",
        "workflow_status": "In Progress",
    },
    {
        "hoa_code": "5",
        "name": "450 Sutter",
        "units": 150,
        "tax_id": "23-4567890",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2026,
        "city": "San Francisco",
        "workflow_status": "Not Started",
    },
    {
        "hoa_code": "6",
        "name": "880 Market",
        "units": 180,
        "tax_id": "34-5678901",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2024,
        "city": "San Francisco",
        "workflow_status": "Completed",
    },
    {
        "hoa_code": "7",
        "name": "22 Fremont",
        "units": 95,
        "tax_id": "56-7890123",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2025,
        "city": "San Jose",
        "workflow_status": "In Progress",
    },
    {
        "hoa_code": "8",
        "name": "555 Mission",
        "units": 110,
        "tax_id": "67-8901234",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2026,
        "city": "San Francisco",
        "workflow_status": "Not Started",
    },
    {
        "hoa_code": "9",
        "name": "401 HOA",
        "units": 48,
        "tax_id": "89-0123456",
        "fiscal_year_start_month": 1,
        "fiscal_year_end_month": 12,
        "portfolio_year": 2025,
        "city": "Palo Alto",
        "workflow_status": "In Progress",
    },
]


def sync_portfolio_properties(db: sqlite3.Connection) -> int:
    """Backfill the current HOA portfolio without overwriting user-edited fields."""
    seeded = 0
    for seed in PORTFOLIO_SEED:
        candidate_names = [seed["name"], *seed.get("legacy_names", [])]
        placeholders = ", ".join("?" for _ in candidate_names)
        row = db.execute(
            f"""
            SELECT id, hoa_code, name, tax_id, units, fiscal_year_start_month,
                   fiscal_year_end_month, city, portfolio_year, workflow_status
              FROM properties
             WHERE hoa_code = ?
                OR name IN ({placeholders})
             LIMIT 1
            """,
            (seed["hoa_code"], *candidate_names),
        ).fetchone()

        if row is None:
            db.execute(
                """
                INSERT INTO properties (
                    hoa_code,
                    name,
                    tax_id,
                    units,
                    fiscal_year_start_month,
                    fiscal_year_end_month,
                    city,
                    portfolio_year,
                    workflow_status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    seed["hoa_code"],
                    seed["name"],
                    seed["tax_id"],
                    seed["units"],
                    seed["fiscal_year_start_month"],
                    seed["fiscal_year_end_month"],
                    seed["city"],
                    seed["portfolio_year"],
                    seed["workflow_status"],
                ),
            )
            seeded += 1
            continue

        if not row["hoa_code"]:
            db.execute(
                """
                UPDATE properties
                   SET hoa_code = ?,
                       tax_id = ?,
                       units = ?,
                       fiscal_year_start_month = ?,
                       fiscal_year_end_month = ?,
                       city = ?,
                       portfolio_year = ?,
                       workflow_status = ?
                 WHERE id = ?
                """,
                (
                    seed["hoa_code"],
                    seed["tax_id"],
                    seed["units"],
                    seed["fiscal_year_start_month"],
                    seed["fiscal_year_end_month"],
                    seed["city"],
                    seed["portfolio_year"],
                    seed["workflow_status"],
                    row["id"],
                ),
            )
            continue

        db.execute(
            """
            UPDATE properties
               SET fiscal_year_end_month = COALESCE(fiscal_year_end_month, ?),
                   city = CASE
                     WHEN city IS NULL OR TRIM(city) = '' THEN ?
                     ELSE city
                   END,
                   portfolio_year = COALESCE(portfolio_year, ?),
                   workflow_status = CASE
                     WHEN workflow_status IS NULL OR TRIM(workflow_status) = '' THEN ?
                     ELSE workflow_status
                   END
             WHERE id = ?
            """,
            (
                seed["fiscal_year_end_month"],
                seed["city"],
                seed["portfolio_year"],
                seed["workflow_status"],
                row["id"],
            ),
        )

    db.commit()
    return seeded


def get_or_create_property(db: sqlite3.Connection) -> int:
    """Get or create the Esprit Park property. Returns property_id."""
    sync_portfolio_properties(db)
    row = db.execute(
        "SELECT id FROM properties WHERE hoa_code = ? OR name = ? LIMIT 1",
        (PRIMARY_PORTFOLIO_HOA_CODE, PROPERTY_NAME),
    ).fetchone()
    if row:
        return row[0]
    cur = db.execute(
        """
        INSERT INTO properties (
            hoa_code, name, tax_id, units, fiscal_year_start_month,
            fiscal_year_end_month, city, portfolio_year, workflow_status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            PRIMARY_PORTFOLIO_HOA_CODE,
            PROPERTY_NAME,
            "12-3456789",
            120,
            4,
            3,
            "San Francisco",
            2025,
            "Not Started",
        ),
    )
    db.commit()
    return cur.lastrowid


def create_seed_run(db: sqlite3.Connection, property_id: int, source: str, statement_month: int = 8) -> int:
    """Create a suggestion_run for seed data. Returns run_id."""
    cur = db.execute("""
        INSERT INTO suggestion_runs (property_id, source, statement_month, fiscal_year)
        VALUES (?, ?, ?, ?)
    """, (property_id, source, statement_month, 2025))
    db.commit()
    return cur.lastrowid


def compute_features(account_code: int, annual_budget: float, ytd_actual: float,
                     statement_month: int = 8) -> dict:
    """Compute all engineered features for a line item."""
    hierarchy = encode_account_hierarchy(account_code)
    level_2 = hierarchy["account_level_2"]
    seasonality_index = get_seasonality_index(level_2, statement_month)
    adjusted = compute_adjusted_features(ytd_actual, annual_budget, seasonality_index)
    return {**hierarchy, "seasonality_index": seasonality_index, **adjusted}


def insert_feedback_case(
    db: sqlite3.Connection,
    run_id: int,
    property_id: int,
    account_code: int,
    account_name: str,
    annual_budget: float,
    ytd_actual: float,
    user_final_pct_change: float,
    features: dict,
    created_at: Optional[str] = None,
    normalized_annual_budget: float = 0.5,
) -> int:
    """Insert a single feedback case. Returns case_id."""
    label = f"{account_code} - {account_name}"
    params = (
        run_id, property_id,
        account_code, account_name, label, "operating",
        features.get("account_level_1"), features.get("account_level_2"), features.get("account_level_3"),
        int(features.get("is_income", False)), int(features.get("is_reserve", False)), int(features.get("is_admin", False)),
        annual_budget, ytd_actual,
        features.get("adjusted_projection"),
        features.get("pct_diff"), features.get("coverage_ratio"),
        features.get("adjusted_pct_diff"), features.get("adjusted_coverage_ratio"),
        features.get("seasonality_index"), normalized_annual_budget,
        user_final_pct_change,
    )
    if created_at:
        cur = db.execute("""
            INSERT INTO feedback_cases (
                run_id, property_id, account_code, account_name, label, category,
                account_level_1, account_level_2, account_level_3,
                is_income, is_reserve, is_admin,
                annual_budget, ytd_actual, projection,
                pct_diff, coverage_ratio, adjusted_pct_diff, adjusted_coverage_ratio,
                seasonality_index, normalized_annual_budget,
                user_final_pct_change, user_decision, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'accepted', ?)
        """, params + (created_at,))
    else:
        cur = db.execute("""
            INSERT INTO feedback_cases (
                run_id, property_id, account_code, account_name, label, category,
                account_level_1, account_level_2, account_level_3,
                is_income, is_reserve, is_admin,
                annual_budget, ytd_actual, projection,
                pct_diff, coverage_ratio, adjusted_pct_diff, adjusted_coverage_ratio,
                seasonality_index, normalized_annual_budget,
                user_final_pct_change, user_decision
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'accepted')
        """, params)
    return cur.lastrowid


def seed_casebase(db: sqlite3.Connection, property_id: int) -> int:
    """Seed 18 cases from ai_agent_casebase.json. Returns count inserted."""
    casebase_path = DATA_DIR / "ai_agent_casebase.json"
    if not casebase_path.exists():
        logger.warning(f"Not found: {casebase_path}")
        return 0

    with open(casebase_path) as f:
        cases = json.load(f)

    run_id = create_seed_run(db, property_id, "casebase_seed", statement_month=8)
    count = 0

    # Compute max budget for normalization
    budgets = [c.get("annual_budget", 1000) for c in cases if c.get("annual_budget", 0) > 0]
    max_budget = max(budgets) if budgets else 1.0

    with write_lock():
        for case in cases:
            account_code = int(case.get("account_code", 0))
            if account_code == 0:
                continue
            account_name = case.get("label", str(account_code)).split(" - ", 1)[-1] if " - " in case.get("label", "") else case.get("label", str(account_code))
            annual_budget = float(case.get("annual_budget", 0))
            ytd_actual = float(case.get("ytd_actual", 0))
            accepted_pct = float(case.get("accepted_pct_change", 0))

            features = compute_features(account_code, annual_budget, ytd_actual)
            norm_budget = annual_budget / max_budget if max_budget > 0 else 0.5

            insert_feedback_case(db, run_id, property_id, account_code, account_name,
                                  annual_budget, ytd_actual, accepted_pct, features,
                                  normalized_annual_budget=norm_budget)
            count += 1

        db.commit()

    logger.info(f"Seeded {count} casebase cases")
    return count


def seed_feedback_jsonl(db: sqlite3.Connection, property_id: int) -> int:
    """Seed 34 cases from ai_agent_feedback.jsonl. Returns count inserted."""
    feedback_path = DATA_DIR / "ai_agent_feedback.jsonl"
    if not feedback_path.exists():
        logger.warning(f"Not found: {feedback_path}")
        return 0

    lines = [line.strip() for line in feedback_path.read_text().splitlines() if line.strip()]
    run_id = create_seed_run(db, property_id, "feedback_seed", statement_month=8)

    budgets = []
    cases = []
    for line in lines:
        try:
            c = json.loads(line)
            cases.append(c)
            budgets.append(float(c.get("annual_budget", 0)))
        except json.JSONDecodeError:
            continue

    max_budget = max(b for b in budgets if b > 0) if budgets else 1.0
    count = 0

    with write_lock():
        for case in cases:
            account_code = int(case.get("account_code", 0))
            if account_code == 0:
                continue
            label = case.get("label", str(account_code))
            account_name = label.split(" - ", 1)[-1] if " - " in label else label
            annual_budget = float(case.get("annual_budget", 0))
            ytd_actual = float(case.get("ytd_actual", 0))
            accepted_pct = float(case.get("accepted_pct_change", 0))
            ts = case.get("ts")

            features = compute_features(account_code, annual_budget, ytd_actual)
            norm_budget = annual_budget / max_budget if max_budget > 0 else 0.5

            insert_feedback_case(db, run_id, property_id, account_code, account_name,
                                  annual_budget, ytd_actual, accepted_pct, features,
                                  created_at=ts, normalized_annual_budget=norm_budget)
            count += 1

        db.commit()

    logger.info(f"Seeded {count} feedback JSONL cases")
    return count


def seed_excel_2025(db: sqlite3.Connection, property_id: int) -> int:
    """Seed ~17 cases from 401 - 2025 Income Statement FINAL.xlsx (non-zero AM column)."""
    xlsx_path = DATA_DIR / "401 - 2025 Income Statement FINAL.xlsx"
    if not xlsx_path.exists():
        logger.warning(f"Not found: {xlsx_path}")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        # Try 'Income Statement' sheet first
        sheet = wb["Income Statement"] if "Income Statement" in wb.sheetnames else wb.active

        run_id = create_seed_run(db, property_id, "historical_import", statement_month=8)
        count = 0
        rows_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Column B (index 1) = account label like "60000 - Electricity & Gas"
            # Column AM = % change column (index 38 = col AM in 0-based)
            # We need to find the right columns by header first
            label_val = row[1] if len(row) > 1 else None
            if label_val is None:
                continue

            label_str = str(label_val).strip()
            # Check if it looks like an account code line: "NNNNN - ..."
            parts = label_str.split(" - ", 1)
            if len(parts) < 2:
                continue
            try:
                account_code = int(parts[0].strip())
            except ValueError:
                continue

            account_name = parts[1].strip()

            # Find AM column - try index 38 (column AM = 1+26+13 = 39th col, 0-based=38)
            # AM is the 39th column in Excel (A=1..Z=26, AA=27..AM=39)
            # 0-based: col AM = index 38
            am_val = row[38] if len(row) > 38 else None

            if am_val is None or am_val == 0 or am_val == "" :
                continue
            try:
                pct_change = float(am_val)
                if pct_change == 0:
                    continue
            except (TypeError, ValueError):
                continue

            # Get annual budget from column AG (0-based: 32)
            budget_val = row[32] if len(row) > 32 else 0
            annual_budget = float(budget_val) if budget_val else 0.0

            # YTD actual - column T (0-based: 19)
            ytd_val = row[19] if len(row) > 19 else 0
            ytd_actual = float(ytd_val) if ytd_val else 0.0

            rows_data.append({
                "account_code": account_code,
                "account_name": account_name,
                "annual_budget": annual_budget,
                "ytd_actual": ytd_actual,
                "pct_change": pct_change,
            })

        if not rows_data:
            logger.warning("No valid rows found in 2025 Excel file, trying header detection")
            # Try to find AM column by header
            rows_data = _extract_excel_by_headers(sheet, source_year=2025)

        max_budget = max((r["annual_budget"] for r in rows_data if r["annual_budget"] > 0), default=1.0)

        with write_lock():
            for r in rows_data:
                features = compute_features(r["account_code"], r["annual_budget"], r["ytd_actual"])
                norm_budget = r["annual_budget"] / max_budget if max_budget > 0 else 0.5
                insert_feedback_case(db, run_id, property_id, r["account_code"], r["account_name"],
                                      r["annual_budget"], r["ytd_actual"], r["pct_change"],
                                      features, normalized_annual_budget=norm_budget)
                count += 1
            db.commit()

        logger.info(f"Seeded {count} 2025 Excel cases")
        return count
    finally:
        wb.close()


def seed_excel_2024(db: sqlite3.Connection, property_id: int) -> int:
    """Seed ~40 cases from 2024_Income_Statement.xlsx (Proposed != Budget)."""
    xlsx_path = DATA_DIR / "2024_Income_Statement.xlsx"
    if not xlsx_path.exists():
        logger.warning(f"Not found: {xlsx_path}")
        return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        sheet = wb["Income Statement"] if "Income Statement" in wb.sheetnames else wb.active

        run_id = create_seed_run(db, property_id, "historical_import", statement_month=12)
        count = 0
        rows_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            label_val = row[1] if len(row) > 1 else None
            if label_val is None:
                continue

            label_str = str(label_val).strip()
            parts = label_str.split(" - ", 1)
            if len(parts) < 2:
                continue
            try:
                account_code = int(parts[0].strip())
            except ValueError:
                continue

            account_name = parts[1].strip()

            # T = YTD Actual (col T = index 19)
            ytd_val = row[19] if len(row) > 19 else 0
            ytd_actual = float(ytd_val) if ytd_val else 0.0

            # AG = Annual Budget (col AG = index 32)
            budget_val = row[32] if len(row) > 32 else 0
            annual_budget = float(budget_val) if budget_val else 0.0

            # AN = Proposed (col AN = index 39)
            proposed_val = row[39] if len(row) > 39 else None

            if proposed_val is None:
                continue
            try:
                proposed = float(proposed_val)
            except (TypeError, ValueError):
                continue

            if annual_budget == 0:
                continue

            if abs(proposed - annual_budget) < 0.01:
                continue  # Proposed == Budget, skip

            pct_change = (proposed - annual_budget) / annual_budget
            # Clamp to reasonable range
            if abs(pct_change) > 1.0:
                continue

            rows_data.append({
                "account_code": account_code,
                "account_name": account_name,
                "annual_budget": annual_budget,
                "ytd_actual": ytd_actual,
                "pct_change": pct_change,
            })

        if not rows_data:
            logger.warning("No valid rows found in 2024 Excel, trying header detection")
            rows_data = _extract_excel_by_headers(sheet, source_year=2024)

        max_budget = max((r["annual_budget"] for r in rows_data if r["annual_budget"] > 0), default=1.0)

        with write_lock():
            for r in rows_data:
                features = compute_features(r["account_code"], r["annual_budget"], r["ytd_actual"])
                norm_budget = r["annual_budget"] / max_budget if max_budget > 0 else 0.5
                insert_feedback_case(db, run_id, property_id, r["account_code"], r["account_name"],
                                      r["annual_budget"], r["ytd_actual"], r["pct_change"],
                                      features, normalized_annual_budget=norm_budget)
                count += 1
            db.commit()

        logger.info(f"Seeded {count} 2024 Excel cases")
        return count
    finally:
        wb.close()


def _extract_excel_by_headers(sheet, source_year: int) -> list[dict]:
    """Fallback: detect column positions from header row."""
    rows_data = []
    header_row = None

    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(str(c).upper() in ("ACCOUNT", "LABEL", "DESCRIPTION") for c in row if c):
            header_row = [str(c).strip().upper() if c else "" for c in row]
            break

    if not header_row:
        return rows_data

    col = {h: i for i, h in enumerate(header_row)}

    def get_col(*keys):
        for k in keys:
            if k in col:
                return col[k]
        return None

    label_idx = get_col("LABEL", "DESCRIPTION", "ACCOUNT")
    budget_idx = get_col("ANNUAL BUDGET", "BUDGET", "AG")
    ytd_idx = get_col("YTD ACTUAL", "YTD", "T")
    proposed_idx = get_col("PROPOSED", "AN") if source_year == 2024 else None
    pct_idx = get_col("% CHANGE", "AM", "PCT CHANGE") if source_year == 2025 else None

    if label_idx is None:
        return rows_data

    for row in sheet.iter_rows(min_row=2, values_only=True):
        label_val = row[label_idx] if label_idx is not None and len(row) > label_idx else None
        if not label_val:
            continue
        parts = str(label_val).split(" - ", 1)
        if len(parts) < 2:
            continue
        try:
            ac = int(parts[0].strip())
        except ValueError:
            continue

        budget = float(row[budget_idx]) if budget_idx is not None and len(row) > budget_idx and row[budget_idx] else 0.0
        ytd = float(row[ytd_idx]) if ytd_idx is not None and len(row) > ytd_idx and row[ytd_idx] else 0.0

        if source_year == 2025 and pct_idx is not None:
            pct_val = row[pct_idx] if len(row) > pct_idx else None
            if not pct_val or pct_val == 0:
                continue
            try:
                pct = float(pct_val)
            except (TypeError, ValueError):
                continue
        elif source_year == 2024 and proposed_idx is not None:
            prop_val = row[proposed_idx] if len(row) > proposed_idx else None
            if not prop_val or budget == 0:
                continue
            try:
                prop = float(prop_val)
                pct = (prop - budget) / budget
                if abs(pct) < 0.001 or abs(pct) > 1.0:
                    continue
            except (TypeError, ValueError):
                continue
        else:
            continue

        rows_data.append({
            "account_code": ac,
            "account_name": parts[1].strip(),
            "annual_budget": budget,
            "ytd_actual": ytd,
            "pct_change": pct,
        })

    return rows_data


def seed_sop_rules(db: sqlite3.Connection) -> int:
    """Seed 6 SOP rules from ai_agent_sop_notes.txt."""
    sop_path = DATA_DIR / "ai_agent_sop_notes.txt"
    if not sop_path.exists():
        logger.warning(f"Not found: {sop_path}")
        return 0

    rules = [line.strip() for line in sop_path.read_text().splitlines() if line.strip()]

    with write_lock():
        # Check if already seeded
        existing = db.execute("SELECT COUNT(*) FROM sop_rules").fetchone()[0]
        if existing > 0:
            logger.info(f"SOP rules already seeded ({existing} rules), skipping")
            return existing

        for rule in rules:
            db.execute("INSERT INTO sop_rules (rule_text) VALUES (?)", (rule,))
        db.commit()

    logger.info(f"Seeded {len(rules)} SOP rules")
    return len(rules)


def run_seed(force: bool = False) -> dict:
    """Run the full seed operation. Returns counts per source."""
    init_db()
    db = get_db()

    try:
        portfolio_seeded = sync_portfolio_properties(db)

        # Check if already seeded
        if not force:
            existing = db.execute(
                "SELECT COUNT(*) FROM feedback_cases WHERE user_decision = 'accepted'"
            ).fetchone()[0]
            if existing >= 50:
                logger.info(f"Database already has {existing} accepted cases, skipping seed. Use force=True to re-seed.")
                return {"existing": existing, "portfolio_seeded": portfolio_seeded, "skipped": True}

        property_id = get_or_create_property(db)
        logger.info(f"Property ID: {property_id}")

        counts = {
            "portfolio_seeded": portfolio_seeded,
            "sop_rules": seed_sop_rules(db),
            "casebase": seed_casebase(db, property_id),
            "excel_2025": seed_excel_2025(db, property_id),
            "excel_2024": seed_excel_2024(db, property_id),
        }

        total = db.execute("SELECT COUNT(*) FROM feedback_cases WHERE user_decision = 'accepted'").fetchone()[0]
        counts["total_accepted"] = total
        logger.info(f"Seed complete. Total accepted cases: {total}")
        logger.info(f"Counts: {counts}")

        return counts

    finally:
        db.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Seed the AI pipeline database")
    parser.add_argument("--force", action="store_true", help="Re-seed even if data exists")
    args = parser.parse_args()
    result = run_seed(force=args.force)
    print(f"\nSeed result: {result}")
