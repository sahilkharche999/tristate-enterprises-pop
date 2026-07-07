"""Tests for GET /hoa/{hoa_id}/budget/uploads/{upload_id}/income-statement-file-html.

Renders every sheet of an Excel-family income-statement upload as a plain HTML
table (no header-row detection, no cell/page jump); backs the Enriched tab's
"Compare with source" split view for Excel-sourced drafts (change:
add-income-statement-pdf-compare-view). Mirrors the cross-HOA/404/auth test
shape already established in test_income_statement_file_endpoint.py for the
sibling PDF file-serving route.
"""

from __future__ import annotations

from io import BytesIO


def _xlsx_bytes(sheet_rows: dict) -> bytes:
    """Build real .xlsx bytes with one sheet per key, rows as given."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheet_rows.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _xls_bytes(sheet_name: str, rows: list) -> bytes:
    """Build real legacy .xls bytes (BIFF format) via xlwt."""
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload_excel_income_statement(
    client, *, filename: str, content_type: str, file_bytes: bytes, hoa_id: int = 1
) -> int:
    response = client.post(
        f"/hoa/{hoa_id}/budget/upload",
        data={"source_mode": "income_statement"},
        files={"file": (filename, file_bytes, content_type)},
    )
    assert response.status_code == 200, response.text
    upload_id = response.json()["upload_id"]
    assert upload_id is not None
    return upload_id


_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS_CONTENT_TYPE = "application/vnd.ms-excel"


def test_get_income_statement_html_renders_all_sheets(client, budget_history_test_harness):
    file_bytes = _xlsx_bytes(
        {
            "Income Statement": [["Account", "Label", "Amount"], ["40000", "Assessment Income", 150000]],
            "Notes": [["This is a note sheet"]],
        }
    )
    upload_id = _upload_excel_income_statement(
        client, filename="income.xlsx", content_type=_XLSX_CONTENT_TYPE, file_bytes=file_bytes
    )

    response = client.get(f"/hoa/1/budget/uploads/{upload_id}/income-statement-file-html")

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("text/html")
    body = response.text
    assert "Assessment Income" in body
    assert "This is a note sheet" in body
    assert "Income Statement" in body
    assert "Notes" in body


def test_get_income_statement_html_renders_legacy_xls(client, budget_history_test_harness):
    # Deliberately NOT named "Income Statement" — _ensure_xlsx's .xls conversion
    # falls back to sheet_by_index(0) when no sheet matches that literal name;
    # a fixture that happens to be named "Income Statement" would hide a bug
    # where any differently-named real-world .xls sheet renders blank.
    file_bytes = _xls_bytes("Q1 Financials", [["Account", "Label", "Amount"], ["40000", "Assessment Income", 150000]])
    upload_id = _upload_excel_income_statement(
        client, filename="income.xls", content_type=_XLS_CONTENT_TYPE, file_bytes=file_bytes
    )

    response = client.get(f"/hoa/1/budget/uploads/{upload_id}/income-statement-file-html")

    assert response.status_code == 200, response.text
    assert "Assessment Income" in response.text


def test_get_income_statement_html_escapes_cell_values(client, budget_history_test_harness):
    file_bytes = _xlsx_bytes({"Sheet1": [["<script>alert(1)</script>"]]})
    upload_id = _upload_excel_income_statement(
        client, filename="income.xlsx", content_type=_XLSX_CONTENT_TYPE, file_bytes=file_bytes
    )

    response = client.get(f"/hoa/1/budget/uploads/{upload_id}/income-statement-file-html")

    assert response.status_code == 200, response.text
    assert "<script>alert(1)</script>" not in response.text
    assert "&lt;script&gt;" in response.text


def test_get_income_statement_html_404_when_upload_belongs_to_other_hoa(
    client, db_session, budget_history_test_harness
):
    from app.ai_implementation.db.models import Property

    file_bytes = _xlsx_bytes({"Sheet1": [["a"]]})
    upload_id = _upload_excel_income_statement(
        client, filename="income.xlsx", content_type=_XLSX_CONTENT_TYPE, file_bytes=file_bytes
    )

    other = Property(name="Different HOA", units=20, hoa_code="X4")
    db_session.add(other)
    db_session.commit()
    db_session.refresh(other)

    response = client.get(f"/hoa/{other.id}/budget/uploads/{upload_id}/income-statement-file-html")

    assert response.status_code == 404


def test_get_income_statement_html_404_when_file_missing_from_disk(client, db_session, budget_history_test_harness):
    from app.ai_implementation.db.models import BudgetUpload
    from app.services.budget_history_service import _budget_storage_path

    file_bytes = _xlsx_bytes({"Sheet1": [["a"]]})
    upload_id = _upload_excel_income_statement(
        client, filename="income.xlsx", content_type=_XLSX_CONTENT_TYPE, file_bytes=file_bytes
    )

    upload = db_session.get(BudgetUpload, upload_id)
    _budget_storage_path(upload.storage_key).unlink()

    response = client.get(f"/hoa/1/budget/uploads/{upload_id}/income-statement-file-html")

    assert response.status_code == 404


def test_get_income_statement_html_requires_auth(client, budget_history_test_harness):
    file_bytes = _xlsx_bytes({"Sheet1": [["a"]]})
    upload_id = _upload_excel_income_statement(
        client, filename="income.xlsx", content_type=_XLSX_CONTENT_TYPE, file_bytes=file_bytes
    )

    response = client.get(
        f"/hoa/1/budget/uploads/{upload_id}/income-statement-file-html",
        headers={"Authorization": ""},
    )

    assert response.status_code in (401, 403)
