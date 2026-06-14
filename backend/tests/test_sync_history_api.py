import sqlite3
import json
from pathlib import Path

from app.ai_implementation.db import BUDGET_DRAFT_ACTIVE, BudgetDraft, BudgetUpload
from app.models.financial_document_extraction import ExtractedFinancialStatement
from app.services import macros_service
from app.services.budget_history_service import _table_to_line_items


def _seed_assessment_rule_for_history(db_session) -> int:
    raw = db_session.connection().connection
    raw.execute(
        """
        INSERT INTO assessment_setups
            (property_id, setup_type, display_mode, status, approved_at)
        VALUES
            (1, 'grouped', 'grouped', 'approved', datetime('now'))
        """
    )
    setup_id = raw.execute("SELECT last_insert_rowid()").fetchone()[0]
    raw.execute(
        """
        INSERT INTO allocation_pools
            (assessment_setup_id, pool_key, pool_name, allocation_method,
             recipient_scope)
        VALUES
            (?, 'total_budget_prorated', 'Prorated', 'square_footage',
             'all_units')
        """,
        (setup_id,),
    )
    raw.execute(
        """
        INSERT INTO assessment_budget_mapping_rules
            (property_id, assessment_setup_id, pool_key, match_label,
             normalized_label, account_code, match_type, rule_source,
             approval_status, review_state)
        VALUES
            (1, ?, 'total_budget_prorated', 'Insurance', 'insurance',
             '6100', 'account_code', 'operator', 'approved', 'ready')
        """,
        (setup_id,),
    )
    raw.execute(
        "UPDATE properties SET default_assessment_setup_id = ? WHERE id = 1",
        (setup_id,),
    )
    db_session.commit()
    return setup_id


def _upload_income_statement(client, *, source_mode: str = "income_statement"):
    return client.post(
        "/hoa/1/budget/upload",
        data={"source_mode": source_mode},
        files={
            "file": (
                "income-statement.xlsx",
                b"placeholder workbook bytes",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def _generate_version(client, draft, *, global_note="Generated for test"):
    return client.post(
        "/hoa/1/budget/generate",
        json={
            "draft_id": draft["id"],
            "line_items": draft["line_items"],
            "global_note": global_note,
        },
    )


def _raw_and_short_label_duplicates() -> list[dict[str, object]]:
    return [
        {
            "line_item_key": "55000",
            "account_code": 55000,
            "category": "operating",
            "label": "55000 - General Insurance",
            "annual_budget": 15000.0,
            "projection": 15000.0,
            "percent_change": 0.0,
            "read_only": False,
            "raw": {"section": "Operating Expenses"},
        },
        {
            "line_item_key": "Insurance",
            "account_code": None,
            "category": "operating",
            "label": "Insurance",
            "annual_budget": 15000.0,
            "projection": 15000.0,
            "percent_change": 0.0,
            "read_only": False,
            "raw": {"section": "Operating Expenses"},
        },
    ]


def test_write_percent_changes_matches_normalized_pdf_label_column(tmp_path):
    from openpyxl import Workbook, load_workbook

    path = tmp_path / "normalized.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Income Statement"
    worksheet.append(["Section", "Account Code", "Label", "% Change"])
    worksheet.append(["income", "40000", "Assessment Income", None])
    workbook.save(path)
    workbook.close()

    matched = macros_service.write_percent_changes_by_label(
        str(path),
        "Income Statement",
        {"Assessment Income": 0.05},
        pct_change_col=4,
    )

    workbook = load_workbook(path, data_only=True)
    try:
        assert matched == 1
        assert workbook["Income Statement"].cell(row=2, column=4).value == 0.05
    finally:
        workbook.close()


def _sheet_row(*values, width: int = 39):
    row = [None] * width
    for index, value in enumerate(values):
        row[index] = value
    return row


def test_table_to_line_items_supports_headerless_income_statement_layout():
    table = {
        "headers": ["Esprit Park Owners Association"] + [None] * 38,
        "rows": [
            _sheet_row("Statement of Revenues and Expenses 8/1/2025 - 8/31/2025"),
            _sheet_row(),
            _sheet_row(None, None, None, None, None, None, "Current Period"),
            _sheet_row(None, None, None, "Actual", None, None, None, None, None, None, "Budget"),
            _sheet_row("Operating Income"),
            _sheet_row(None, "41170 - Utility Refund (Conservice)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 24340.0, None, None, None, None, None, None, None, None, None, None, None, None, 50000.0, None, None, None, None, 36510.0, 0),
            _sheet_row(None, "40000 - Assessment Income", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 765386.88, None, None, None, None, None, None, None, None, None, None, None, None, 1148080.32, None, None, None, None, 1148080.32, 0),
            _sheet_row(None, "51610 - Bank Charges", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 30.0, None, None, None, None, None, None, None, None, None, None, None, None, "-", None, None, None, None, 45.0, 0),
            _sheet_row("Total Administration Expenses", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 150497.98, None, None, None, None, None, None, None, None, None, None, None, None, 287748, None, None, None, None, 223346.22, None),
            _sheet_row("Allocation to Reserves"),
            _sheet_row(None, "90000 - Reserve - Allocation/Transfer", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 177478.0, None, None, None, None, None, None, None, None, None, None, None, None, 266217.0, None, None, None, None, None, None),
            _sheet_row("Reserve Income"),
            _sheet_row(None, "45000 - Reserve Income", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 177478.0, None, None, None, None, None, None, None, None, None, None, None, None, 266217.0, None, None, None, None, None, None),
            _sheet_row(None, "47001 - Change in Asset Value", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 806.84, None, None, None, None, None, None, None, None, None, None, None, None, 7000.0, None, None, None, None, None, None),
            _sheet_row("Reserve Expenses (Per Reserve Study)"),
            _sheet_row(None, "95220 - Roof", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 1500.0, None, None, None, None, None, None, None, None, None, None, None, None, 3000.0, None, None, None, None, 2250.0, 0),
        ],
    }

    line_items, _warnings = _table_to_line_items(table)

    assert [item["label"] for item in line_items[:2]] == [
        "41170 - Utility Refund (Conservice)",
        "40000 - Assessment Income",
    ]
    utility_refund = next(item for item in line_items if item["label"] == "41170 - Utility Refund (Conservice)")
    assert utility_refund["category"] == "income"
    assert utility_refund["annual_budget"] == 50000.0
    assert utility_refund["read_only"] is False
    assert [item["label"] for item in line_items[1:3]] == [
        "40000 - Assessment Income",
        "51610 - Bank Charges",
    ]
    assessment_income = next(item for item in line_items if item["label"] == "40000 - Assessment Income")
    assert assessment_income["account_code"] == 40000
    assert assessment_income["ytd_actual"] == 765386.88
    assert assessment_income["annual_budget"] == 1148080.32
    assert assessment_income["projection"] == 1148080.32
    assert assessment_income["percent_change"] == 0
    assert assessment_income["read_only"] is False
    # Section-based classification: "Allocation to Reserves" is not a recognized section header,
    # so this item inherits the current section state ("income"). Not read_only (editable).
    reserve_transfer = next(item for item in line_items if item["label"] == "90000 - Reserve - Allocation/Transfer")
    assert reserve_transfer["read_only"] is False
    # "Reserve Income" section → category=reserve, read_only=True (Reserve
    # Income items come from the reserve study extraction, not editable
    # budget input). See READ_ONLY_SECTIONS in income_statement_parser.py.
    reserve_income = next(item for item in line_items if item["label"] == "45000 - Reserve Income")
    assert reserve_income["category"] in ("reserve", "reserve_income", "reserve_expense")
    assert reserve_income["read_only"] is True
    reserve_asset_value = next(item for item in line_items if item["label"] == "47001 - Change in Asset Value")
    assert reserve_asset_value["category"] in ("reserve", "reserve_income", "reserve_expense")
    assert reserve_asset_value["read_only"] is True
    assert all(item["label"] != "Total Administration Expenses" for item in line_items)
    # "Reserve Expenses (Per Reserve Study)" triggers read_only=True for items in that sub-section
    reserve_item = next(item for item in line_items if item["label"] == "95220 - Roof")
    assert reserve_item["read_only"] is True
    assert reserve_item["category"] in ("reserve", "reserve_income", "reserve_expense")


def test_table_to_line_items_marks_reserve_expense_block_rows_as_reserve_components():
    table = {
        "headers": ["Income Statement"] + [None] * 38,
        "rows": [
            _sheet_row("Reserve Expenses (Per Reserve Study)"),
            _sheet_row(None, "91432 - Circulation Pump 5 H.P.", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 0.0, None, None, None, None, None, None, None, None, None, None, None, None, 109008.0, None, None, None, None, None, None),
            _sheet_row(None, "95220 - Roof", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 217758.0, None, None, None, None, None, None, None, None, None, None, None, None, 0.0, None, None, None, None, None, None),
            _sheet_row(None, "47000 - Interest Earned Reserve", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 6322.0, None, None, None, None, None, None, None, None, None, None, None, None, 8600.0, None, None, None, None, None, None),
        ],
    }

    line_items, _warnings = _table_to_line_items(table)

    circulation_pump = next(item for item in line_items if item["label"] == "91432 - Circulation Pump 5 H.P.")
    roof = next(item for item in line_items if item["label"] == "95220 - Roof")
    reserve_interest = next(item for item in line_items if item["label"] == "47000 - Interest Earned Reserve")

    assert circulation_pump["category"] in ("reserve", "reserve_income", "reserve_expense")
    assert circulation_pump["read_only"] is True
    assert roof["category"] in ("reserve", "reserve_income", "reserve_expense")
    assert roof["read_only"] is True
    assert reserve_interest["category"] in ("reserve", "reserve_income", "reserve_expense")
    assert reserve_interest["read_only"] is True


def test_upload_creates_history(client, budget_history_test_harness):
    response = _upload_income_statement(client)

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["upload_id"] > 0
    assert payload["draft"]["source_upload_id"] == payload["upload_id"]
    assert payload["draft"]["status"] == "active"
    assert payload["draft"]["enriched_file_available"] is True
    assert payload["timeline_event"]["event_type"] == "upload_received"

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()

    assert history_payload["active_draft"]["id"] == payload["draft"]["id"]
    assert history_payload["drafts"][0]["id"] == payload["draft"]["id"]
    assert history_payload["drafts"][0]["enriched_file_available"] is True
    assert {event["event_type"] for event in history_payload["timeline"]} >= {
        "upload_received",
        "enrichment_completed",
    }
    retained_files = list(Path(budget_history_test_harness["storage_root"]).rglob("*"))
    assert any(path.is_file() for path in retained_files)


def test_proforma_source_mode_persists_through_history_version_and_reopen(
    client,
    budget_history_test_harness,
    monkeypatch,
):
    monkeypatch.setattr(
        "app.services.budget_history_service._extract_proforma_excel_statement_sync",
        lambda *args, **kwargs: ExtractedFinancialStatement.model_validate(
            {
                "document_family": "excel_budget_workbook",
                "report_type": "income_statement",
                "line_items": [
                    {
                        "account_code_text": "40000",
                        "label": "Assessments",
                        "section_kind": "income",
                        "annual_budget": 1148083.2,
                        "evidence": {"source_column": "annual_budget"},
                    },
                    {
                        "account_code_text": "50050",
                        "label": "Management Service Contract",
                        "section_kind": "operating",
                        "annual_budget": 61740.0,
                        "evidence": {"source_column": "annual_budget"},
                    },
                ],
                "confidence": 0.9,
            }
        ),
        raising=False,
    )

    upload_response = client.post(
        "/hoa/1/budget/upload",
        data={"source_mode": "proforma_final_budget"},
        files={
            "file": (
                "401 Esprit Park 2025 Budget.xlsx",
                b"placeholder workbook bytes",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 200, upload_response.text
    upload_payload = upload_response.json()
    assert upload_payload["draft"]["source_mode"] == "proforma_final_budget"

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert history_payload["active_draft"]["source_mode"] == "proforma_final_budget"
    assert history_payload["drafts"][0]["source_mode"] == "proforma_final_budget"

    draft = upload_payload["draft"]
    generate_response = _generate_version(client, draft, global_note="Pro forma mode version")
    assert generate_response.status_code == 200, generate_response.text
    version_payload = generate_response.json()["version"]
    assert version_payload["source_mode"] == "proforma_final_budget"

    version_detail_response = client.get(f"/hoa/1/versions/{version_payload['id']}")
    assert version_detail_response.status_code == 200
    assert version_detail_response.json()["source_mode"] == "proforma_final_budget"

    reopen_response = client.post(f"/hoa/1/versions/{version_payload['id']}/reopen")
    assert reopen_response.status_code == 200
    assert reopen_response.json()["draft"]["source_mode"] == "proforma_final_budget"


def test_generate_creates_version_snapshot(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    upload_payload = upload_response.json()
    draft = upload_payload["draft"]

    generate_response = _generate_version(
        client,
        draft,
        global_note="Initial board-review draft",
    )

    assert generate_response.status_code == 200
    payload = generate_response.json()
    assert payload["version"]["version_number"] == 1
    assert payload["version"]["version_code"] == "V1"
    assert payload["version"]["stage"] == "Interim"
    assert payload["version"]["growth_factor"] == budget_history_test_harness["expected_growth_factor"]
    assert payload["timeline_event"]["event_type"] == "budget_generated"

    version_response = client.get(f"/hoa/1/versions/{payload['version']['id']}")
    assert version_response.status_code == 200
    version_payload = version_response.json()
    assert version_payload["id"] == payload["version"]["id"]
    assert version_payload["budget_preview"]["sheet"] == "Budget"
    assert len(version_payload["line_items"]) == len(draft["line_items"])

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert history_payload["versions"][0]["version_code"] == "V1"
    assert {event["event_type"] for event in history_payload["timeline"]} >= {"budget_generated"}


def test_generate_materializes_assessment_mappings_before_version_snapshot(
    client,
    budget_history_test_harness,
    db_session,
):
    setup_id = _seed_assessment_rule_for_history(db_session)
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]
    # Simulate a stale draft state where the annual mapping was removed
    # before generation. Generation should re-materialize synchronously.
    db_session.connection().connection.execute(
        "DELETE FROM budget_line_pool_mappings WHERE assessment_setup_id = ?",
        (setup_id,),
    )
    db_session.commit()

    generate_response = _generate_version(client, draft)

    assert generate_response.status_code == 200, generate_response.text
    rows = db_session.connection().connection.execute(
        """
        SELECT budget_line_normalized_label, pool_key, mapping_source
          FROM budget_line_pool_mappings
         WHERE assessment_setup_id = ?
        """,
        (setup_id,),
    ).fetchall()
    assert rows == [("insurance", "total_budget_prorated", "account_code")]


def test_note_save_and_manual_override_audit(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    updated_line_items = [
        {
            **draft["line_items"][0],
            "proposed_amount": 129500.0,
            "manual_override": True,
        },
        *draft["line_items"][1:],
    ]

    draft_response = client.patch(
        "/hoa/1/budget/draft",
        json={
            "draft_id": draft["id"],
            "line_items": updated_line_items,
            "global_note": "Reviewed with updated assessment assumption",
            "statement_month": 12,
            "growth_factor": budget_history_test_harness["expected_growth_factor"],
            "growth_factor_note": "manual confirmation",
        },
    )

    assert draft_response.status_code == 200
    draft_payload = draft_response.json()
    assert draft_payload["draft"]["id"] == draft["id"]
    assert draft_payload["timeline_event"]["event_type"] == "manual_overrides_saved"

    note_response = client.post(
        "/hoa/1/budget/notes",
        json={
            "draft_id": draft["id"],
            "note_scope": "line_item",
            "line_item_key": str(updated_line_items[0].get("line_item_key") or updated_line_items[0].get("label")),
            "title": "Insurance renewal",
            "body": "Carrier communicated a likely 10 percent increase.",
        },
    )

    assert note_response.status_code == 200
    note_payload = note_response.json()
    assert note_payload["note"]["draft_id"] == draft["id"]
    assert note_payload["timeline_event"]["event_type"] == "note_saved"

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert {event["event_type"] for event in history_payload["timeline"]} >= {
        "manual_overrides_saved",
        "note_saved",
    }
    assert history_payload["notes"][0]["title"] == "Insurance renewal"


def test_history_screen_payload_shape(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    note_response = client.post(
        "/hoa/1/budget/notes",
        json={
            "draft_id": draft["id"],
            "note_scope": "global",
            "title": "Fiscal strategy",
            "body": "Board requested conservative operating assumptions.",
        },
    )
    assert note_response.status_code == 200

    generate_response = _generate_version(
        client,
        draft,
        global_note="History payload smoke coverage",
    )
    assert generate_response.status_code == 200
    version_id = generate_response.json()["version"]["id"]

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    payload = history_response.json()

    assert set(payload) == {"active_draft", "timeline", "versions", "notes", "drafts"}
    assert isinstance(payload["timeline"], list)
    assert isinstance(payload["drafts"], list)
    assert isinstance(payload["versions"], list)
    assert isinstance(payload["notes"], list)
    assert payload["notes"][0]["note_scope"] == "global"
    assert payload["versions"][0]["id"] == version_id
    assert payload["active_draft"] is None or payload["active_draft"]["id"] == draft["id"]


def test_generate_version_uses_draft_percent_changes_in_persisted_output(
    client,
    budget_history_test_harness,
    monkeypatch,
):
    preview_by_output_path = {}
    changes_by_input_path = {}

    class PercentAwarePipeline:
        def __init__(self, **kwargs):
            self.input_path = kwargs["input_path"]
            self.intermediate_path = kwargs["intermediate_path"]
            self.output_path = kwargs["output_path"]

        def run(self):
            Path(self.intermediate_path).write_bytes(b"fake enriched workbook")
            changes = changes_by_input_path.get(self.input_path, {})
            total_income = 126000.0
            total_expense = 35200.0 + round(sum(changes.values()) * 1000, 2)
            preview_by_output_path[self.output_path] = {
                "sheet": "Budget",
                "headers": ["Account Code", "Line Item", "YTD Actual", "Annual Budget",
                             "% Change", "Proposed Budget", "Monthly", "Notes"],
                "rows": [
                    ["", "Total Income", None, total_income, None, total_income, round(total_income / 12, 2), None],
                    ["", "Total Expense", None, total_expense, None, total_expense, round(total_expense / 12, 2), None],
                    ["", "Net Operating Income", None, total_income - total_expense, None, total_income - total_expense, round((total_income - total_expense) / 12, 2), None],
                ],
            }
            Path(self.output_path).write_text(str(total_expense), encoding="utf-8")

    def percent_aware_preview(path: str, max_rows: int):
        return preview_by_output_path[path]

    def capture_percent_changes(path: str, sheet: str, changes):
        changes_by_input_path[path] = dict(changes)

    monkeypatch.setattr(
        "app.services.budget_history_service.BudgetPipeline",
        PercentAwarePipeline,
    )
    monkeypatch.setattr(
        "app.services.budget_history_service.macros_service.read_first_sheet_preview",
        percent_aware_preview,
    )
    monkeypatch.setattr(
        "app.services.budget_history_service.macros_service.write_percent_changes_by_label",
        capture_percent_changes,
    )

    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    baseline_generate = _generate_version(client, draft, global_note="Baseline")
    assert baseline_generate.status_code == 200
    baseline_version = baseline_generate.json()["version"]
    baseline_download = client.post(f"/hoa/1/versions/{baseline_version['id']}/download")
    assert baseline_download.status_code == 200

    second_upload = _upload_income_statement(client)
    assert second_upload.status_code == 200
    second_draft = second_upload.json()["draft"]
    edited_items = [{**item} for item in second_draft["line_items"]]
    edited_items[1]["percent_change"] = 250

    save_response = client.patch(
        "/hoa/1/budget/draft",
        json={
            "draft_id": second_draft["id"],
            "line_items": edited_items,
            "global_note": "Edited percent change",
            "statement_month": second_draft["statement_month"],
            "growth_factor": second_draft["growth_factor"],
            "growth_factor_note": second_draft["growth_factor_note"],
        },
    )
    assert save_response.status_code == 200

    edited_generate = client.post(
        "/hoa/1/budget/generate",
        json={
            "draft_id": second_draft["id"],
            "line_items": edited_items,
            "global_note": "Edited percent change",
        },
    )
    assert edited_generate.status_code == 200
    edited_version = edited_generate.json()["version"]
    edited_download = client.post(f"/hoa/1/versions/{edited_version['id']}/download")
    assert edited_download.status_code == 200

    assert edited_version["total_expense"] > baseline_version["total_expense"]
    assert edited_version["net_operating_income"] < baseline_version["net_operating_income"]
    assert baseline_download.content != edited_download.content


def test_non_active_draft_mutations_are_rejected(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    generate_response = _generate_version(client, draft, global_note="Initial version")
    assert generate_response.status_code == 200

    save_response = client.patch(
        "/hoa/1/budget/draft",
        json={
            "draft_id": draft["id"],
            "line_items": draft["line_items"],
            "global_note": "Should not save",
            "statement_month": draft["statement_month"],
            "growth_factor": draft["growth_factor"],
            "growth_factor_note": draft["growth_factor_note"],
        },
    )
    assert save_response.status_code == 409

    note_response = client.post(
        "/hoa/1/budget/notes",
        json={
            "draft_id": draft["id"],
            "note_scope": "global",
            "title": "Should fail",
            "body": "Old generated drafts should not accept notes.",
        },
    )
    assert note_response.status_code == 409

    regenerate_response = client.post(
        "/hoa/1/budget/generate",
        json={
            "draft_id": draft["id"],
            "line_items": draft["line_items"],
            "global_note": "Should not generate",
        },
    )
    assert regenerate_response.status_code == 409


def test_requested_draft_id_loads_exact_active_draft(client, budget_history_test_harness):
    first_upload = _upload_income_statement(client)
    assert first_upload.status_code == 200
    first_draft = first_upload.json()["draft"]

    second_upload = _upload_income_statement(client)
    assert second_upload.status_code == 200
    second_draft = second_upload.json()["draft"]

    active_response = client.get("/hoa/1/budget/draft")
    assert active_response.status_code == 200
    assert active_response.json()["id"] == second_draft["id"]

    exact_active_response = client.get(f"/hoa/1/budget/drafts/{second_draft['id']}")
    assert exact_active_response.status_code == 200
    assert exact_active_response.json()["id"] == second_draft["id"]

    superseded_response = client.get(f"/hoa/1/budget/drafts/{first_draft['id']}")
    assert superseded_response.status_code == 409


def test_download_enriched_draft_artifact_supports_active_generated_and_superseded_statuses(
    client,
    budget_history_test_harness,
):
    first_upload = _upload_income_statement(client)
    assert first_upload.status_code == 200
    first_draft = first_upload.json()["draft"]

    active_download = client.get(f"/hoa/1/budget/drafts/{first_draft['id']}/download-enriched")
    assert active_download.status_code == 200

    second_upload = _upload_income_statement(client)
    assert second_upload.status_code == 200
    second_draft = second_upload.json()["draft"]

    superseded_download = client.get(f"/hoa/1/budget/drafts/{first_draft['id']}/download-enriched")
    assert superseded_download.status_code == 200

    generate_response = _generate_version(client, second_draft, global_note="Generated status draft")
    assert generate_response.status_code == 200

    generated_download = client.get(f"/hoa/1/budget/drafts/{second_draft['id']}/download-enriched")
    assert generated_download.status_code == 200

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    timeline_types = [event["event_type"] for event in history_response.json()["timeline"]]
    assert timeline_types.count("draft_enriched_downloaded") == 3


def test_generate_pdf_origin_draft_uses_persisted_enriched_workbook(
    client,
    db_session,
    budget_history_test_harness,
    monkeypatch,
):
    from app.services import budget_history_service

    storage_root = Path(budget_history_test_harness["storage_root"])
    upload = BudgetUpload(
        property_id=1,
        original_filename="scanned-income-statement.pdf",
        storage_key="hoa/1/uploads/pdf-source/source.pdf",
        content_type="application/pdf",
        byte_size=32,
        sha256="b" * 64,
        enrichment_status="completed",
        line_items_json=json.dumps([]),
        budget_preview_json=json.dumps({}),
        statement_month=8,
        growth_factor=1.0,
        growth_factor_note="stub",
        uploaded_by_user_id=1,
        uploaded_by_name="Test User",
        created_at="2026-03-21T09:00:00+00:00",
    )
    db_session.add(upload)
    db_session.flush()

    source_pdf = storage_root / upload.storage_key
    source_pdf.parent.mkdir(parents=True, exist_ok=True)
    source_pdf.write_bytes(b"scanned pdf bytes")

    line_items = [
        {
            "line_item_key": "40000",
            "account_code": 40000,
            "category": "income",
            "label": "40000 - Assessment Income",
            "annual_budget": 120000.0,
            "percent_change": 0.0,
            "read_only": False,
        },
        {
            "line_item_key": "6100",
            "account_code": 6100,
            "category": "operating",
            "label": "6100 - Insurance",
            "annual_budget": 32000.0,
            "percent_change": 0.0,
            "read_only": False,
        },
    ]
    draft = BudgetDraft(
        property_id=1,
        source_upload_id=upload.id,
        reopened_from_version_id=None,
        status=BUDGET_DRAFT_ACTIVE,
        line_items_json=json.dumps(line_items),
        global_note=None,
        statement_month=8,
        growth_factor=1.0,
        growth_factor_note="stub",
        budget_preview_json=json.dumps({}),
        enriched_storage_key="hoa/1/drafts/pdf-draft/enriched.xlsx",
        created_by_user_id=1,
        updated_by_user_id=1,
        actor_name="Test User",
        created_at="2026-03-21T10:00:00+00:00",
        updated_at="2026-03-21T10:00:00+00:00",
    )
    db_session.add(draft)
    db_session.commit()

    enriched = storage_root / draft.enriched_storage_key
    enriched.parent.mkdir(parents=True, exist_ok=True)
    enriched.write_bytes(b"normalized workbook bytes")

    seen_input_suffixes: list[str] = []
    seen_known_columns: list[dict | None] = []

    class CapturingPipeline:
        def __init__(self, **kwargs):
            self.intermediate_path = kwargs["intermediate_path"]
            self.output_path = kwargs["output_path"]
            seen_known_columns.append(kwargs.get("known_columns"))

        def run(self):
            Path(self.intermediate_path).write_bytes(b"fake enriched workbook")
            Path(self.output_path).write_bytes(b"fake generated workbook")

    def _assert_xlsx_only(path: str) -> str:
        suffix = Path(path).suffix.lower()
        seen_input_suffixes.append(suffix)
        assert suffix == ".xlsx"
        return path

    monkeypatch.setattr(budget_history_service, "_ensure_xlsx", _assert_xlsx_only)
    monkeypatch.setattr(budget_history_service, "BudgetPipeline", CapturingPipeline)

    response = client.post(
        "/hoa/1/budget/generate",
        json={
            "draft_id": draft.id,
            "line_items": line_items,
            "global_note": "Generated from scanned PDF draft",
        },
    )

    assert response.status_code == 200, response.text
    assert seen_input_suffixes == [".xlsx"]
    assert seen_known_columns == [{"ytd_actual": 6, "annual_budget": 9}]

    version_id = response.json()["version"]["id"]
    reopen_response = client.post(f"/hoa/1/versions/{version_id}/reopen")

    assert reopen_response.status_code == 200
    assert seen_input_suffixes == [".xlsx"]
    assert seen_known_columns == [{"ytd_actual": 6, "annual_budget": 9}]
    assert reopen_response.json()["draft"]["enriched_file_available"] is True


def test_save_draft_refreshes_persisted_enriched_download(
    client,
    budget_history_test_harness,
    monkeypatch,
):
    enriched_bytes_by_output_path = {}
    changes_by_input_path = {}

    class PercentAwarePipeline:
        def __init__(self, **kwargs):
            self.input_path = kwargs["input_path"]
            self.intermediate_path = kwargs["intermediate_path"]
            self.output_path = kwargs["output_path"]

        def run(self):
            changes = changes_by_input_path.get(self.input_path, {})
            total_expense = 35200.0 + round(sum(changes.values()) * 1000, 2)
            enriched_bytes = f"enriched:{sorted(changes.items())}".encode("utf-8")
            enriched_bytes_by_output_path[self.intermediate_path] = enriched_bytes
            Path(self.intermediate_path).write_bytes(enriched_bytes)
            Path(self.output_path).write_text(str(total_expense), encoding="utf-8")

    def percent_aware_preview(path: str, max_rows: int):
        total_expense = float(Path(path).read_text(encoding="utf-8"))
        total_income = 126000.0
        return {
            "sheet": "Budget",
            "headers": ["Account Code", "Line Item", "YTD Actual", "Annual Budget",
                         "% Change", "Proposed Budget", "Monthly", "Notes"],
            "rows": [
                ["", "Total Income", None, total_income, None, total_income, round(total_income / 12, 2), None],
                ["", "Total Expense", None, total_expense, None, total_expense, round(total_expense / 12, 2), None],
                ["", "Net Operating Income", None, total_income - total_expense, None, total_income - total_expense, round((total_income - total_expense) / 12, 2), None],
            ],
        }

    def capture_percent_changes(path: str, sheet: str, changes):
        changes_by_input_path[path] = dict(changes)

    monkeypatch.setattr(
        "app.services.budget_history_service.BudgetPipeline",
        PercentAwarePipeline,
    )
    monkeypatch.setattr(
        "app.services.budget_history_service.macros_service.read_first_sheet_preview",
        percent_aware_preview,
    )
    monkeypatch.setattr(
        "app.services.budget_history_service.macros_service.write_percent_changes_by_label",
        capture_percent_changes,
    )

    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    baseline_download = client.get(f"/hoa/1/budget/drafts/{draft['id']}/download-enriched")
    assert baseline_download.status_code == 200

    edited_items = [{**item} for item in draft["line_items"]]
    edited_items[1]["percent_change"] = 250

    save_response = client.patch(
        "/hoa/1/budget/draft",
        json={
            "draft_id": draft["id"],
            "line_items": edited_items,
            "global_note": "Edited percent change",
            "statement_month": draft["statement_month"],
            "growth_factor": draft["growth_factor"],
            "growth_factor_note": draft["growth_factor_note"],
        },
    )
    assert save_response.status_code == 200
    assert save_response.json()["draft"]["enriched_file_available"] is True

    edited_download = client.get(f"/hoa/1/budget/drafts/{draft['id']}/download-enriched")
    assert edited_download.status_code == 200

    assert edited_download.content != baseline_download.content


def test_legacy_draft_download_lazily_rebuilds_missing_enriched_artifact(
    client,
    budget_history_test_harness,
):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    initial_history = client.get("/hoa/1/history")
    assert initial_history.status_code == 200
    initial_draft_summary = next(
        item for item in initial_history.json()["drafts"] if item["id"] == draft["id"]
    )
    assert initial_draft_summary["enriched_file_available"] is True

    draft_path = (
        Path(budget_history_test_harness["storage_root"])
        / "hoa"
        / "1"
        / "drafts"
        / str(draft["id"])
        / "enriched.xlsx"
    )
    assert draft_path.exists()

    db_path = Path(client.app.state.test_db_path)
    with sqlite3.connect(db_path) as connection:
      connection.execute(
          "UPDATE budget_drafts SET enriched_storage_key = NULL WHERE id = ?",
          (draft["id"],),
      )
      connection.commit()
    draft_path.unlink()

    rebuild_download = client.get(f"/hoa/1/budget/drafts/{draft['id']}/download-enriched")
    assert rebuild_download.status_code == 200
    assert draft_path.exists()

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    rebuilt_summary = next(
        item for item in history_response.json()["drafts"] if item["id"] == draft["id"]
    )
    assert rebuilt_summary["enriched_file_available"] is True


def test_compare_and_reopen_preserve_history(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    initial_draft = upload_response.json()["draft"]

    first_generate_response = _generate_version(
        client,
        initial_draft,
        global_note="Initial interim version",
    )
    assert first_generate_response.status_code == 200
    first_version = first_generate_response.json()["version"]

    second_upload_response = _upload_income_statement(client)
    assert second_upload_response.status_code == 200
    second_draft = second_upload_response.json()["draft"]

    second_generate_response = _generate_version(
        client,
        second_draft,
        global_note="Second interim version",
    )
    assert second_generate_response.status_code == 200
    second_version = second_generate_response.json()["version"]

    duplicate_compare = client.get(
        f"/hoa/1/versions/compare?left={first_version['id']}&right={first_version['id']}"
    )
    assert duplicate_compare.status_code == 400

    cross_hoa_compare = client.get(
        f"/hoa/1/versions/compare?left={first_version['id']}&right=99999"
    )
    assert cross_hoa_compare.status_code in {400, 404}

    compare_response = client.get(
        f"/hoa/1/versions/compare?left={first_version['id']}&right={second_version['id']}"
    )
    assert compare_response.status_code == 200
    compare_payload = compare_response.json()

    assert [version["id"] for version in compare_payload["versions"]] == [
        first_version["id"],
        second_version["id"],
    ]
    for version in compare_payload["versions"]:
        assert set(version) >= {
            "id",
            "version_code",
            "stage",
            "label",
            "created_at",
            "created_by_name",
            "source_upload_filename",
            "total_income",
            "total_expense",
            "net_operating_income",
            "growth_factor",
            "growth_factor_note",
            "statement_month",
            "fiscal_year_start_month",
            "fiscal_year_end_month",
        }

    reopen_response = client.post(f"/hoa/1/versions/{first_version['id']}/reopen")
    assert reopen_response.status_code == 200
    reopen_payload = reopen_response.json()

    assert reopen_payload["draft"]["status"] == "active"
    assert reopen_payload["draft"]["reopened_from_version_id"] == first_version["id"]
    assert reopen_payload["timeline_event"]["event_type"] == "version_reopened"
    assert reopen_payload["timeline_event"]["payload"]["source_version_id"] == first_version["id"]
    assert reopen_payload["timeline_event"]["payload"]["new_draft_id"] == reopen_payload["draft"]["id"]

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()

    assert history_payload["active_draft"]["id"] == reopen_payload["draft"]["id"]
    assert [version["id"] for version in history_payload["versions"]] == [
        second_version["id"],
        first_version["id"],
    ]
    reopened_event = next(
        event for event in history_payload["timeline"] if event["event_type"] == "version_reopened"
    )
    assert reopened_event["payload"]["source_version_id"] == first_version["id"]
    assert reopened_event["payload"]["new_draft_id"] == reopen_payload["draft"]["id"]


def test_version_metadata_update_and_download_are_audited(client, budget_history_test_harness):
    upload_response = _upload_income_statement(client)
    assert upload_response.status_code == 200
    draft = upload_response.json()["draft"]

    generate_response = _generate_version(
        client,
        draft,
        global_note="Ready for metadata update",
    )
    assert generate_response.status_code == 200
    version = generate_response.json()["version"]

    metadata_response = client.patch(
        f"/hoa/1/versions/{version['id']}",
        json={
            "stage": "Final",
            "label": "Board Review Draft",
            "summary_note": "Ready for board packet download.",
        },
    )
    assert metadata_response.status_code == 200
    metadata_payload = metadata_response.json()
    assert metadata_payload["version"]["stage"] == "Final"
    assert metadata_payload["version"]["label"] == "Board Review Draft"
    assert metadata_payload["version"]["summary_note"] == "Ready for board packet download."

    invalid_metadata_response = client.patch(
        f"/hoa/1/versions/{version['id']}",
        json={"version_code": "V99"},
    )
    assert invalid_metadata_response.status_code == 422

    download_response = client.post(f"/hoa/1/versions/{version['id']}/download")
    assert download_response.status_code == 200
    assert (
        download_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert download_response.content

    version_response = client.get(f"/hoa/1/versions/{version['id']}")
    assert version_response.status_code == 200
    version_payload = version_response.json()
    assert version_payload["stage"] == "Final"
    assert version_payload["label"] == "Board Review Draft"
    assert version_payload["summary_note"] == "Ready for board packet download."

    history_response = client.get("/hoa/1/history")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert {event["event_type"] for event in history_payload["timeline"]} >= {
        "version_marked_final",
        "file_downloaded",
    }
